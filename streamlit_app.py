"""Streamlit UI for the Valuation Codex biotech financial model."""

from __future__ import annotations

import json
import io
import importlib
import os
import re
from contextlib import contextmanager
from io import BytesIO
from dataclasses import asdict, fields
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import plotly.graph_objects as go
except Exception:  # pragma: no cover - optional dependency
    go = None

try:  # optional optimisation + ML helpers
    from scipy.optimize import minimize
except Exception:  # pragma: no cover - optional dependency
    minimize = None

try:
    from sklearn.cluster import KMeans
    from sklearn.linear_model import LinearRegression, LogisticRegression
    from sklearn.preprocessing import StandardScaler
except Exception:  # pragma: no cover - optional dependency
    KMeans = None
    LinearRegression = None
    LogisticRegression = None
    StandardScaler = None

from valuation_codex_package import (
    ModelConfig,
    Portfolio,
    Product,
    ProductConfig,
    STAGE_SEQUENCE,
    normalize_stage_label,
    Scenario,
    ScenarioEngine,
    ForecastEngine,
    Milestone,
    VCInputs,
    VCValuator,
    ValuationEngine,
    ValuationResult,
    MonteCarloEngine,
    validate_portfolio,
)
from valuation_codex_package.ui_state import (
    _consume_pending_panel_state,
    _consume_pending_selection,
    _panel_state_key,
    _pending_panel_state_key,
    _resolve_selected_index,
    _resolve_selected_index_from_value,
    _row_identifier,
    _selection_identity_column,
    _set_pending_panel_state,
    _set_pending_selection,
    _validate_selection,
)


STAGE_OPTIONS = list(STAGE_SEQUENCE)
STAGE_TRANSITION_COLUMNS = [
    f"{from_stage}->{to_stage}"
    for from_stage, to_stage in zip(STAGE_SEQUENCE[:-1], STAGE_SEQUENCE[1:])
]
STAGE_DURATION_COLUMNS = [f"{stage} duration (years)" for stage in STAGE_SEQUENCE]
STAGE_TRANSITION_ANNUAL_COLUMNS = [
    f"{transition} annual success %" for transition in STAGE_TRANSITION_COLUMNS
]

RAMP_SHAPE_OPTIONS = ["Linear", "S-curve", "Step"]

STAGE_COST_WEIGHT_COLUMNS = [f"{stage} R&D weight %" for stage in STAGE_SEQUENCE[:-1]]
STAGE_CAPEX_WEIGHT_COLUMNS = [f"{stage} CAPEX weight %" for stage in STAGE_SEQUENCE[:-1]]
STAGE_MILESTONE_COLUMNS = [
    f"{stage} completion milestone (USD)" for stage in STAGE_SEQUENCE[:-1]
]

SELECTOR_OPTIONS = [
    "Base case",
    "Upside",
    "Downside",
    "Aggressive expansion",
    "Defensive posture",
]

DEBT_REPAYMENT_LABELS = {
    "straight_line": "Straight-line",
    "sculpted_dscr": "Sculpted to DSCR",
    "bullet": "Bullet maturity",
    "manual": "Manual schedule",
}
DEBT_REPAYMENT_CODES = {label: code for code, label in DEBT_REPAYMENT_LABELS.items()}

DISCOUNT_TIMING_LABELS = {
    "year_end": "Year-end",
    "mid_year": "Mid-year",
    "year_0": "Year-0",
}
DISCOUNT_TIMING_CODES = {label: code for code, label in DISCOUNT_TIMING_LABELS.items()}

TERMINAL_METHOD_LABELS = {
    "exit_multiple": "Exit multiple",
    "perpetuity_growth": "Perpetuity growth",
}
TERMINAL_METHOD_CODES = {label: code for code, label in TERMINAL_METHOD_LABELS.items()}


def _inject_app_theme() -> None:
    st.markdown(
        """
        <style>
        :root {
            --bio-ink: #0f172a;
            --bio-muted: #475569;
            --bio-brand: #14532d;
            --bio-brand-soft: #e9f7ef;
            --bio-panel: rgba(255, 255, 255, 0.9);
        }
        .stApp {
            background:
                radial-gradient(circle at top left, rgba(187, 247, 208, 0.30), transparent 33%),
                radial-gradient(circle at top right, rgba(191, 219, 254, 0.22), transparent 28%),
                linear-gradient(180deg, #f6fbf7 0%, #f4f6fb 58%, #edf5f3 100%);
        }
        .block-container {
            padding-top: 1.35rem;
            padding-bottom: 3rem;
            max-width: 1450px;
        }
        .designer-hero {
            margin-bottom: 1.2rem;
            padding: 1.8rem 1.9rem;
            border-radius: 28px;
            border: 1px solid rgba(20, 83, 45, 0.12);
            background:
                linear-gradient(135deg, rgba(233, 247, 239, 0.96), rgba(255, 255, 255, 0.94)),
                linear-gradient(135deg, rgba(20, 83, 45, 0.05), rgba(30, 64, 175, 0.07));
            box-shadow: 0 24px 48px rgba(15, 23, 42, 0.08);
        }
        .designer-kicker {
            margin: 0 0 0.45rem 0;
            font-size: 0.78rem;
            letter-spacing: 0.16em;
            text-transform: uppercase;
            color: var(--bio-brand);
            font-weight: 700;
        }
        .designer-title {
            margin: 0;
            font-size: clamp(2rem, 2.8vw, 3.15rem);
            line-height: 1.02;
            color: var(--bio-ink);
            font-weight: 800;
        }
        .designer-copy {
            max-width: 55rem;
            margin: 0.7rem 0 0 0;
            color: var(--bio-muted);
            font-size: 1rem;
            line-height: 1.6;
        }
        .designer-badges {
            display: flex;
            flex-wrap: wrap;
            gap: 0.55rem;
            margin-top: 1rem;
        }
        .designer-badge {
            padding: 0.42rem 0.78rem;
            border-radius: 999px;
            border: 1px solid rgba(15, 23, 42, 0.08);
            background: rgba(255, 255, 255, 0.92);
            color: var(--bio-brand);
            font-size: 0.82rem;
            font-weight: 700;
        }
        div[data-baseweb="tab-list"] {
            gap: 0.55rem;
            margin-bottom: 1rem;
        }
        div[data-baseweb="tab-list"] button {
            min-height: 3rem;
            border-radius: 999px;
            border: 1px solid rgba(15, 23, 42, 0.08);
            background: rgba(255, 255, 255, 0.72);
            color: var(--bio-muted);
            padding: 0.25rem 1rem;
        }
        div[data-baseweb="tab-list"] button[aria-selected="true"] {
            background: linear-gradient(135deg, #14532d, #1d4ed8);
            color: white;
            border-color: transparent;
            box-shadow: 0 12px 24px rgba(29, 78, 216, 0.16);
        }
        div[data-testid="stMetric"],
        div[data-testid="stDataFrame"],
        div[data-testid="stExpander"] {
            border-radius: 20px;
        }
        div[data-testid="stMetric"] {
            border: 1px solid rgba(15, 23, 42, 0.08);
            background: var(--bio-panel);
            padding: 0.6rem 0.7rem;
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.05);
        }
        .stage-mapping-editor-controls + div[data-testid="stHorizontalBlock"] > div:nth-child(2) > div {
            padding: 0.55rem;
            border-radius: 18px;
            border: 1px solid rgba(29, 78, 216, 0.22);
            background: linear-gradient(135deg, rgba(219, 234, 254, 0.95), rgba(233, 247, 239, 0.98));
            box-shadow: 0 14px 30px rgba(29, 78, 216, 0.12);
        }
        .stage-mapping-editor-controls + div[data-testid="stHorizontalBlock"] > div:nth-child(2) button {
            width: 100%;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _render_model_hero() -> None:
    badges = "".join(
        f'<span class="designer-badge">{label}</span>'
        for label in (
            "Portfolio valuation",
            "Professional workbook",
            "Scenario stress testing",
            "RAG support",
        )
    )
    st.markdown(
        f"""
        <section class="designer-hero">
            <p class="designer-kicker">Pipeline valuation studio</p>
            <h1 class="designer-title">Biotech Financial Model</h1>
            <p class="designer-copy">
                Configure the asset mix, stage assumptions, and portfolio economics in a cleaner
                executive shell, then export an investor-grade workbook instead of flat worksheets.
            </p>
            <div class="designer-badges">{badges}</div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def _style_workbook_sheet(ws, *, accent: str, accent_soft: str, is_overview: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    if is_overview:
        ws.freeze_panes = "A6"
    elif ws.max_row > 1:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=accent)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.auto_filter.ref = ws.dimensions
        for row_idx in range(2, min(ws.max_row, 120) + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill("solid", fgColor=accent_soft)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row_idx in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 36)


def _format_overview_value(value: object) -> object:
    if isinstance(value, (int, float, np.floating)) and np.isfinite(value):
        amount = float(value)
        if abs(amount) <= 1.0 and amount != 0:
            return f"{amount:.1%}"
        if abs(amount) >= 1_000_000:
            return f"${amount / 1_000_000:,.2f}M"
        if abs(amount) >= 1_000:
            return f"${amount / 1_000:,.1f}K"
        return f"{amount:,.2f}"
    return value


def _style_professional_workbook(
    workbook_bytes: bytes,
    *,
    cons: pd.DataFrame,
    model_cfg: Optional[ModelConfig],
) -> bytes:
    workbook = load_workbook(BytesIO(workbook_bytes))
    accent = "14532D"
    accent_soft = "E9F7EF"
    if "Overview" in workbook.sheetnames:
        del workbook["Overview"]
    overview = workbook.create_sheet("Overview", 0)
    overview["A1"] = "Biotech Financial Model"
    overview["A1"].font = Font(size=20, bold=True, color="0F172A")
    overview["A2"] = "Executive overview for portfolio valuation, statements, advanced analytics, and scenario comparison."
    overview["A2"].font = Font(size=11, color="475569")
    overview["A4"] = "Executive Snapshot"
    overview["A4"].font = Font(size=12, bold=True, color=accent)
    overview["A5"] = "Metric"
    overview["B5"] = "Value"
    for cell in overview[5]:
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_rows: List[Tuple[str, object]] = []
    if cons is not None and not cons.empty:
        latest = cons.iloc[-1]
        for label, column in (
            ("Latest Revenue", "revenue"),
            ("Latest EBITDA", "ebitda"),
            ("Latest FCFF", "fcff_after_wc"),
        ):
            if column in latest.index and pd.notna(latest[column]):
                summary_rows.append((label, _format_overview_value(float(latest[column]))))
    if model_cfg is not None:
        summary_rows.append(("Tax Rate", _format_overview_value(getattr(model_cfg, "tax_rate", ""))))
        summary_rows.append(("Discount Rate", _format_overview_value(getattr(model_cfg, "discount_rate", ""))))
    summary_rows.append(("Included Sheets", len(workbook.sheetnames)))
    for row_idx, (label, value) in enumerate(summary_rows[:8], start=6):
        overview.cell(row=row_idx, column=1, value=label)
        overview.cell(row=row_idx, column=2, value=value)
    overview["D4"] = "Workbook Notes"
    overview["D4"].font = Font(size=12, bold=True, color=accent)
    notes = [
        "Portfolio valuation outputs sit alongside board-ready statements and scenario views.",
        "Advanced analytics, margin trends, and break-even analysis remain available on dedicated tabs.",
        "Use the workbook as the investor hand-off version of the live model run.",
    ]
    for row_idx, note in enumerate(notes, start=5):
        overview.cell(row=row_idx, column=4, value=f"• {note}")
    overview.column_dimensions["A"].width = 26
    overview.column_dimensions["B"].width = 18
    overview.column_dimensions["D"].width = 58

    for sheet in workbook.worksheets:
        _style_workbook_sheet(
            sheet,
            accent=accent,
            accent_soft=accent_soft,
            is_overview=sheet.title == "Overview",
        )

    output = BytesIO()
    workbook.save(output)
    return _style_professional_workbook(
        output.getvalue(),
        cons=cons,
        model_cfg=model_cfg,
    )


def _default_products() -> pd.DataFrame:
    """Seed table with two representative products."""

    data = [
        {
            "name": "AgSeed-101",
            "stage": "Approval",
            "success_prob": 1.0,
            "sales_ramp_length": 3,
            "sales_ramp_shape": "Linear",
            "include_in_consolidation": True,
            "time_to_market": 1,
            "patent_years": 18,
            "patent_revenue_target": 220_000_000,
            "post_patent_revenue_target": 180_000_000,
            "market_growth_patent": 0.03,
            "market_growth_post": 0.01,
            "cogs_patent": 0.26,
            "cogs_post": 0.32,
            "labor_pct": 0.09,
            "overhead_pct": 0.06,
            "material_pct": 0.07,
            "sales_marketing_pct": 0.12,
            "gna_pct": 0.08,
            "rd_remaining_pre_launch": 80_000_000,
            "rd_annual_post_launch": 6_000_000,
            "capex_remaining_pre_launch": 25_000_000,
            "capex_annual_post_launch": 2_500_000,
        },
        {
            "name": "BioYield-Plus",
            "stage": "Commercial",
            "success_prob": 1.0,
            "sales_ramp_length": 1,
            "sales_ramp_shape": "Step",
            "include_in_consolidation": True,
            "time_to_market": 0,
            "patent_years": 20,
            "patent_revenue_target": 300_000_000,
            "post_patent_revenue_target": 240_000_000,
            "market_growth_patent": 0.03,
            "market_growth_post": 0.01,
            "cogs_patent": 0.24,
            "cogs_post": 0.30,
            "labor_pct": 0.08,
            "overhead_pct": 0.05,
            "material_pct": 0.06,
            "sales_marketing_pct": 0.11,
            "gna_pct": 0.07,
            "rd_remaining_pre_launch": 0.0,
            "rd_annual_post_launch": 4_000_000,
            "capex_remaining_pre_launch": 0.0,
            "capex_annual_post_launch": 2_000_000,
        },
    ]
    return pd.DataFrame(data)


def _stage_visibility_flags(selected_stage: str) -> Dict[str, bool]:
    """Return visibility flags for stage-gated sections based on the pipeline stage."""

    stage_index = STAGE_SEQUENCE.index(selected_stage)
    show_precommercial = stage_index <= 4
    show_approval_or_later = stage_index >= 5
    show_forecast_ramp = stage_index in {0, 5, 6}
    return {
        "show_forecast_ramp": show_forecast_ramp,
        "show_vaccine_sales": stage_index == 6,
        "show_uses_sources": show_precommercial or stage_index == 5,
        "show_relevant_market_sizes": stage_index in {1, 2, 3, 4},
        "show_market_size_estimation": show_approval_or_later,
        "show_revenue_estimation": show_approval_or_later,
        "show_cost_assumptions": show_approval_or_later,
        "show_royalties": show_approval_or_later,
        "show_market_share": show_approval_or_later,
        "show_rd": show_precommercial,
        "show_capex": True,
    }


def _stage_mapping_sanity_checks(mapping_df: pd.DataFrame) -> List[str]:
    """Validate stage mapping inputs for scientific and commercial plausibility."""

    warnings: List[str] = []
    if mapping_df is None or mapping_df.empty:
        return warnings
    for _, row in mapping_df.iterrows():
        stage = normalize_stage_label(row.get("Stage"))
        if not stage:
            continue
        time_to_market = row.get("Time to market (years)")
        if pd.notna(time_to_market):
            time_to_market = float(time_to_market)
        else:
            time_to_market = None
        duration_sum = 0
        durations = {}
        for col in STAGE_DURATION_COLUMNS:
            value = row.get(col)
            if pd.isna(value):
                continue
            duration = max(0, int(value))
            stage_name = col.replace(" duration (years)", "")
            durations[stage_name] = duration
            duration_sum += duration
        if time_to_market is not None and duration_sum and abs(time_to_market - duration_sum) > 1:
            warnings.append(
                f"{stage}: time-to-market ({time_to_market:.0f}y) should align with "
                f"the stage durations total ({duration_sum}y)."
            )
        for col, label in [
            ("Success Probability %", "success probability"),
            ("R&D remaining pre-launch (USD)", "R&D remaining"),
            ("R&D annual post-launch (USD/year)", "R&D annual post-launch"),
        ]:
            value = row.get(col)
            if pd.isna(value):
                continue
            if col == "Success Probability %" and not (0 <= float(value) <= 100):
                warnings.append(f"{stage}: {label} should be between 0% and 100%.")
        for weight_cols, label in [
            (STAGE_COST_WEIGHT_COLUMNS, "R&D weight"),
            (STAGE_CAPEX_WEIGHT_COLUMNS, "CAPEX weight"),
        ]:
            weight_total = 0.0
            for col in weight_cols:
                value = row.get(col)
                if pd.isna(value):
                    continue
                weight_total += float(value)
            if weight_total and abs(weight_total - 100.0) > 5.0:
                warnings.append(
                    f"{stage}: {label} totals {weight_total:.0f}%. Target ~100% so spend allocation is coherent."
                )
        if stage in {"Approval", "Commercial"}:
            if durations.get("Discovery", 0) or durations.get("Preclinical", 0):
                warnings.append(
                    f"{stage}: early-stage durations should typically be 0 once in {stage}."
                )
            if stage == "Commercial" and time_to_market not in (0, None):
                warnings.append(
                    f"{stage}: time-to-market should be 0 for commercial assets."
                )
    return warnings


def _render_section_warnings(title: str, warnings: List[str]) -> None:
    if not warnings:
        return
    st.warning(f"{title}: please review the inputs below for scientific, commercial, or financial validity.")
    for warning in warnings:
        st.write(f"- {warning}")


@contextmanager
def _section_block(title: str, *, heading_level: int = 3, caption: Optional[str] = None):
    heading_level = min(max(int(heading_level), 1), 6)
    st.markdown(f"{'#' * heading_level} {title}")
    if caption:
        st.caption(caption)
    with st.container():
        yield


def _template_library() -> Dict[str, pd.DataFrame]:
    """Pre-built product templates for quick setup."""

    templates = {
        "Phase II oncology asset": pd.DataFrame(
            [
                {
                    "name": "Onco-Phase2",
                    "stage": "Phase II",
                    "success_prob": 0.35,
                    "include_in_consolidation": True,
                    "time_to_market": 4,
                    "patent_years": 12,
                    "patent_revenue_target": 250_000_000,
                    "post_patent_revenue_target": 120_000_000,
                    "market_growth_patent": 0.03,
                    "market_growth_post": -0.02,
                    "cogs_patent": 0.28,
                    "cogs_post": 0.5,
                    "labor_pct": 0.12,
                    "overhead_pct": 0.08,
                    "material_pct": 0.1,
                    "sales_marketing_pct": 0.18,
                    "gna_pct": 0.12,
                    "rd_remaining_pre_launch": 220_000_000,
                    "rd_annual_post_launch": 15_000_000,
                    "capex_remaining_pre_launch": 70_000_000,
                    "capex_annual_post_launch": 7_500_000,
                }
            ]
        ),
        "Pre-clinical platform": pd.DataFrame(
            [
                {
                    "name": "Platform-Preclinical",
                    "stage": "Preclinical",
                    "success_prob": 0.2,
                    "include_in_consolidation": True,
                    "time_to_market": 6,
                    "patent_years": 15,
                    "patent_revenue_target": 150_000_000,
                    "post_patent_revenue_target": 80_000_000,
                    "market_growth_patent": 0.04,
                    "market_growth_post": 0.0,
                    "cogs_patent": 0.3,
                    "cogs_post": 0.55,
                    "labor_pct": 0.14,
                    "overhead_pct": 0.1,
                    "material_pct": 0.12,
                    "sales_marketing_pct": 0.16,
                    "gna_pct": 0.12,
                    "rd_remaining_pre_launch": 280_000_000,
                    "rd_annual_post_launch": 10_000_000,
                    "capex_remaining_pre_launch": 40_000_000,
                    "capex_annual_post_launch": 5_000_000,
                }
            ]
        ),
        "Commercial asset": pd.DataFrame(
            [
                {
                    "name": "Commercial-Asset",
                    "stage": "Commercial",
                    "success_prob": 1.0,
                    "include_in_consolidation": True,
                    "time_to_market": 0,
                    "patent_years": 8,
                    "patent_revenue_target": 300_000_000,
                    "post_patent_revenue_target": 140_000_000,
                    "market_growth_patent": 0.02,
                    "market_growth_post": -0.03,
                    "cogs_patent": 0.25,
                    "cogs_post": 0.45,
                    "labor_pct": 0.1,
                    "overhead_pct": 0.08,
                    "material_pct": 0.08,
                    "sales_marketing_pct": 0.14,
                    "gna_pct": 0.1,
                    "rd_remaining_pre_launch": 0.0,
                    "rd_annual_post_launch": 8_000_000,
                    "capex_remaining_pre_launch": 0.0,
                    "capex_annual_post_launch": 6_000_000,
                }
            ]
        ),
    }
    return templates


def _blank_product_row(name: str = "New vaccine") -> Dict:
    """Return a ProductConfig-like dict for initializing new rows."""

    cfg = ProductConfig(
        name=name,
        stage="Discovery",
        success_prob=0.2,
        include_in_consolidation=True,
        patent_revenue_target=50_000_000,
        post_patent_revenue_target=25_000_000,
        cogs_patent=0.35,
        cogs_post=0.5,
        labor_pct=0.12,
        overhead_pct=0.08,
        material_pct=0.1,
        sales_marketing_pct=0.15,
        gna_pct=0.1,
        rd_remaining_pre_launch=25_000_000,
        rd_annual_post_launch=5_000_000,
        capex_remaining_pre_launch=10_000_000,
        capex_annual_post_launch=2_000_000,
    )
    return asdict(cfg)


def _default_vaccine_sales_table(first_year: int = 2024, horizon_years: int = 5) -> pd.DataFrame:
    years = [first_year + i for i in range(max(horizon_years, 1))]
    vaccine_rows = _default_vaccine_revenue_table()[
        ["ID_vaccine", "Vaccine name", "Patent customers per year", "Patent price (USD/customer)"]
    ]
    rows: List[Dict[str, Any]] = []
    for _, vaccine in vaccine_rows.iterrows():
        doses = float(vaccine.get("Patent customers per year", 0.0) or 0.0) / 1e6
        price = float(vaccine.get("Patent price (USD/customer)", 0.0) or 0.0)
        for year in years:
            rows.append(
                {
                    "ID_vaccine": vaccine["ID_vaccine"],
                    "Vaccine name": vaccine["Vaccine name"],
                    "Year": year,
                    "Doses (M)": doses,
                    "Price per dose": price,
                    "Comments": "",
                }
            )
    return pd.DataFrame(rows)


def _blank_vaccine_sales_row(df: pd.DataFrame, first_year: int) -> Dict:
    next_year = first_year
    if "Year" in df.columns and not df.empty:
        with pd.option_context("mode.use_inf_as_na", True):
            existing_years = pd.to_numeric(df["Year"], errors="coerce").dropna()
        if not existing_years.empty:
            next_year = int(existing_years.max()) + 1
    doses = 5.0
    price = 25.0
    if "Doses (M)" in df.columns and not df.empty:
        last_doses = pd.to_numeric(df["Doses (M)"], errors="coerce").dropna()
        if not last_doses.empty:
            doses = float(last_doses.iloc[-1])
    if "Price per dose" in df.columns and not df.empty:
        last_price = pd.to_numeric(df["Price per dose"], errors="coerce").dropna()
        if not last_price.empty:
            price = float(last_price.iloc[-1])
    vaccine_id = _next_vaccine_id(df)
    vaccine_name = "New vaccine"
    if "ID_vaccine" in df.columns and not df.empty:
        last_id = df["ID_vaccine"].dropna()
        if not last_id.empty:
            vaccine_id = str(last_id.iloc[-1])
    if "Vaccine name" in df.columns and not df.empty:
        last_name = df["Vaccine name"].dropna()
        if not last_name.empty:
            vaccine_name = str(last_name.iloc[-1])
    return {
        "ID_vaccine": vaccine_id,
        "Vaccine name": vaccine_name,
        "Year": next_year,
        "Doses (M)": doses,
        "Price per dose": price,
        "Comments": "",
    }


def _default_uses_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Item": "Final approval, launch readiness, and market access",
            "Amount": 110_000_000,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Item": "Commercial capacity, channels, and support programs",
            "Amount": 90_000_000,
        },
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Item": "Working capital buffer and contingency",
            "Amount": 50_000_000,
        },
    ]
    return pd.DataFrame(data)


def _blank_use_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    vaccine_name = "New vaccine"
    if "ID_vaccine" in df.columns and not df.empty:
        last_id = df["ID_vaccine"].dropna()
        if not last_id.empty:
            next_id = str(last_id.iloc[-1])
    if "Vaccine name" in df.columns and not df.empty:
        last_name = df["Vaccine name"].dropna()
        if not last_name.empty:
            vaccine_name = str(last_name.iloc[-1])
    return {
        "ID_vaccine": next_id,
        "Vaccine name": vaccine_name,
        "Item": "New use",
        "Amount": 0.0,
    }


def _default_sources_table() -> pd.DataFrame:
    data = [
        {"Item": "Existing cash", "Amount": 35_000_000},
        {"Item": "Strategic grant", "Amount": 10_000_000},
        {"Item": "New equity", "Amount": 85_000_000},
    ]
    return pd.DataFrame(data)


def _blank_source_row(df: pd.DataFrame) -> Dict:
    return {"Item": "New source", "Amount": 0.0}


def _default_shareholders_table() -> pd.DataFrame:
    data = [
        {
            "Shareholder": "Founders",
            "Security": "Common",
            "Seniority": 3,
            "Ownership %": 0.35,
            "Investment": 30_000_000,
            "Liquidation preference (x)": 0.0,
            "Participating preferred": False,
        },
        {
            "Shareholder": "Growth fund",
            "Security": "Preferred",
            "Seniority": 1,
            "Ownership %": 0.65,
            "Investment": 55_000_000,
            "Liquidation preference (x)": 1.0,
            "Participating preferred": False,
        },
    ]
    return pd.DataFrame(data)


def _blank_shareholder_row(df: pd.DataFrame) -> Dict:
    return {
        "Shareholder": "New investor",
        "Security": "Preferred",
        "Seniority": 1,
        "Ownership %": 0.05,
        "Investment": 0.0,
        "Liquidation preference (x)": 1.0,
        "Participating preferred": False,
    }


def _default_market_sizes_table() -> pd.DataFrame:
    data = [
        {"Segment": "Crop protection biologics", "Value": 2_750_000_000},
        {"Segment": "Soil and yield enhancement platforms", "Value": 3_750_000_000},
    ]
    return pd.DataFrame(data)


def _blank_relevant_market_row(df: pd.DataFrame) -> Dict:
    return {"Segment": "New segment", "Value": 1_000_000}


def _default_vaccine_development_table(first_year: int = 2024) -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Stage": "Approval",
            "Success Probability %": 100.0,
            "Consolidation": True,
            "First year forecast": first_year,
            "Time to market": 1,
            "Market entry year": first_year + 1,
            "Patent duration years": 18,
            "End patent year": first_year + 18,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Stage": "Commercial",
            "Success Probability %": 100.0,
            "Consolidation": True,
            "First year forecast": first_year,
            "Time to market": 0,
            "Market entry year": first_year,
            "Patent duration years": 20,
            "End patent year": first_year + 19,
        },
    ]
    return pd.DataFrame(data)


def _default_market_size_estimation_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Market size (# customers)": 7_000_000,
            "Average spend (USD/customer)": 65,
            "Serviceable Available Market (% TAM)": 80.0,
            "Serviceable Available Market (% Market size)": 70.0,
            "Serviceable Obtainable Market (%)": 55.0,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Market size (# customers)": 6_500_000,
            "Average spend (USD/customer)": 75,
            "Serviceable Available Market (% TAM)": 80.0,
            "Serviceable Available Market (% Market size)": 75.0,
            "Serviceable Obtainable Market (%)": 70.0,
        },
    ]
    return pd.DataFrame(data)


def _default_vaccine_revenue_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Patent customers per year": 4_000_000,
            "Patent price (USD/customer)": 55,
            "Post patent customer adj. %": 90.0,
            "Post patent price adj. %": 91.0,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Patent customers per year": 5_000_000,
            "Patent price (USD/customer)": 60,
            "Post patent customer adj. %": 100.0,
            "Post patent price adj. %": 80.0,
        },
    ]
    return pd.DataFrame(data)


def _default_royalty_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Monetization model": "Product Sale",
            "Royalty rate (%)": 0.0,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Monetization model": "Product Sale",
            "Royalty rate (%)": 0.0,
        },
    ]
    return pd.DataFrame(data)


def _default_market_share_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Relevant market type": "Crop protection biologics",
            "Relevant market size (USD)": 2_750_000_000,
            "Revenue target - patent %": 8.0,
            "Revenue target - post %": 6.55,
            "Market share patent %": 5.5,
            "Market share post %": 4.5,
            "Market growth %": 1.5,
            "Sales growth %": 3.0,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Relevant market type": "Soil and yield enhancement platforms",
            "Relevant market size (USD)": 3_750_000_000,
            "Revenue target - patent %": 8.0,
            "Revenue target - post %": 6.4,
            "Market share patent %": 6.0,
            "Market share post %": 4.8,
            "Market growth %": 1.5,
            "Sales growth %": 3.0,
        },
    ]
    return pd.DataFrame(data)


def _default_vaccine_cost_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "COGS patent % of sales": 26.0,
            "COGS post % of sales": 32.0,
            "Marketing annual % of sales": 12.0,
            "Marketing launch cost (USD)": 10_000_000,
            "Indirect staff cost (USD)": 10_000_000,
            "Electricity (USD)": 1_600_000,
            "Depreciation (USD)": 4_000_000,
            "Interest & amortization (USD)": 2_000_000,
            "Royalties cost % of sales": 0.0,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "COGS patent % of sales": 24.0,
            "COGS post % of sales": 30.0,
            "Marketing annual % of sales": 11.0,
            "Marketing launch cost (USD)": 8_000_000,
            "Indirect staff cost (USD)": 12_000_000,
            "Electricity (USD)": 1_500_000,
            "Depreciation (USD)": 4_500_000,
            "Interest & amortization (USD)": 3_000_000,
            "Royalties cost % of sales": 0.0,
        },
    ]
    return pd.DataFrame(data)


def _default_vaccine_rd_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Cost accounting (capitalisation)": "55% capitalised",
            "Pre-GTM spent to date (USD)": 60_000_000,
            "Pre-GTM remaining (USD)": 80_000_000,
            "Post-GTM annual cost (USD/year)": 6_000_000,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Cost accounting (capitalisation)": "45% capitalised",
            "Pre-GTM spent to date (USD)": 150_000_000,
            "Pre-GTM remaining (USD)": 0.0,
            "Post-GTM annual cost (USD/year)": 4_000_000,
        },
    ]
    return pd.DataFrame(data)


def _default_vaccine_capex_table() -> pd.DataFrame:
    data = [
        {
            "ID_vaccine": "VAC-001",
            "Vaccine name": "AgSeed-101",
            "Manufacturing & Scale-up Assets (Pre-GTM, USD)": 8_000_000,
            "Manufacturing & Scale-up Assets (Post-GTM, USD/year)": 700_000,
            "Quality & Compliance Infrastructure (Pre-GTM, USD)": 4_000_000,
            "Quality & Compliance Infrastructure (Post-GTM, USD/year)": 350_000,
            "Cold-chain / Distribution Assets (Pre-GTM, USD)": 2_000_000,
            "Cold-chain / Distribution Assets (Post-GTM, USD/year)": 200_000,
            "IT / Data / Digital Infrastructure (Pre-GTM, USD)": 1_500_000,
            "IT / Data / Digital Infrastructure (Post-GTM, USD/year)": 150_000,
            "Facility Build-out / Leasehold Improvements (Pre-GTM, USD)": 5_500_000,
            "Facility Build-out / Leasehold Improvements (Post-GTM, USD/year)": 600_000,
            "Process Development & Tech-Transfer Assets (Pre-GTM, USD)": 4_000_000,
            "Process Development & Tech-Transfer Assets (Post-GTM, USD/year)": 500_000,
        },
        {
            "ID_vaccine": "VAC-002",
            "Vaccine name": "BioYield-Plus",
            "Manufacturing & Scale-up Assets (Pre-GTM, USD)": 0.0,
            "Manufacturing & Scale-up Assets (Post-GTM, USD/year)": 600_000,
            "Quality & Compliance Infrastructure (Pre-GTM, USD)": 0.0,
            "Quality & Compliance Infrastructure (Post-GTM, USD/year)": 250_000,
            "Cold-chain / Distribution Assets (Pre-GTM, USD)": 0.0,
            "Cold-chain / Distribution Assets (Post-GTM, USD/year)": 150_000,
            "IT / Data / Digital Infrastructure (Pre-GTM, USD)": 0.0,
            "IT / Data / Digital Infrastructure (Post-GTM, USD/year)": 150_000,
            "Facility Build-out / Leasehold Improvements (Pre-GTM, USD)": 0.0,
            "Facility Build-out / Leasehold Improvements (Post-GTM, USD/year)": 500_000,
            "Process Development & Tech-Transfer Assets (Pre-GTM, USD)": 0.0,
            "Process Development & Tech-Transfer Assets (Post-GTM, USD/year)": 350_000,
        },
    ]
    return pd.DataFrame(data)


def _default_shared_capex_pools_table() -> pd.DataFrame:
    data = [
        {
            "Pool name": "Core manufacturing facility",
            "Applies to (IDs or ALL)": "ALL",
            "Allocation method": "Equal",
            "Manufacturing & Scale-up Assets (Pre-GTM, USD)": 20_000_000,
            "Manufacturing & Scale-up Assets (Post-GTM, USD/year)": 2_500_000,
            "Quality & Compliance Infrastructure (Pre-GTM, USD)": 5_000_000,
            "Quality & Compliance Infrastructure (Post-GTM, USD/year)": 600_000,
            "Cold-chain / Distribution Assets (Pre-GTM, USD)": 3_000_000,
            "Cold-chain / Distribution Assets (Post-GTM, USD/year)": 400_000,
            "IT / Data / Digital Infrastructure (Pre-GTM, USD)": 2_000_000,
            "IT / Data / Digital Infrastructure (Post-GTM, USD/year)": 250_000,
            "Facility Build-out / Leasehold Improvements (Pre-GTM, USD)": 8_000_000,
            "Facility Build-out / Leasehold Improvements (Post-GTM, USD/year)": 850_000,
            "Process Development & Tech-Transfer Assets (Pre-GTM, USD)": 4_000_000,
            "Process Development & Tech-Transfer Assets (Post-GTM, USD/year)": 350_000,
        }
    ]
    return pd.DataFrame(data)


def _default_shared_capex_allocations_table() -> pd.DataFrame:
    data = [
        {"Pool name": "Core manufacturing facility", "ID_vaccine": "VAC-001", "Weight": 0.5},
        {"Pool name": "Core manufacturing facility", "ID_vaccine": "VAC-002", "Weight": 0.5},
    ]
    return pd.DataFrame(data)


def _next_vaccine_id(df: pd.DataFrame) -> str:
    """Return the next sequential vaccine identifier (VAC-XXX)."""

    existing = set()
    if "ID_vaccine" in df.columns:
        existing = {
            str(val)
            for val in df["ID_vaccine"].astype(str).tolist()
            if val and val != "nan"
        }
    idx = 1
    while True:
        candidate = f"VAC-{idx:03d}"
        if candidate not in existing:
            return candidate
        idx += 1


def _blank_vaccine_development_row(df: pd.DataFrame, first_year: int) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "Stage": "Discovery",
        "Success Probability %": 30.0,
        "Consolidation": True,
        "First year forecast": first_year,
        "Time to market": 3,
        "Market entry year": first_year + 3,
        "Patent duration years": 15,
        "End patent year": first_year + 17,
    }


def _blank_market_size_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "Market size (# customers)": 1_000_000,
        "Average spend (USD/customer)": 100.0,
        "Serviceable Available Market (% TAM)": 50.0,
        "Serviceable Available Market (% Market size)": 40.0,
        "Serviceable Obtainable Market (%)": 20.0,
    }


def _blank_vaccine_revenue_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "Patent customers per year": 1_000_000,
        "Patent price (USD/customer)": 50.0,
        "Post patent customer adj. %": 80.0,
        "Post patent price adj. %": 85.0,
    }


def _blank_vaccine_cost_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "COGS patent % of sales": 30.0,
        "COGS post % of sales": 45.0,
        "Marketing annual % of sales": 15.0,
        "Marketing launch cost (USD)": 10_000_000,
        "Indirect staff cost (USD)": 5_000_000,
        "Electricity (USD)": 1_000_000,
        "Depreciation (USD)": 2_000_000,
        "Interest & amortization (USD)": 1_000_000,
        "Royalties cost % of sales": 3.0,
    }


def _blank_vaccine_rd_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "Cost accounting (capitalisation)": "50% capitalised",
        "Pre-GTM spent to date (USD)": 20_000_000,
        "Pre-GTM remaining (USD)": 10_000_000,
        "Post-GTM annual cost (USD/year)": 5_000_000,
    }


def _blank_vaccine_capex_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "Manufacturing & Scale-up Assets (Pre-GTM, USD)": 8_000_000,
        "Manufacturing & Scale-up Assets (Post-GTM, USD/year)": 1_200_000,
        "Quality & Compliance Infrastructure (Pre-GTM, USD)": 3_000_000,
        "Quality & Compliance Infrastructure (Post-GTM, USD/year)": 300_000,
        "Cold-chain / Distribution Assets (Pre-GTM, USD)": 2_000_000,
        "Cold-chain / Distribution Assets (Post-GTM, USD/year)": 250_000,
        "IT / Data / Digital Infrastructure (Pre-GTM, USD)": 1_500_000,
        "IT / Data / Digital Infrastructure (Post-GTM, USD/year)": 200_000,
        "Facility Build-out / Leasehold Improvements (Pre-GTM, USD)": 4_000_000,
        "Facility Build-out / Leasehold Improvements (Post-GTM, USD/year)": 350_000,
        "Process Development & Tech-Transfer Assets (Pre-GTM, USD)": 2_500_000,
        "Process Development & Tech-Transfer Assets (Post-GTM, USD/year)": 250_000,
    }


def _blank_vaccine_royalty_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "Monetization model": "Licensing",
        "Royalty rate (%)": 5.0,
    }


def _blank_vaccine_market_share_row(df: pd.DataFrame) -> Dict:
    next_id = _next_vaccine_id(df)
    return {
        "ID_vaccine": next_id,
        "Vaccine name": "New vaccine",
        "Relevant market type": "New segment",
        "Relevant market size (USD)": 1_000_000_000,
        "Revenue target - patent %": 10.0,
        "Revenue target - post %": 5.0,
        "Market share patent %": 5.0,
        "Market share post %": 3.0,
        "Market growth %": 5.0,
        "Sales growth %": 8.0,
    }


def _ensure_table_state(key: str, default_factory: Callable[[], pd.DataFrame]) -> pd.DataFrame:
    if key not in st.session_state or st.session_state[key] is None:
        st.session_state[key] = default_factory()
    return st.session_state[key]


def _parse_pool_targets(raw_value: str, fallback_ids: List[str]) -> List[str]:
    if not raw_value:
        return fallback_ids
    cleaned = str(raw_value).strip()
    if not cleaned:
        return fallback_ids
    if cleaned.upper() == "ALL":
        return fallback_ids
    targets = [item.strip() for item in cleaned.split(",") if item.strip()]
    return targets or fallback_ids


def _build_shared_capex_allocations(
    dev_df: pd.DataFrame,
    pools_df: pd.DataFrame,
    allocations_df: pd.DataFrame,
) -> pd.DataFrame:
    if dev_df.empty or "ID_vaccine" not in dev_df.columns:
        return pd.DataFrame()
    vaccine_ids = (
        dev_df["ID_vaccine"].astype(str).dropna().tolist()
        if "ID_vaccine" in dev_df.columns
        else []
    )
    if pools_df.empty:
        return pd.DataFrame()

    rows: List[Dict[str, Any]] = []
    alloc_df = allocations_df.copy()
    for _, pool in pools_df.iterrows():
        pool_name = str(pool.get("Pool name", "")).strip() or "Shared pool"
        method = str(pool.get("Allocation method", "Equal")).strip() or "Equal"
        targets = _parse_pool_targets(pool.get("Applies to (IDs or ALL)", ""), vaccine_ids)
        targets = [t for t in targets if t in vaccine_ids]
        if not targets:
            continue

        if method.lower().startswith("by weight"):
            weights_df = alloc_df.loc[
                alloc_df.get("Pool name", "") == pool_name, ["ID_vaccine", "Weight"]
            ].copy()
            weights_df["ID_vaccine"] = weights_df["ID_vaccine"].astype(str)
            weights_df = weights_df[weights_df["ID_vaccine"].isin(targets)]
            weights = _coerce_numeric(weights_df.get("Weight", pd.Series(dtype=float)), 0.0)
            weight_map = dict(zip(weights_df["ID_vaccine"], weights))
            total_weight = sum(weight_map.values())
            if total_weight <= 0:
                weight_map = {vid: 1.0 for vid in targets}
                total_weight = float(len(targets))
        else:
            weight_map = {vid: 1.0 for vid in targets}
            total_weight = float(len(targets))

        for vid in targets:
            weight = weight_map.get(vid, 0.0)
            if total_weight <= 0:
                share = 0.0
            else:
                share = weight / total_weight
            rows.append(
                {
                    "ID_vaccine": vid,
                    "Pool name": pool_name,
                    "Share": share,
                }
            )

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def _format_row_label(
    df: pd.DataFrame,
    idx,
    id_column: Optional[str],
    name_column: Optional[str],
) -> str:
    parts: List[str] = []
    if id_column and id_column in df.columns:
        val = df.at[idx, id_column]
        if pd.notna(val):
            parts.append(str(val))
    if name_column and name_column in df.columns:
        val = df.at[idx, name_column]
        if pd.notna(val):
            parts.append(str(val))
    if not parts:
        pos = df.index.get_loc(idx) if idx in df.index else 0
        parts.append(f"Row {pos + 1}")
    return " - ".join(parts)


def _render_row_selector(
    df: pd.DataFrame,
    select_key: str,
    id_column: Optional[str],
    name_column: Optional[str],
) -> Optional[int]:
    pending = _consume_pending_selection(select_key)

    if df.empty:
        st.caption("No rows available yet.")
        st.session_state.pop(select_key, None)
        st.session_state.pop(_pending_selection_key(select_key), None)
        return None

    row_indexes = list(df.index)
    identity_column = _selection_identity_column(df, id_column, name_column)
    option_values = [_row_identifier(df, idx, identity_column) for idx in row_indexes]
    option_to_index = dict(zip(option_values, row_indexes))
    selected_id = pending if pending is not None else st.session_state.get(select_key)
    default_idx = _resolve_selected_index_from_value(df, selected_id, identity_column)
    if default_idx is None or default_idx not in row_indexes:
        default_idx = row_indexes[0]
    default_value = _row_identifier(df, default_idx, identity_column)
    if st.session_state.get(select_key) != default_value:
        st.session_state[select_key] = default_value

    def _format(option_value):
        idx = option_to_index.get(option_value, default_idx)
        return _format_row_label(df, idx, id_column, name_column)

    selected_value = st.selectbox(
        "Select row",
        options=option_values,
        format_func=_format,
        index=option_values.index(default_value),
        key=select_key,
    )
    return option_to_index.get(selected_value, default_idx)


def _render_yearly_increment_helper(
    *,
    section_key: str,
    df: pd.DataFrame,
    year_column: str,
    target_columns: List[str],
    filter_builder: Callable[[pd.DataFrame, Optional[str], int], pd.Series],
    id_column: Optional[str] = None,
    id_label: str = "ID",
    start_year_label: str = "Start year",
    start_year_default: int = 0,
    periods_default: int = 5,
    increment_default: float = 1.0,
    allow_compound: bool = True,
    create_missing_rows: bool = False,
    base_value_mode: str = "first_row",
    start_value_default: float = 0.0,
) -> pd.DataFrame:
    st.markdown("**Yearly Increment Helper**")
    st.caption(
        "Apply a fixed change or % growth from a start year onward. "
        "'Increment per year' is the step size (or growth rate when compounding). "
        "'Years to apply' controls how many consecutive rows are updated."
    )
    if df.empty:
        st.caption("Add rows to apply increments.")
        return df
    if year_column not in df.columns:
        st.caption(f"Missing '{year_column}' column.")
        return df

    selected_id: Optional[str] = None
    if id_column and id_column in df.columns:
        id_values = df[id_column].dropna().astype(str).unique().tolist()
        if not id_values:
            st.caption(f"Add {id_label} values to use the helper.")
            return df
        selected_id = st.selectbox(
            id_label,
            options=id_values,
            key=f"{section_key}_inc_id",
        )

    available_cols = [col for col in target_columns if col in df.columns]
    if not available_cols:
        st.caption("No target columns available.")
        return df

    target_col = st.selectbox(
        "Column",
        options=available_cols,
        key=f"{section_key}_inc_col",
    )
    start_year = st.number_input(
        start_year_label,
        value=int(start_year_default),
        step=1,
        key=f"{section_key}_inc_start",
    )
    years = st.number_input(
        "Years to apply",
        min_value=1,
        max_value=50,
        value=int(periods_default),
        key=f"{section_key}_inc_years",
    )
    increment = st.number_input(
        "Increment per year",
        value=float(increment_default),
        step=0.1,
        key=f"{section_key}_inc_value",
    )
    compound = False
    if allow_compound:
        compound = st.checkbox(
            "Compound annually (apply % growth)",
            value=False,
            key=f"{section_key}_inc_compound",
        )

    base_value = None
    if base_value_mode == "input":
        base_value = st.number_input(
            "Starting value",
            value=float(start_value_default),
            step=0.1,
            key=f"{section_key}_inc_start_value",
        )

    if st.button("Apply increment", key=f"{section_key}_inc_apply", use_container_width=True):
        df = df.copy()
        mask = filter_builder(df, selected_id, int(start_year))
        subset = df.loc[mask, [year_column, target_col]].copy()
        subset[year_column] = pd.to_numeric(subset[year_column], errors="coerce")
        subset[target_col] = pd.to_numeric(subset[target_col], errors="coerce").fillna(0.0)
        subset = subset.dropna(subset=[year_column]).sort_values(year_column)

        if subset.empty and not create_missing_rows:
            st.warning("No matching rows found for the selected filters.")
            return df

        if base_value_mode != "input":
            base_value = float(subset[target_col].iloc[0]) if not subset.empty else 0.0

        if create_missing_rows:
            existing_years = (
                pd.to_numeric(df[year_column], errors="coerce")
                .fillna(-1)
                .astype(int)
                .tolist()
            )
            for offset in range(int(years)):
                year_value = int(start_year + offset)
                if year_value not in existing_years:
                    new_row = {col: np.nan for col in df.columns}
                    new_row[year_column] = year_value
                    if id_column and selected_id is not None:
                        new_row[id_column] = selected_id
                    df.loc[len(df)] = new_row
            mask = filter_builder(df, selected_id, int(start_year))
            subset = df.loc[mask, [year_column, target_col]].copy()
            subset[year_column] = pd.to_numeric(subset[year_column], errors="coerce")
            subset = subset.dropna(subset=[year_column]).sort_values(year_column)

        for i, year_value in enumerate(subset[year_column].iloc[: int(years)]):
            if compound:
                value = float(base_value) * ((1 + increment) ** i)
            else:
                value = float(base_value) + increment * i
            row_mask = df[year_column].astype(int) == int(year_value)
            if id_column and selected_id is not None:
                row_mask &= df[id_column].astype(str) == str(selected_id)
            df.loc[row_mask, target_col] = value

        st.session_state[section_key] = df
        st.success("Increment applied")

    return st.session_state.get(section_key, df)


def _apply_yearly_increment(
    section_key: str,
    df: pd.DataFrame,
    selected_idx: Optional[int],
) -> pd.DataFrame:
    if df.empty or selected_idx is None or selected_idx not in df.index:
        st.caption("Select a row to apply increments.")
        return df

    temp_col = "__row_index__"
    df = df.copy()
    df[temp_col] = np.arange(len(df))
    start_pos = int(df.at[selected_idx, temp_col])

    def _filter(df: pd.DataFrame, _selected_id: Optional[str], start_year: int) -> pd.Series:
        return df[temp_col] >= int(start_year)

    numeric_cols = [
        col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])
    ]
    if not numeric_cols:
        st.caption("No numeric columns available.")
        return df

    updated = _render_yearly_increment_helper(
        section_key=section_key,
        df=df,
        year_column=temp_col,
        target_columns=numeric_cols,
        filter_builder=_filter,
        start_year_label="Start row",
        start_year_default=start_pos,
        periods_default=1,
        increment_default=1.0,
        allow_compound=True,
        create_missing_rows=False,
        base_value_mode="first_row",
    )
    updated = updated.drop(columns=[temp_col], errors="ignore")
    st.session_state[section_key] = updated
    return updated


def _widget_value(label: str, value, key: str):
    """Render an input widget based on the inferred data type of ``value``."""

    label_lower = label.lower()
    if label == "Stage":
        current = value if value in STAGE_OPTIONS else STAGE_OPTIONS[0]
        return st.selectbox(label, options=STAGE_OPTIONS, index=STAGE_OPTIONS.index(current), key=key)

    bool_like = isinstance(value, (bool, np.bool_)) or label_lower in {
        "include_in_consolidation",
        "consolidation",
        "participating preferred",
    }
    if bool_like:
        return st.checkbox(label, value=bool(value), key=key)

    # Treat missing numeric values as zero for editing convenience.
    numeric_like = isinstance(value, (int, float, np.number)) or (
        isinstance(value, str) and value.strip().replace(".", "", 1).isdigit()
    )
    if numeric_like:
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            numeric_value = 0.0
        min_value: Optional[float] = None
        max_value: Optional[float] = None
        step = 0.1 if abs(numeric_value) < 1 else 1.0
        if "%" in label or "prob" in label_lower or "growth" in label_lower or "share" in label_lower:
            min_value = 0.0
        if "%" in label or "prob" in label_lower or "probability" in label_lower:
            max_value = 100.0
        kwargs = {"value": float(numeric_value), "step": step, "key": key}
        if min_value is not None:
            kwargs["min_value"] = float(min_value)
        if max_value is not None:
            kwargs["max_value"] = float(max_value)
        return st.number_input(label, **kwargs)

    safe_value = "" if value is None or (isinstance(value, float) and pd.isna(value)) else str(value)
    return st.text_input(label, value=safe_value, key=key)


def _render_row_form(
    *,
    section_key: str,
    form_key: str,
    title: str,
    columns: List[str],
    initial_values: Dict,
    submit_label: str,
) -> Optional[Dict]:
    """Generic helper that renders a form for editing/adding a row."""

    with st.form(f"{section_key}_{form_key}"):
        st.caption(title)
        new_values: Dict = {}
        for col in columns:
            val = initial_values.get(col, "")
            widget_key = f"{section_key}_{form_key}_{col}"
            new_values[col] = _widget_value(col, val, widget_key)
        submitted = st.form_submit_button(submit_label, use_container_width=True)
    if submitted:
        return new_values
    return None


def _edit_selected_row(
    section_key: str,
    df: pd.DataFrame,
    selected_idx: Optional[int],
) -> pd.DataFrame:
    """Allow inline editing of the currently selected row."""

    if df.empty or selected_idx is None:
        st.caption("Select a row to edit.")
        return df

    columns = list(df.columns)
    initial_values = df.loc[selected_idx].to_dict()
    edited_values = _render_row_form(
        section_key=section_key,
        form_key="edit",
        title="Edit selected row",
        columns=columns,
        initial_values=initial_values,
        submit_label="Save changes",
    )
    if edited_values is not None:
        for col, val in edited_values.items():
            _set_dataframe_cell(df, selected_idx, col, val)
        st.session_state[section_key] = df
        st.success("Row updated")
    return st.session_state.get(section_key, df)


def _add_row_via_form(
    section_key: str,
    df: pd.DataFrame,
    blank_row_factory: Callable[[pd.DataFrame], Dict],
    select_key: str,
    id_column: Optional[str],
    name_column: Optional[str],
) -> pd.DataFrame:
    """Render an add-row form so users can insert new entries with custom values."""

    template_row = blank_row_factory(df.copy())
    columns = list(df.columns) if not df.empty else list(template_row.keys())
    initial_values = {col: template_row.get(col, "") for col in columns}
    new_row = _render_row_form(
        section_key=section_key,
        form_key="add",
        title="Add a new row",
        columns=columns,
        initial_values=initial_values,
        submit_label="Add row",
    )
    if new_row is not None:
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        st.session_state[section_key] = df
        identity_column = _selection_identity_column(df, id_column, name_column)
        _set_pending_selection(select_key, _row_identifier(df, df.index[-1], identity_column))
        _set_pending_panel_state(_panel_state_key(section_key, "add"), False)
        st.success("Row added")
    return st.session_state.get(section_key, df)


def _remove_selected_row(
    section_key: str,
    df: pd.DataFrame,
    selected_idx: Optional[int],
    select_key: str,
    id_column: Optional[str],
    name_column: Optional[str],
) -> pd.DataFrame:
    """Delete the selected row when the user confirms the removal."""

    disabled = df.empty or selected_idx is None or selected_idx not in df.index
    if st.button(
        "Remove row",
        key=f"{section_key}_remove",
        use_container_width=True,
        disabled=disabled,
    ):
        if selected_idx is not None and selected_idx in df.index:
            df = df.drop(index=selected_idx).reset_index(drop=True)
            st.session_state[section_key] = df
            if not df.empty:
                identity_column = _selection_identity_column(df, id_column, name_column)
                _set_pending_selection(select_key, _row_identifier(df, df.index[-1], identity_column))
            else:
                _set_pending_selection(select_key, None)
            st.success("Row removed")
    return st.session_state.get(section_key, df)


def _render_product_assumption_table(
    *,
    session_key: str,
    default_factory: Callable[[], pd.DataFrame],
    blank_row_factory: Callable[[pd.DataFrame], Dict],
    column_config: Optional[Dict] = None,
    id_column: Optional[str] = "ID_vaccine",
    name_column: Optional[str] = "Vaccine name",
) -> pd.DataFrame:
    df = _ensure_table_state(session_key, default_factory).copy()
    select_key = f"{session_key}_row_select"
    selected_idx = _render_row_selector(df, select_key, id_column, name_column)
    edit_panel_key = _panel_state_key(session_key, "edit")
    add_panel_key = _panel_state_key(session_key, "add")
    increment_panel_key = _panel_state_key(session_key, "increment")
    _consume_pending_panel_state(edit_panel_key)
    _consume_pending_panel_state(add_panel_key)
    _consume_pending_panel_state(increment_panel_key)

    action_cols = st.columns(4)
    with action_cols[0]:
        edit_open = st.checkbox(
            "Edit",
            value=bool(st.session_state.get(edit_panel_key, False)),
            key=edit_panel_key,
            help="Open the focused row editor for the selected row.",
        )
    with action_cols[1]:
        add_open = st.checkbox(
            "Add Row",
            value=bool(st.session_state.get(add_panel_key, False)),
            key=add_panel_key,
            help="Open the add-row form.",
        )
    with action_cols[2]:
        df = _remove_selected_row(
            session_key,
            df,
            selected_idx,
            select_key,
            id_column,
            name_column,
        )
    with action_cols[3]:
        increment_open = st.checkbox(
            "Yearly Increment",
            value=bool(st.session_state.get(increment_panel_key, False)),
            key=increment_panel_key,
            help="Open the yearly increment helper for the selected row.",
        )

    if edit_open:
        st.caption("Focused row editor: update one selected row at a time.")
        df = _edit_selected_row(session_key, df, selected_idx)
    else:
        st.caption("Tick `Edit` to open the focused row editor for the selected row.")

    if add_open:
        st.caption("Add a new row with all fields visible before it is inserted into the table.")
        df = _add_row_via_form(session_key, df, blank_row_factory, select_key, id_column, name_column)

    if increment_open:
        st.caption("Apply a fixed yearly change or compounded growth from the selected row onward.")
        df = _apply_yearly_increment(session_key, df, selected_idx)

    df = st.session_state.get(session_key, df)
    if session_key == "vaccine_sales_table":
        df = _recompute_vaccine_sales_implied_revenue(df)
    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        hide_index=True,
        key=f"{session_key}_editor",
        column_config=column_config,
    )
    if session_key == "vaccine_sales_table":
        edited_df = _recompute_vaccine_sales_implied_revenue(edited_df)
    st.session_state[session_key] = edited_df
    _validate_selection(edited_df, select_key, id_column, name_column)
    return edited_df


def _default_ramp_schedule() -> pd.DataFrame:
    """Return the seed schedule for global sales ramp factors."""

    default_ramp = [0.2, 0.6, 1.0, 1.0, 1.0]
    data = {
        "Year offset": list(range(len(default_ramp))),
        "Ramp factor": default_ramp,
    }
    return pd.DataFrame(data)


def _default_stage_schedule_mapping() -> pd.DataFrame:
    """Default mapping from pipeline stage to schedule assumptions."""

    data = [
        {
            "Stage": "Discovery",
            "Success Probability %": 10.0,
            "Time to market (years)": 7,
            "Sales ramp length (years)": 5,
            "Ramp shape": "Linear",
            "R&D remaining pre-launch (USD)": 300_000_000,
            "R&D annual post-launch (USD/year)": 15_000_000,
            "Discovery duration (years)": 1,
            "Preclinical duration (years)": 1,
            "Phase I duration (years)": 1,
            "Phase II duration (years)": 2,
            "Phase III duration (years)": 1,
            "Approval duration (years)": 1,
            "Commercial duration (years)": 0,
            "Discovery R&D weight %": 10.0,
            "Preclinical R&D weight %": 15.0,
            "Phase I R&D weight %": 15.0,
            "Phase II R&D weight %": 25.0,
            "Phase III R&D weight %": 25.0,
            "Approval R&D weight %": 10.0,
            "Discovery CAPEX weight %": 5.0,
            "Preclinical CAPEX weight %": 10.0,
            "Phase I CAPEX weight %": 10.0,
            "Phase II CAPEX weight %": 25.0,
            "Phase III CAPEX weight %": 35.0,
            "Approval CAPEX weight %": 15.0,
            "Discovery completion milestone (USD)": 0.0,
            "Preclinical completion milestone (USD)": 0.0,
            "Phase I completion milestone (USD)": 5_000_000,
            "Phase II completion milestone (USD)": 15_000_000,
            "Phase III completion milestone (USD)": 25_000_000,
            "Approval completion milestone (USD)": 30_000_000,
            "Discovery->Preclinical": 60.0,
            "Preclinical->Phase I": 70.0,
            "Phase I->Phase II": 65.0,
            "Phase II->Phase III": 55.0,
            "Phase III->Approval": 65.0,
            "Approval->Commercial": 100.0,
            "Discovery->Preclinical annual success %": 60.0,
            "Preclinical->Phase I annual success %": 70.0,
            "Phase I->Phase II annual success %": 65.0,
            "Phase II->Phase III annual success %": 55.0,
            "Phase III->Approval annual success %": 65.0,
            "Approval->Commercial annual success %": 100.0,
        },
        {
            "Stage": "Preclinical",
            "Success Probability %": 20.0,
            "Time to market (years)": 6,
            "Sales ramp length (years)": 5,
            "Ramp shape": "Linear",
            "R&D remaining pre-launch (USD)": 250_000_000,
            "R&D annual post-launch (USD/year)": 12_000_000,
            "Discovery duration (years)": 1,
            "Preclinical duration (years)": 1,
            "Phase I duration (years)": 1,
            "Phase II duration (years)": 2,
            "Phase III duration (years)": 1,
            "Approval duration (years)": 1,
            "Commercial duration (years)": 0,
            "Discovery R&D weight %": 10.0,
            "Preclinical R&D weight %": 15.0,
            "Phase I R&D weight %": 15.0,
            "Phase II R&D weight %": 25.0,
            "Phase III R&D weight %": 25.0,
            "Approval R&D weight %": 10.0,
            "Discovery CAPEX weight %": 5.0,
            "Preclinical CAPEX weight %": 10.0,
            "Phase I CAPEX weight %": 10.0,
            "Phase II CAPEX weight %": 25.0,
            "Phase III CAPEX weight %": 35.0,
            "Approval CAPEX weight %": 15.0,
            "Discovery completion milestone (USD)": 0.0,
            "Preclinical completion milestone (USD)": 0.0,
            "Phase I completion milestone (USD)": 5_000_000,
            "Phase II completion milestone (USD)": 15_000_000,
            "Phase III completion milestone (USD)": 25_000_000,
            "Approval completion milestone (USD)": 30_000_000,
            "Discovery->Preclinical": 100.0,
            "Preclinical->Phase I": 70.0,
            "Phase I->Phase II": 65.0,
            "Phase II->Phase III": 55.0,
            "Phase III->Approval": 65.0,
            "Approval->Commercial": 100.0,
            "Discovery->Preclinical annual success %": 100.0,
            "Preclinical->Phase I annual success %": 70.0,
            "Phase I->Phase II annual success %": 65.0,
            "Phase II->Phase III annual success %": 55.0,
            "Phase III->Approval annual success %": 65.0,
            "Approval->Commercial annual success %": 100.0,
        },
        {
            "Stage": "Phase I",
            "Success Probability %": 35.0,
            "Time to market (years)": 5,
            "Sales ramp length (years)": 4,
            "Ramp shape": "Linear",
            "R&D remaining pre-launch (USD)": 200_000_000,
            "R&D annual post-launch (USD/year)": 10_000_000,
            "Discovery duration (years)": 1,
            "Preclinical duration (years)": 1,
            "Phase I duration (years)": 1,
            "Phase II duration (years)": 2,
            "Phase III duration (years)": 1,
            "Approval duration (years)": 1,
            "Commercial duration (years)": 0,
            "Discovery R&D weight %": 10.0,
            "Preclinical R&D weight %": 15.0,
            "Phase I R&D weight %": 15.0,
            "Phase II R&D weight %": 25.0,
            "Phase III R&D weight %": 25.0,
            "Approval R&D weight %": 10.0,
            "Discovery CAPEX weight %": 5.0,
            "Preclinical CAPEX weight %": 10.0,
            "Phase I CAPEX weight %": 10.0,
            "Phase II CAPEX weight %": 25.0,
            "Phase III CAPEX weight %": 35.0,
            "Approval CAPEX weight %": 15.0,
            "Discovery completion milestone (USD)": 0.0,
            "Preclinical completion milestone (USD)": 0.0,
            "Phase I completion milestone (USD)": 5_000_000,
            "Phase II completion milestone (USD)": 15_000_000,
            "Phase III completion milestone (USD)": 25_000_000,
            "Approval completion milestone (USD)": 30_000_000,
            "Discovery->Preclinical": 100.0,
            "Preclinical->Phase I": 100.0,
            "Phase I->Phase II": 65.0,
            "Phase II->Phase III": 55.0,
            "Phase III->Approval": 65.0,
            "Approval->Commercial": 100.0,
            "Discovery->Preclinical annual success %": 100.0,
            "Preclinical->Phase I annual success %": 100.0,
            "Phase I->Phase II annual success %": 65.0,
            "Phase II->Phase III annual success %": 55.0,
            "Phase III->Approval annual success %": 65.0,
            "Approval->Commercial annual success %": 100.0,
        },
        {
            "Stage": "Phase II",
            "Success Probability %": 45.0,
            "Time to market (years)": 4,
            "Sales ramp length (years)": 4,
            "Ramp shape": "Linear",
            "R&D remaining pre-launch (USD)": 150_000_000,
            "R&D annual post-launch (USD/year)": 9_000_000,
            "Discovery duration (years)": 1,
            "Preclinical duration (years)": 1,
            "Phase I duration (years)": 1,
            "Phase II duration (years)": 2,
            "Phase III duration (years)": 1,
            "Approval duration (years)": 1,
            "Commercial duration (years)": 0,
            "Discovery R&D weight %": 10.0,
            "Preclinical R&D weight %": 15.0,
            "Phase I R&D weight %": 15.0,
            "Phase II R&D weight %": 25.0,
            "Phase III R&D weight %": 25.0,
            "Approval R&D weight %": 10.0,
            "Discovery CAPEX weight %": 5.0,
            "Preclinical CAPEX weight %": 10.0,
            "Phase I CAPEX weight %": 10.0,
            "Phase II CAPEX weight %": 25.0,
            "Phase III CAPEX weight %": 35.0,
            "Approval CAPEX weight %": 15.0,
            "Discovery completion milestone (USD)": 0.0,
            "Preclinical completion milestone (USD)": 0.0,
            "Phase I completion milestone (USD)": 5_000_000,
            "Phase II completion milestone (USD)": 15_000_000,
            "Phase III completion milestone (USD)": 25_000_000,
            "Approval completion milestone (USD)": 30_000_000,
            "Discovery->Preclinical": 100.0,
            "Preclinical->Phase I": 100.0,
            "Phase I->Phase II": 100.0,
            "Phase II->Phase III": 55.0,
            "Phase III->Approval": 65.0,
            "Approval->Commercial": 100.0,
            "Discovery->Preclinical annual success %": 100.0,
            "Preclinical->Phase I annual success %": 100.0,
            "Phase I->Phase II annual success %": 100.0,
            "Phase II->Phase III annual success %": 55.0,
            "Phase III->Approval annual success %": 65.0,
            "Approval->Commercial annual success %": 100.0,
        },
        {
            "Stage": "Phase III",
            "Success Probability %": 60.0,
            "Time to market (years)": 3,
            "Sales ramp length (years)": 3,
            "Ramp shape": "Linear",
            "R&D remaining pre-launch (USD)": 100_000_000,
            "R&D annual post-launch (USD/year)": 8_000_000,
            "Discovery duration (years)": 1,
            "Preclinical duration (years)": 1,
            "Phase I duration (years)": 1,
            "Phase II duration (years)": 2,
            "Phase III duration (years)": 1,
            "Approval duration (years)": 1,
            "Commercial duration (years)": 0,
            "Discovery R&D weight %": 10.0,
            "Preclinical R&D weight %": 15.0,
            "Phase I R&D weight %": 15.0,
            "Phase II R&D weight %": 25.0,
            "Phase III R&D weight %": 25.0,
            "Approval R&D weight %": 10.0,
            "Discovery CAPEX weight %": 5.0,
            "Preclinical CAPEX weight %": 10.0,
            "Phase I CAPEX weight %": 10.0,
            "Phase II CAPEX weight %": 25.0,
            "Phase III CAPEX weight %": 35.0,
            "Approval CAPEX weight %": 15.0,
            "Discovery completion milestone (USD)": 0.0,
            "Preclinical completion milestone (USD)": 0.0,
            "Phase I completion milestone (USD)": 5_000_000,
            "Phase II completion milestone (USD)": 15_000_000,
            "Phase III completion milestone (USD)": 25_000_000,
            "Approval completion milestone (USD)": 30_000_000,
            "Discovery->Preclinical": 100.0,
            "Preclinical->Phase I": 100.0,
            "Phase I->Phase II": 100.0,
            "Phase II->Phase III": 100.0,
            "Phase III->Approval": 65.0,
            "Approval->Commercial": 100.0,
            "Discovery->Preclinical annual success %": 100.0,
            "Preclinical->Phase I annual success %": 100.0,
            "Phase I->Phase II annual success %": 100.0,
            "Phase II->Phase III annual success %": 100.0,
            "Phase III->Approval annual success %": 65.0,
            "Approval->Commercial annual success %": 100.0,
        },
        {
            "Stage": "Approval",
            "Success Probability %": 80.0,
            "Time to market (years)": 1,
            "Sales ramp length (years)": 2,
            "Ramp shape": "Linear",
            "R&D remaining pre-launch (USD)": 50_000_000,
            "R&D annual post-launch (USD/year)": 6_000_000,
            "Discovery duration (years)": 1,
            "Preclinical duration (years)": 1,
            "Phase I duration (years)": 1,
            "Phase II duration (years)": 2,
            "Phase III duration (years)": 1,
            "Approval duration (years)": 1,
            "Commercial duration (years)": 0,
            "Discovery R&D weight %": 10.0,
            "Preclinical R&D weight %": 15.0,
            "Phase I R&D weight %": 15.0,
            "Phase II R&D weight %": 25.0,
            "Phase III R&D weight %": 25.0,
            "Approval R&D weight %": 10.0,
            "Discovery CAPEX weight %": 5.0,
            "Preclinical CAPEX weight %": 10.0,
            "Phase I CAPEX weight %": 10.0,
            "Phase II CAPEX weight %": 25.0,
            "Phase III CAPEX weight %": 35.0,
            "Approval CAPEX weight %": 15.0,
            "Discovery completion milestone (USD)": 0.0,
            "Preclinical completion milestone (USD)": 0.0,
            "Phase I completion milestone (USD)": 5_000_000,
            "Phase II completion milestone (USD)": 15_000_000,
            "Phase III completion milestone (USD)": 25_000_000,
            "Approval completion milestone (USD)": 30_000_000,
            "Discovery->Preclinical": 100.0,
            "Preclinical->Phase I": 100.0,
            "Phase I->Phase II": 100.0,
            "Phase II->Phase III": 100.0,
            "Phase III->Approval": 100.0,
            "Approval->Commercial": 100.0,
            "Discovery->Preclinical annual success %": 100.0,
            "Preclinical->Phase I annual success %": 100.0,
            "Phase I->Phase II annual success %": 100.0,
            "Phase II->Phase III annual success %": 100.0,
            "Phase III->Approval annual success %": 100.0,
            "Approval->Commercial annual success %": 100.0,
        },
        {
            "Stage": "Commercial",
            "Success Probability %": 100.0,
            "Time to market (years)": 0,
            "Sales ramp length (years)": 1,
            "Ramp shape": "Step",
            "R&D remaining pre-launch (USD)": 0.0,
            "R&D annual post-launch (USD/year)": 5_000_000,
            "Discovery duration (years)": 1,
            "Preclinical duration (years)": 1,
            "Phase I duration (years)": 1,
            "Phase II duration (years)": 2,
            "Phase III duration (years)": 1,
            "Approval duration (years)": 1,
            "Commercial duration (years)": 0,
            "Discovery R&D weight %": 10.0,
            "Preclinical R&D weight %": 15.0,
            "Phase I R&D weight %": 15.0,
            "Phase II R&D weight %": 25.0,
            "Phase III R&D weight %": 25.0,
            "Approval R&D weight %": 10.0,
            "Discovery CAPEX weight %": 5.0,
            "Preclinical CAPEX weight %": 10.0,
            "Phase I CAPEX weight %": 10.0,
            "Phase II CAPEX weight %": 25.0,
            "Phase III CAPEX weight %": 35.0,
            "Approval CAPEX weight %": 15.0,
            "Discovery completion milestone (USD)": 0.0,
            "Preclinical completion milestone (USD)": 0.0,
            "Phase I completion milestone (USD)": 5_000_000,
            "Phase II completion milestone (USD)": 15_000_000,
            "Phase III completion milestone (USD)": 25_000_000,
            "Approval completion milestone (USD)": 30_000_000,
            "Discovery->Preclinical": 100.0,
            "Preclinical->Phase I": 100.0,
            "Phase I->Phase II": 100.0,
            "Phase II->Phase III": 100.0,
            "Phase III->Approval": 100.0,
            "Approval->Commercial": 100.0,
            "Discovery->Preclinical annual success %": 100.0,
            "Preclinical->Phase I annual success %": 100.0,
            "Phase I->Phase II annual success %": 100.0,
            "Phase II->Phase III annual success %": 100.0,
            "Phase III->Approval annual success %": 100.0,
            "Approval->Commercial annual success %": 100.0,
        },
    ]
    return pd.DataFrame(data)


def _stage_duration_years_from_row(row: pd.Series) -> Dict[str, int]:
    durations: Dict[str, int] = {}
    for col in STAGE_DURATION_COLUMNS:
        if col not in row:
            continue
        stage = col.replace(" duration (years)", "")
        value = row.get(col)
        if pd.isna(value):
            continue
        durations[stage] = max(0, int(value))
    return durations


def _stage_cost_weights_from_row(row: pd.Series) -> Dict[str, float]:
    weights: Dict[str, float] = {}
    for col in STAGE_COST_WEIGHT_COLUMNS:
        if col not in row:
            continue
        stage = col.replace(" R&D weight %", "")
        value = row.get(col)
        if pd.isna(value):
            continue
        weight = float(value)
        if weight > 1.0:
            weight = weight / 100.0
        weights[stage] = max(0.0, weight)
    return weights


def _stage_capex_weights_from_row(row: pd.Series) -> Dict[str, float]:
    weights: Dict[str, float] = {}
    for col in STAGE_CAPEX_WEIGHT_COLUMNS:
        if col not in row:
            continue
        stage = col.replace(" CAPEX weight %", "")
        value = row.get(col)
        if pd.isna(value):
            continue
        weight = float(value)
        if weight > 1.0:
            weight = weight / 100.0
        weights[stage] = max(0.0, weight)
    return weights


def _compute_time_to_market_from_durations(stage: str, durations: Dict[str, int]) -> Optional[int]:
    stage = normalize_stage_label(stage)
    if stage not in STAGE_SEQUENCE or not durations:
        return None
    stage_idx = STAGE_SEQUENCE.index(stage)
    total = 0
    for idx in range(stage_idx, len(STAGE_SEQUENCE) - 1):
        from_stage = STAGE_SEQUENCE[idx]
        total += int(durations.get(from_stage, 0))
    return total if total > 0 else 0


def _stage_mapping_row(mapping_df: pd.DataFrame, stage: str) -> Optional[pd.Series]:
    if mapping_df is None or mapping_df.empty or not stage:
        return None
    if "Stage" not in mapping_df.columns:
        return None
    normalized_stage = normalize_stage_label(stage)
    matches = mapping_df[mapping_df["Stage"].astype(str).map(normalize_stage_label) == normalized_stage]
    if matches.empty:
        return None
    return matches.iloc[0]


def _stage_transition_probabilities_from_row(row: pd.Series) -> Dict[str, float]:
    transitions: Dict[str, float] = {}
    for col in STAGE_TRANSITION_COLUMNS:
        value = row.get(col)
        if pd.isna(value):
            continue
        prob = float(value)
        if prob > 1.0:
            prob = prob / 100.0
        transitions[col] = max(0.0, min(1.0, prob))
    return transitions


def _stage_transition_curve_from_row(
    row: pd.Series,
    durations: Dict[str, int],
) -> Dict[str, List[float]]:
    curves: Dict[str, List[float]] = {}
    for col in STAGE_TRANSITION_ANNUAL_COLUMNS:
        value = row.get(col)
        if pd.isna(value):
            continue
        transition = col.replace(" annual success %", "")
        from_stage = transition.split("->")[0]
        duration = int(durations.get(from_stage, 0))
        if duration <= 0:
            continue
        prob = float(value)
        if prob > 1.0:
            prob = prob / 100.0
        prob = max(0.0, min(1.0, prob))
        curves[transition] = [prob] * duration
    return curves


def _stage_milestones_from_row(
    row: pd.Series,
    durations: Dict[str, int],
    transitions: Dict[str, float],
) -> List[Milestone]:
    milestones: List[Milestone] = []
    if not durations:
        return milestones
    cumulative_years = 0
    for stage in STAGE_SEQUENCE:
        duration = int(durations.get(stage, 0))
        cumulative_years += duration
        col = f"{stage} completion milestone (USD)"
        if col not in row:
            continue
        amount = row.get(col)
        if pd.isna(amount) or float(amount) == 0.0:
            continue
        milestone = Milestone(
            name=f"{stage} completion milestone",
            year_offset=cumulative_years,
            amount=float(amount),
            # Stage-transition schedules already risk-adjust the product cash flows
            # over time, so generated milestones should not embed the same
            # transition probability a second time.
            probability=1.0,
            timing="from_start",
        )
        milestones.append(milestone)
    return milestones


def _apply_stage_schedule_defaults(
    df: pd.DataFrame,
    mapping_df: pd.DataFrame,
    *,
    stage_column: str,
    overwrite: bool,
) -> pd.DataFrame:
    if df.empty or mapping_df is None or mapping_df.empty or stage_column not in df.columns:
        return df
    updated = df.copy()
    for idx, row in updated.iterrows():
        stage = row.get(stage_column)
        mapping_row = _stage_mapping_row(mapping_df, stage)
        if mapping_row is None:
            continue
        durations = _stage_duration_years_from_row(mapping_row)
        derived_time_to_market = _compute_time_to_market_from_durations(stage, durations)
        defaults = {
            "success_prob": mapping_row.get("Success Probability %"),
            "time_to_market": derived_time_to_market
            if derived_time_to_market is not None
            else mapping_row.get("Time to market (years)"),
            "sales_ramp_length": mapping_row.get("Sales ramp length (years)"),
            "sales_ramp_shape": mapping_row.get("Ramp shape"),
            "rd_remaining_pre_launch": mapping_row.get("R&D remaining pre-launch (USD)"),
            "rd_annual_post_launch": mapping_row.get("R&D annual post-launch (USD/year)"),
        }
        for col, value in defaults.items():
            if pd.isna(value):
                continue
            existing = row.get(col)
            if overwrite or pd.isna(existing) or existing in (None, ""):
                if col == "success_prob" and value > 1.0:
                    updated.at[idx, col] = float(value) / 100.0
                else:
                    updated.at[idx, col] = value
    return updated


def _stage_mapping_editor_token(value: str) -> str:
    token = re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_")
    return token or "field"


def _stage_mapping_input_key(stage: str, revision: int, column: str) -> str:
    return (
        "stage_mapping_editor_"
        f"{_stage_mapping_editor_token(stage)}_{revision}_{_stage_mapping_editor_token(column)}"
    )


def _build_stage_mapping_candidate_row(
    base_row: pd.Series,
    updates: Dict[str, Any],
) -> pd.Series:
    candidate = base_row.copy()
    for col, value in updates.items():
        candidate.loc[col] = value

    stage_value = normalize_stage_label(candidate.get("Stage")) or normalize_stage_label(base_row.get("Stage"))
    if stage_value:
        candidate.loc["Stage"] = stage_value

    derived_time = _compute_time_to_market_from_durations(
        str(candidate.get("Stage") or ""),
        _stage_duration_years_from_row(candidate),
    )
    if derived_time is None:
        derived_time = max(0, int(_as_float(candidate.get("Time to market (years)"), 0.0)))
    candidate.loc["Time to market (years)"] = int(derived_time)
    return candidate


def _stage_mapping_row_warnings(
    mapping_df: pd.DataFrame,
    row_idx: int,
    candidate_row: pd.Series,
) -> List[str]:
    preview_df = mapping_df.copy()
    for col in preview_df.columns:
        if col in candidate_row.index:
            preview_df.at[row_idx, col] = candidate_row.get(col)

    stage_label = normalize_stage_label(candidate_row.get("Stage"))
    if not stage_label:
        return _stage_mapping_sanity_checks(preview_df)

    prefix = f"{stage_label}:"
    return [warning for warning in _stage_mapping_sanity_checks(preview_df) if warning.startswith(prefix)]


def _default_debt_schedule(first_year: int, n_years: int) -> pd.DataFrame:
    years = list(range(int(first_year), int(first_year) + int(n_years)))
    seed_drawdowns = [60_000_000.0, 20_000_000.0, 20_000_000.0, 20_000_000.0]
    drawdowns = seed_drawdowns[: len(years)] + [0.0] * max(0, len(years) - len(seed_drawdowns))
    return pd.DataFrame(
        {
            "Year": years,
            "Debt drawdowns": drawdowns,
            "Manual debt repayments": [0.0] * len(years),
        }
    )


def _blank_debt_schedule_row(df: pd.DataFrame, first_year: int, n_years: int) -> Dict:
    if df.empty or "Year" not in df.columns:
        year = int(first_year)
    else:
        year = int(pd.to_numeric(df["Year"], errors="coerce").max() or first_year) + 1
    return {
        "Year": year,
        "Debt drawdowns": 0.0,
        "Manual debt repayments": 0.0,
    }


def _coerce_numeric(series: pd.Series, default: float = 0.0) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(default)


def _set_dataframe_cell(df: pd.DataFrame, row_idx: int, column_name: str, value: object) -> None:
    dtype = df[column_name].dtype
    if value is None and not pd.api.types.is_object_dtype(dtype):
        df[column_name] = df[column_name].astype("object")
    elif isinstance(value, str) and pd.api.types.is_numeric_dtype(dtype):
        df[column_name] = df[column_name].astype("object")

    try:
        df.at[row_idx, column_name] = value
    except (TypeError, ValueError):
        df[column_name] = df[column_name].astype("object")
        df.at[row_idx, column_name] = value


def _coerce_frame_column(df: pd.DataFrame, column: str, default: float = 0.0) -> pd.Series:
    if column in df.columns:
        return pd.to_numeric(df[column], errors="coerce").fillna(default)
    return pd.Series(default, index=df.index, dtype=float)


def _align_table_to_template(df: Optional[pd.DataFrame], template: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return template.copy()
    aligned = df.copy()
    if aligned.empty and not list(aligned.columns):
        return template.copy()
    for col in template.columns:
        if col not in aligned.columns:
            default_value = template[col].iloc[0] if not template.empty else ""
            aligned[col] = default_value
    ordered_cols = list(template.columns) + [col for col in aligned.columns if col not in template.columns]
    return aligned[ordered_cols]


def _roll_cash_balances(cash_flow_df: pd.DataFrame, opening_cash: float = 0.0) -> pd.DataFrame:
    updated = cash_flow_df.copy()
    net_cash = _coerce_frame_column(updated, "Net change in cash")
    beginning_cash: List[float] = []
    ending_cash: List[float] = []
    current_cash = float(opening_cash or 0.0)
    for year in updated.index:
        beginning_cash.append(current_cash)
        current_cash += float(net_cash.loc[year])
        ending_cash.append(current_cash)
    updated["Beginning cash balance"] = pd.Series(beginning_cash, index=updated.index)
    updated["Ending cash balance"] = pd.Series(ending_cash, index=updated.index)
    return updated


def _recompute_vaccine_sales_implied_revenue(df: pd.DataFrame) -> pd.DataFrame:
    if "Doses (M)" not in df.columns or "Price per dose" not in df.columns:
        return df
    doses = pd.to_numeric(df["Doses (M)"], errors="coerce").fillna(0.0)
    price = pd.to_numeric(df["Price per dose"], errors="coerce").fillna(0.0)
    df = df.copy()
    df["Implied revenue"] = doses * 1e6 * price
    return df


def _render_schedule_editor(title: str, session_key: str) -> pd.DataFrame:
    """Render a reusable schedule editor with manual controls.

    The widget exposes explicit Edit / Add Row / Remove Row controls in addition to a
    "Yearly Increment Helper" that can seed values from a starting point.
    """

    if session_key not in st.session_state:
        st.session_state[session_key] = _default_ramp_schedule().copy()

    schedule_df: pd.DataFrame = st.session_state[session_key]
    st.markdown(f"**{title}**")
    toolbar_cols = st.columns(4)
    edit_mode = toolbar_cols[0].toggle("Edit", value=True, key=f"{session_key}_edit")

    if toolbar_cols[1].button("Add Row", key=f"{session_key}_add"):
        next_year = int(schedule_df["Year offset"].max() + 1) if not schedule_df.empty else 0
        last_value = (
            float(schedule_df["Ramp factor"].iloc[-1]) if not schedule_df.empty else 1.0
        )
        schedule_df.loc[len(schedule_df)] = [next_year, last_value]
        st.session_state[session_key] = schedule_df

    if toolbar_cols[2].button("Remove Row", key=f"{session_key}_remove") and not schedule_df.empty:
        schedule_df = schedule_df.iloc[:-1]
        st.session_state[session_key] = schedule_df

    helper_open = toolbar_cols[3].toggle(
        "Yearly Increment Helper",
        value=False,
        key=f"{session_key}_helper_open",
    )
    if helper_open:
        with _section_block("Yearly Increment Helper", heading_level=5):
            def _filter(df: pd.DataFrame, _selected_id: Optional[str], start_year: int) -> pd.Series:
                return pd.to_numeric(df["Year offset"], errors="coerce").fillna(0).astype(int) >= int(start_year)

            schedule_df = _render_yearly_increment_helper(
                section_key=session_key,
                df=schedule_df,
                year_column="Year offset",
                target_columns=["Ramp factor"],
                filter_builder=_filter,
                start_year_label="Start year offset",
                start_year_default=0,
                periods_default=5,
                increment_default=0.2,
                allow_compound=False,
                create_missing_rows=True,
                base_value_mode="input",
                start_value_default=0.2,
            )
            st.session_state[session_key] = schedule_df

    edited_df = st.data_editor(
        schedule_df,
        hide_index=True,
        disabled=not edit_mode,
        key=f"{session_key}_editor",
    )
    st.session_state[session_key] = edited_df
    return edited_df


def _validate_product_df(df: pd.DataFrame) -> pd.DataFrame:
    """Clamp probability/percentage fields to avoid invalid model assumptions."""

    validated = df.copy()
    if "success_prob" in validated.columns:
        validated["success_prob"] = (
            validated["success_prob"].fillna(0.0).clip(0.0, 1.0)
        )
    if "time_to_market" in validated.columns:
        validated["time_to_market"] = (
            pd.to_numeric(validated["time_to_market"], errors="coerce")
            .fillna(0.0)
            .clip(lower=0.0)
        )
    if "sales_ramp_length" in validated.columns:
        validated["sales_ramp_length"] = (
            pd.to_numeric(validated["sales_ramp_length"], errors="coerce")
            .fillna(0.0)
            .clip(lower=0.0)
        )

    percent_cols = [
        "cogs_patent",
        "cogs_post",
        "labor_pct",
        "overhead_pct",
        "material_pct",
        "sales_marketing_pct",
        "gna_pct",
        "royalty_pct",
        "rd_capitalization_ratio",
    ]
    for col in percent_cols:
        if col in validated.columns:
            upper = 1.0 if col != "royalty_pct" else None
            series = validated[col].fillna(0.0)
            if upper is None:
                validated[col] = series.clip(lower=0.0)
            else:
                validated[col] = series.clip(0.0, upper)

    if "include_in_consolidation" in validated.columns:
        validated["include_in_consolidation"] = validated[
            "include_in_consolidation"
        ].fillna(True)

    return validated


def _normalized_label(value: Any) -> str:
    return str(value or "").strip().lower()


def _as_probability(value: Any) -> float:
    if value is None or pd.isna(value):
        return 0.0
    prob = float(value)
    if prob > 1.0:
        prob = prob / 100.0
    return max(0.0, min(1.0, prob))


def _as_float(value: Any, default: float = 0.0) -> float:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return float(default)
    return float(numeric)


def _capitalization_ratio_from_label(value: Any) -> float:
    text = str(value or "").strip().lower()
    if not text:
        return 0.5
    match = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    if match:
        return max(0.0, min(1.0, float(match.group(1)) / 100.0))
    if "expense" in text:
        return 0.0
    if "capital" in text:
        return 1.0
    return 0.5


def _find_detail_row(df: Optional[pd.DataFrame], vaccine_name: str) -> Optional[pd.Series]:
    if df is None or df.empty or "Vaccine name" not in df.columns:
        return None
    normalized = _normalized_label(vaccine_name)
    matches = df[df["Vaccine name"].astype(str).map(_normalized_label) == normalized]
    if matches.empty:
        return None
    return matches.iloc[0]


def _detail_tables_from_state() -> Dict[str, pd.DataFrame]:
    return {
        "development": st.session_state.get("vaccine_development_table", pd.DataFrame()),
        "market_size_estimation": st.session_state.get("market_size_estimation", pd.DataFrame()),
        "revenue": st.session_state.get("vaccine_revenue_table", pd.DataFrame()),
        "cost": st.session_state.get("vaccine_cost_table", pd.DataFrame()),
        "rd": st.session_state.get("vaccine_rd_table", pd.DataFrame()),
        "capex": st.session_state.get("vaccine_capex_table", pd.DataFrame()),
        "royalty": st.session_state.get("vaccine_royalty_table", pd.DataFrame()),
        "market_share": st.session_state.get("vaccine_market_share_table", pd.DataFrame()),
    }


def _apply_detail_assumption_overrides(
    cleaned: Dict[str, Any],
    detail_tables: Optional[Dict[str, pd.DataFrame]],
) -> Dict[str, Any]:
    if not detail_tables:
        return cleaned

    updated = dict(cleaned)
    product_name = str(updated.get("name") or "").strip()
    if not product_name:
        return updated

    development_row = _find_detail_row(detail_tables.get("development"), product_name)
    market_size_row = _find_detail_row(detail_tables.get("market_size_estimation"), product_name)
    revenue_row = _find_detail_row(detail_tables.get("revenue"), product_name)
    cost_row = _find_detail_row(detail_tables.get("cost"), product_name)
    rd_row = _find_detail_row(detail_tables.get("rd"), product_name)
    capex_row = _find_detail_row(detail_tables.get("capex"), product_name)
    royalty_row = _find_detail_row(detail_tables.get("royalty"), product_name)
    market_share_row = _find_detail_row(detail_tables.get("market_share"), product_name)

    if development_row is not None:
        if pd.notna(development_row.get("Stage")):
            updated["stage"] = str(development_row.get("Stage"))
        if pd.notna(development_row.get("Success Probability %")):
            updated["success_prob"] = _as_probability(development_row.get("Success Probability %"))
        if pd.notna(development_row.get("Consolidation")):
            updated["include_in_consolidation"] = bool(development_row.get("Consolidation"))
        if pd.notna(development_row.get("Time to market")):
            updated["time_to_market"] = max(0, int(_as_float(development_row.get("Time to market"))))
        if pd.notna(development_row.get("Patent duration years")):
            updated["patent_years"] = max(1, int(_as_float(development_row.get("Patent duration years"))))

    patient_population_patent = float(updated.get("patient_population_patent") or 0.0)
    penetration_patent = float(updated.get("penetration_patent") or 0.0)

    if market_size_row is not None:
        tam_customers = _as_float(market_size_row.get("Market size (# customers)"))
        sam_pct = _as_probability(market_size_row.get("Serviceable Available Market (% TAM)"))
        serviceable_customers = tam_customers * sam_pct if tam_customers > 0 and sam_pct > 0 else 0.0
        if serviceable_customers > 0:
            patient_population_patent = serviceable_customers
            updated["patient_population_patent"] = serviceable_customers
            updated["patient_population_post"] = serviceable_customers

    if revenue_row is not None:
        patent_customers = _as_float(revenue_row.get("Patent customers per year"))
        patent_price = _as_float(revenue_row.get("Patent price (USD/customer)"))
        post_customer_adj = _as_probability(revenue_row.get("Post patent customer adj. %"))
        post_price_adj = _as_probability(revenue_row.get("Post patent price adj. %"))
        post_patent_customers = revenue_row.get("Post patent customers per year")
        if pd.isna(post_patent_customers):
            post_patent_customers = patent_customers * (post_customer_adj or 1.0)
        post_patent_price = revenue_row.get("Post patent price (USD/customer)")
        if pd.isna(post_patent_price):
            post_patent_price = patent_price * (post_price_adj or 1.0)
        post_patent_customers = _as_float(post_patent_customers)
        post_patent_price = _as_float(post_patent_price)

        if patient_population_patent > 0 and patent_customers > 0:
            penetration_patent = min(1.0, patent_customers / patient_population_patent)
        else:
            penetration_patent = 1.0 if patent_customers > 0 and patent_price > 0 else 0.0

        updated["patient_population_patent"] = patient_population_patent or patent_customers
        updated["price_per_patient_patent"] = patent_price
        updated["penetration_patent"] = penetration_patent
        updated["patient_population_post"] = post_patent_customers
        updated["price_per_patient_post"] = post_patent_price
        updated["penetration_post"] = 1.0 if post_patent_customers > 0 and post_patent_price > 0 else 0.0
        updated["patent_revenue_target"] = patent_customers * patent_price
        updated["post_patent_revenue_target"] = post_patent_customers * post_patent_price

    if market_share_row is not None:
        market_growth = _as_probability(market_share_row.get("Market growth %"))
        sales_growth = _as_probability(market_share_row.get("Sales growth %"))
        if sales_growth > 0:
            updated["market_growth_patent"] = sales_growth
        if market_growth > 0 or pd.notna(market_share_row.get("Market growth %")):
            updated["market_growth_post"] = market_growth
        if float(updated.get("patent_revenue_target") or 0.0) <= 0.0 and pd.notna(
            market_share_row.get("Revenue target patent (USD)")
        ):
            updated["patent_revenue_target"] = _as_float(market_share_row.get("Revenue target patent (USD)"))
        if float(updated.get("post_patent_revenue_target") or 0.0) <= 0.0 and pd.notna(
            market_share_row.get("Revenue target post (USD)")
        ):
            updated["post_patent_revenue_target"] = _as_float(market_share_row.get("Revenue target post (USD)"))

    if cost_row is not None:
        updated["cogs_patent"] = _as_probability(cost_row.get("COGS patent % of sales"))
        updated["cogs_post"] = _as_probability(cost_row.get("COGS post % of sales"))
        updated["sales_marketing_pct"] = _as_probability(cost_row.get("Marketing annual % of sales"))
        updated["royalty_pct"] = _as_probability(cost_row.get("Royalties cost % of sales"))
        gna_total = _as_float(cost_row.get("G&A total (USD)"))
        revenue_base = float(updated.get("patent_revenue_target") or 0.0)
        if gna_total > 0 and revenue_base > 0:
            updated["gna_pct"] = min(1.0, gna_total / revenue_base)

    if rd_row is not None:
        updated["rd_remaining_pre_launch"] = float(
            _as_float(rd_row.get("Pre-GTM remaining (USD)"))
        )
        updated["rd_annual_post_launch"] = float(
            _as_float(rd_row.get("Post-GTM annual cost (USD/year)"))
        )
        updated["rd_capitalization_ratio"] = _capitalization_ratio_from_label(
            rd_row.get("Cost accounting (capitalisation)")
        )

    if capex_row is not None:
        updated["capex_remaining_pre_launch"] = float(
            _as_float(capex_row.get("Total Pre-GTM capex (USD)"))
        )
        updated["capex_annual_post_launch"] = float(
            _as_float(capex_row.get("Total Post-GTM capex (USD/year)"))
        )

    if royalty_row is not None:
        monetization_model = str(royalty_row.get("Monetization model") or "Product Sale").strip() or "Product Sale"
        updated["commercialization_model"] = monetization_model
        if monetization_model.lower() == "licensing":
            patent_royalty_income = float(
                _as_float(royalty_row.get("Royalty income (USD)"))
            )
            post_patent_revenue = float(
                _as_float(royalty_row.get("Post patent revenue (USD)"))
            )
            royalty_rate = _as_probability(royalty_row.get("Royalty rate (%)"))
            if patent_royalty_income > 0:
                updated["patent_revenue_target"] = patent_royalty_income
                updated["patient_population_patent"] = 0.0
                updated["price_per_patient_patent"] = 0.0
                updated["penetration_patent"] = 0.0
            if post_patent_revenue > 0 and royalty_rate > 0:
                updated["post_patent_revenue_target"] = post_patent_revenue * royalty_rate
                updated["patient_population_post"] = 0.0
                updated["price_per_patient_post"] = 0.0
                updated["penetration_post"] = 0.0

    stage_weights = updated.get("stage_cost_weights") or {}
    rd_remaining = float(updated.get("rd_remaining_pre_launch") or 0.0)
    if stage_weights and rd_remaining > 0:
        current_stage = str(updated.get("stage") or "").strip()
        if current_stage in STAGE_SEQUENCE:
            remaining_stages = set(STAGE_SEQUENCE[STAGE_SEQUENCE.index(current_stage) : -1])
        else:
            remaining_stages = set(stage_weights.keys())
        relevant_weights = {
            stage: float(weight)
            for stage, weight in stage_weights.items()
            if stage in remaining_stages and float(weight) > 0
        }
        total_weight = sum(relevant_weights.values())
        if total_weight > 0:
            updated["trial_costs_by_phase"] = {
                stage: rd_remaining * (weight / total_weight)
                for stage, weight in relevant_weights.items()
            }

    return updated


def _build_probability_preview(
    product_df: pd.DataFrame,
    model_cfg: ModelConfig,
    stage_mapping: Optional[pd.DataFrame],
    *,
    overwrite_defaults: bool,
    detail_tables: Optional[Dict[str, pd.DataFrame]],
) -> pd.DataFrame:
    def _probability_path_label(source: str) -> str:
        if source == "stage_transitions":
            return "Stage-transition path"
        if source == "success_prob_stage_fallback":
            return "Single success probability fallback"
        return "Single success probability"

    rows: List[Dict[str, Any]] = []
    preview_records = _sanitize_product_records(
        product_df,
        stage_mapping=stage_mapping,
        overwrite_defaults=overwrite_defaults,
        detail_tables=detail_tables,
    )
    for record in preview_records:
        try:
            product = Product(ProductConfig(**record), model_cfg)
        except Exception:
            continue
        rows.append(
            {
                "Product": record.get("name"),
                "Stage": record.get("stage"),
                "Probability source": _probability_path_label(product.probability_source()),
                "Probability path used": _probability_path_label(product.probability_source()),
                "Input success probability": float(record.get("success_prob", 0.0) or 0.0),
                "Effective cumulative success probability": product.effective_success_probability(),
                "Time to market (years)": int(record.get("time_to_market", 0) or 0),
            }
        )
    return pd.DataFrame(rows)


def _sanitize_product_records(
    df: pd.DataFrame,
    stage_mapping: Optional[pd.DataFrame] = None,
    overwrite_defaults: bool = False,
    detail_tables: Optional[Dict[str, pd.DataFrame]] = None,
) -> List[Dict]:
    records: List[Dict] = []
    cfg_fields = {f.name for f in fields(ProductConfig)}
    for raw in df.to_dict("records"):
        if not raw.get("name"):
            continue
        cleaned: Dict = {}
        for key, value in raw.items():
            if key not in cfg_fields:
                continue
            if isinstance(value, float) and np.isnan(value):
                continue
            cleaned[key] = value
        cleaned["stage"] = normalize_stage_label(cleaned.get("stage")) or "Unspecified"
        cleaned.setdefault("success_prob", 0.5)
        cleaned.setdefault("include_in_consolidation", True)
        mapping_row = _stage_mapping_row(stage_mapping, cleaned.get("stage"))
        if mapping_row is not None:
            durations = _stage_duration_years_from_row(mapping_row)
            cost_weights = _stage_cost_weights_from_row(mapping_row)
            capex_weights = _stage_capex_weights_from_row(mapping_row)
            if overwrite_defaults or "success_prob" not in cleaned:
                mapped_prob = mapping_row.get("Success Probability %")
                if pd.notna(mapped_prob):
                    mapped_prob = float(mapped_prob)
                    if mapped_prob > 1.0:
                        mapped_prob = mapped_prob / 100.0
                    cleaned["success_prob"] = mapped_prob
            if overwrite_defaults or "time_to_market" not in cleaned:
                mapped_time = _compute_time_to_market_from_durations(cleaned.get("stage"), durations)
                if mapped_time is None:
                    mapped_time = mapping_row.get("Time to market (years)")
                if pd.notna(mapped_time):
                    cleaned["time_to_market"] = mapped_time
            if overwrite_defaults or "sales_ramp_length" not in cleaned:
                mapped_ramp = mapping_row.get("Sales ramp length (years)")
                if pd.notna(mapped_ramp):
                    cleaned["sales_ramp_length"] = mapped_ramp
            if overwrite_defaults or "sales_ramp_shape" not in cleaned:
                mapped_shape = mapping_row.get("Ramp shape")
                if pd.notna(mapped_shape):
                    cleaned["sales_ramp_shape"] = mapped_shape
            if overwrite_defaults or "rd_remaining_pre_launch" not in cleaned:
                mapped_rd = mapping_row.get("R&D remaining pre-launch (USD)")
                if pd.notna(mapped_rd):
                    cleaned["rd_remaining_pre_launch"] = mapped_rd
            if overwrite_defaults or "rd_annual_post_launch" not in cleaned:
                mapped_rd_annual = mapping_row.get("R&D annual post-launch (USD/year)")
                if pd.notna(mapped_rd_annual):
                    cleaned["rd_annual_post_launch"] = mapped_rd_annual
            if durations:
                cleaned["stage_duration_years"] = durations
            if cost_weights:
                cleaned["stage_cost_weights"] = cost_weights
            if capex_weights:
                cleaned["stage_capex_weights"] = capex_weights
            transition_curve = _stage_transition_curve_from_row(mapping_row, durations)
            if transition_curve:
                cleaned["stage_transition_curve"] = transition_curve
            transitions = _stage_transition_probabilities_from_row(mapping_row)
            if transitions:
                cleaned["stage_transition_probabilities"] = transitions
            milestones = _stage_milestones_from_row(mapping_row, durations, transitions)
            if milestones:
                cleaned["milestones"] = [asdict(milestone) for milestone in milestones]
        cleaned = _apply_detail_assumption_overrides(cleaned, detail_tables)
        records.append(cleaned)
    return records


def _build_portfolio(
    product_df: pd.DataFrame,
    model_cfg: ModelConfig,
    stage_mapping: Optional[pd.DataFrame] = None,
    overwrite_defaults: bool = False,
    detail_tables: Optional[Dict[str, pd.DataFrame]] = None,
) -> Portfolio | None:
    product_records = _sanitize_product_records(
        product_df,
        stage_mapping=stage_mapping,
        overwrite_defaults=overwrite_defaults,
        detail_tables=detail_tables,
    )
    if not product_records:
        return None
    products = [Product(ProductConfig(**record), model_cfg) for record in product_records]
    return Portfolio(products, model_cfg)


def _compute_financial_statements(
    cons: pd.DataFrame, model_cfg: ModelConfig
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    years = cons.index
    da_positive = -cons["da"]

    perf_df = pd.DataFrame(
        {
            "Revenue": cons["revenue"],
            "COGS": cons["cogs"],
            "Materials": cons["materials"],
            "Labor": cons["labor"],
            "Overhead": cons["overhead"],
            "Sales & Marketing": cons["sales_marketing"],
            "G&A": cons["gna"],
            "Royalty": cons["royalty"],
            "R&D expense": cons["rd_expense_pnl"],
            "Milestones": cons.get("milestones", pd.Series(0.0, index=years)),
            "EBITDA": cons["ebitda"],
            "EBIT": cons["ebit"],
            "Tax": cons["tax"],
            "NOPAT": cons["nopat"],
        }
    )

    wc = model_cfg.working_capital_pct_sales * cons["revenue"]
    wc_diff = wc.diff().fillna(wc)

    intangible = []
    ppe = []
    working_capital_asset = []
    retained = []
    paid_in = []

    intangible_val = 0.0
    ppe_val = 0.0
    wc_val = 0.0
    retained_val = 0.0

    for year in years:
        rd_cap_add = cons.loc[year, "rd_cap_add"]
        rd_amort = cons.loc[year, "rd_amort"]
        capex_cash = cons.loc[year, "capex_cash"]
        depreciation = cons.loc[year, "depreciation"]
        nopat = cons.loc[year, "nopat"]

        intangible_val += -rd_cap_add + rd_amort
        ppe_val += -capex_cash + depreciation
        wc_val += wc_diff.loc[year]
        retained_val += nopat

        total_assets = intangible_val + ppe_val + wc_val
        paid_in_val = max(0.0, total_assets - retained_val)

        intangible.append(intangible_val)
        ppe.append(ppe_val)
        working_capital_asset.append(wc_val)
        retained.append(retained_val)
        paid_in.append(paid_in_val)

    position_df = pd.DataFrame(
        {
            "Intangibles": intangible,
            "Property & equipment": ppe,
            "Working capital": working_capital_asset,
            "Total assets": np.array(intangible) + np.array(ppe) + np.array(working_capital_asset),
            "Retained earnings": retained,
            "Paid-in capital": paid_in,
            "Total equity": np.array(retained) + np.array(paid_in),
        },
        index=years,
    )

    cash_from_ops = cons["nopat"] + da_positive - wc_diff
    capex_cash = cons["capex_cash"]
    rd_cap_add = cons["rd_cap_add"]
    cash_from_investing = capex_cash + rd_cap_add
    equity_issuance = pd.Series(0.0, index=years)
    debt_draw = pd.Series(0.0, index=years)
    debt_repay = pd.Series(0.0, index=years)
    interest_paid = pd.Series(0.0, index=years)
    cash_from_financing = equity_issuance + debt_draw - debt_repay - interest_paid
    net_cash = cash_from_ops + cash_from_investing + cash_from_financing
    starting_cash = pd.Series(0.0, index=years)
    ending_cash = starting_cash + net_cash.cumsum()

    receivables_change = -wc_diff * 0.5
    inventory_change = -wc_diff * 0.3
    payables_change = -wc_diff * 0.2

    cash_flow_df = pd.DataFrame(
        {
            "EBIT": cons["ebit"],
            "Materials": cons["materials"],
            "Labor": cons["labor"],
            "Overhead": cons["overhead"],
            "Cash taxes paid": cons["tax"],
            "Depreciation & amortization": da_positive,
            "Receivables change": receivables_change,
            "Inventory change": inventory_change,
            "Payables change": payables_change,
            "Working capital change": -wc_diff,
            "Net cash from operations": cash_from_ops,
            "Capital expenditure": capex_cash,
            "R&D capitalization": rd_cap_add,
            "Net cash from investing": cash_from_investing,
            "Equity issuance": equity_issuance,
            "Debt drawdowns": debt_draw,
            "Debt repayments": debt_repay,
            "Interest paid": interest_paid,
            "Net cash from financing": cash_from_financing,
            "Net change in cash": net_cash,
            "Beginning cash balance": starting_cash,
            "Ending cash balance": ending_cash,
        }
    )
    cash_flow_df = _roll_cash_balances(cash_flow_df, opening_cash=0.0)

    return perf_df, position_df, cash_flow_df


def _build_ratio_table(cons: pd.DataFrame) -> pd.DataFrame:
    revenue = cons["revenue"].replace(0, np.nan)
    gross_profit = cons["revenue"] + cons["cogs"]
    ratios = pd.DataFrame(index=cons.index)
    ratios["Gross margin"] = gross_profit / revenue
    ratios["EBITDA margin"] = cons["ebitda"] / revenue
    ratios["NOPAT margin"] = cons["nopat"] / revenue
    ratios["R&D intensity"] = cons["rd_cash"].abs() / revenue
    ratios["Capex intensity"] = (-cons["capex_cash"]) / revenue
    return ratios.fillna(0.0)


def _build_vaccine_break_even_inputs(model_cfg: Optional[ModelConfig]) -> pd.DataFrame:
    if model_cfg is None:
        return pd.DataFrame()
    dev_df = st.session_state.get("vaccine_development_table", pd.DataFrame()).copy()
    if dev_df.empty or "ID_vaccine" not in dev_df.columns:
        return pd.DataFrame()

    revenue_df = st.session_state.get("vaccine_revenue_table", pd.DataFrame()).copy()
    for col in ["ID_vaccine", "Vaccine name"]:
        if col not in revenue_df.columns:
            revenue_df[col] = dev_df.get(col, pd.Series(dtype=str))
    if "Patent revenue target (USD)" not in revenue_df.columns:
        revenue_df["Patent revenue target (USD)"] = _coerce_numeric(
            revenue_df.get("Patent customers per year", pd.Series(dtype=float))
        ) * _coerce_numeric(revenue_df.get("Patent price (USD/customer)", pd.Series(dtype=float)))
    if "Post patent revenue target (USD)" not in revenue_df.columns:
        revenue_df["Post patent revenue target (USD)"] = _coerce_numeric(
            revenue_df.get("Post patent customers per year", pd.Series(dtype=float))
        ) * _coerce_numeric(revenue_df.get("Post patent price (USD/customer)", pd.Series(dtype=float)))

    cost_df = st.session_state.get("vaccine_cost_table", pd.DataFrame()).copy()
    for col in ["ID_vaccine", "Vaccine name"]:
        if col not in cost_df.columns:
            cost_df[col] = dev_df.get(col, pd.Series(dtype=str))
    gna_cols = [
        "Indirect staff cost (USD)",
        "Electricity (USD)",
        "Depreciation (USD)",
        "Interest & amortization (USD)",
    ]
    if "G&A total (USD)" not in cost_df.columns:
        available_gna = [col for col in gna_cols if col in cost_df.columns]
        if available_gna:
            cost_df["G&A total (USD)"] = cost_df[available_gna].sum(axis=1)
        else:
            cost_df["G&A total (USD)"] = 0.0
    if "Patent operating cost %" not in cost_df.columns:
        cost_df["Patent operating cost %"] = (
            _coerce_numeric(cost_df.get("COGS patent % of sales", pd.Series(dtype=float)))
            + _coerce_numeric(cost_df.get("Marketing annual % of sales", pd.Series(dtype=float)))
            + _coerce_numeric(cost_df.get("Royalties cost % of sales", pd.Series(dtype=float)))
        )

    rd_df = st.session_state.get("vaccine_rd_table", pd.DataFrame()).copy()
    for col in ["ID_vaccine", "Vaccine name"]:
        if col not in rd_df.columns:
            rd_df[col] = dev_df.get(col, pd.Series(dtype=str))
    if "Pre-GTM total (USD)" not in rd_df.columns:
        rd_df["Pre-GTM total (USD)"] = _coerce_numeric(
            rd_df.get("Pre-GTM spent to date (USD)", pd.Series(dtype=float))
        ) + _coerce_numeric(rd_df.get("Pre-GTM remaining (USD)", pd.Series(dtype=float)))

    capex_df = st.session_state.get("vaccine_capex_table", pd.DataFrame()).copy()
    for col in ["ID_vaccine", "Vaccine name"]:
        if col not in capex_df.columns:
            capex_df[col] = dev_df.get(col, pd.Series(dtype=str))
    capex_pre_cols = [
        "Manufacturing & Scale-up Assets (Pre-GTM, USD)",
        "Quality & Compliance Infrastructure (Pre-GTM, USD)",
        "Cold-chain / Distribution Assets (Pre-GTM, USD)",
        "IT / Data / Digital Infrastructure (Pre-GTM, USD)",
        "Facility Build-out / Leasehold Improvements (Pre-GTM, USD)",
        "Process Development & Tech-Transfer Assets (Pre-GTM, USD)",
    ]
    capex_post_cols = [
        "Manufacturing & Scale-up Assets (Post-GTM, USD/year)",
        "Quality & Compliance Infrastructure (Post-GTM, USD/year)",
        "Cold-chain / Distribution Assets (Post-GTM, USD/year)",
        "IT / Data / Digital Infrastructure (Post-GTM, USD/year)",
        "Facility Build-out / Leasehold Improvements (Post-GTM, USD/year)",
        "Process Development & Tech-Transfer Assets (Post-GTM, USD/year)",
    ]
    if "Total Pre-GTM capex (USD)" not in capex_df.columns:
        capex_pre = capex_df.get(capex_pre_cols, pd.DataFrame()).apply(
            pd.to_numeric, errors="coerce"
        )
        capex_df["Total Pre-GTM capex (USD)"] = capex_pre.fillna(0.0).sum(axis=1)
    if "Total Post-GTM capex (USD/year)" not in capex_df.columns:
        capex_post = capex_df.get(capex_post_cols, pd.DataFrame()).apply(
            pd.to_numeric, errors="coerce"
        )
        capex_df["Total Post-GTM capex (USD/year)"] = capex_post.fillna(0.0).sum(axis=1)

    pools_df = st.session_state.get("shared_capex_pools_table", pd.DataFrame()).copy()
    allocations_df = st.session_state.get("shared_capex_allocations_table", pd.DataFrame()).copy()
    shared_allocations = _build_shared_capex_allocations(dev_df, pools_df, allocations_df)
    if not shared_allocations.empty:
        pool_values = pools_df.copy()
        pool_values["Pool name"] = pool_values.get("Pool name", "").astype(str)
        pool_values["Pre-GTM total (USD)"] = pool_values.get(capex_pre_cols, pd.DataFrame()).apply(
            pd.to_numeric, errors="coerce"
        ).fillna(0.0).sum(axis=1)
        pool_values["Post-GTM total (USD/year)"] = pool_values.get(
            capex_post_cols, pd.DataFrame()
        ).apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)
        shared_totals = shared_allocations.merge(
            pool_values[["Pool name", "Pre-GTM total (USD)", "Post-GTM total (USD/year)"]],
            on="Pool name",
            how="left",
        )
        shared_totals["Shared Pre-GTM capex (USD)"] = (
            shared_totals["Share"] * shared_totals["Pre-GTM total (USD)"].fillna(0.0)
        )
        shared_totals["Shared Post-GTM capex (USD/year)"] = (
            shared_totals["Share"] * shared_totals["Post-GTM total (USD/year)"].fillna(0.0)
        )
        shared_summary = (
            shared_totals.groupby("ID_vaccine", as_index=False)[
                ["Shared Pre-GTM capex (USD)", "Shared Post-GTM capex (USD/year)"]
            ]
            .sum()
        )
        capex_df = capex_df.merge(shared_summary, on="ID_vaccine", how="left")
        capex_df["Shared Pre-GTM capex (USD)"] = capex_df.get(
            "Shared Pre-GTM capex (USD)", pd.Series(0.0, index=capex_df.index)
        ).fillna(0.0)
        capex_df["Shared Post-GTM capex (USD/year)"] = capex_df.get(
            "Shared Post-GTM capex (USD/year)", pd.Series(0.0, index=capex_df.index)
        ).fillna(0.0)
        capex_df["Total Pre-GTM capex (USD)"] = (
            capex_df["Total Pre-GTM capex (USD)"] + capex_df["Shared Pre-GTM capex (USD)"]
        )
        capex_df["Total Post-GTM capex (USD/year)"] = (
            capex_df["Total Post-GTM capex (USD/year)"]
            + capex_df["Shared Post-GTM capex (USD/year)"]
        )

    merged = dev_df.merge(revenue_df, on=["ID_vaccine", "Vaccine name"], how="left")
    merged = merged.merge(cost_df, on=["ID_vaccine", "Vaccine name"], how="left")
    merged = merged.merge(rd_df, on=["ID_vaccine", "Vaccine name"], how="left")
    merged = merged.merge(capex_df, on=["ID_vaccine", "Vaccine name"], how="left")

    inputs = []
    for _, row in merged.iterrows():
        price_candidates = _coerce_numeric(
            pd.Series(
                [
                    row.get("Patent price (USD/customer)"),
                    row.get("Post patent price (USD/customer)"),
                ]
            ),
            0.0,
        )
        unit_price = next((float(value) for value in price_candidates if float(value) > 0.0), 0.0)
        units_per_year = _coerce_numeric(pd.Series([row.get("Patent customers per year")]), 0.0).iloc[0]
        if not unit_price:
            unit_price = _coerce_numeric(pd.Series([row.get("Patent revenue target (USD)")]), 0.0).iloc[0]
            unit_price = unit_price / units_per_year if units_per_year else 0.0
        operating_cost_pct = float(row.get("Patent operating cost %", 0.0) or 0.0) / 100.0
        unit_variable_cost = unit_price * operating_cost_pct
        unit_fixed_cost = float(row.get("G&A total (USD)", 0.0) or 0.0) + float(
            row.get("Post-GTM annual cost (USD/year)", 0.0) or 0.0
        ) + float(row.get("Total Post-GTM capex (USD/year)", 0.0) or 0.0)

        inputs.append(
            {
                "ID_vaccine": row.get("ID_vaccine"),
                "Vaccine name": row.get("Vaccine name"),
                "Unit price (USD)": unit_price,
                "Unit variable cost (USD)": unit_variable_cost,
                "Unit fixed cost (USD/year)": unit_fixed_cost,
                "Units per year": units_per_year,
            }
        )

    return pd.DataFrame(inputs)


def _build_vaccine_break_even_table(
    model_cfg: Optional[ModelConfig],
    *,
    inputs_df: Optional[pd.DataFrame] = None,
    ai_assist: Optional[bool] = None,
    ai_target_years: Optional[int] = None,
) -> pd.DataFrame:
    if model_cfg is None:
        return pd.DataFrame()

    base_inputs = _build_vaccine_break_even_inputs(model_cfg)
    if base_inputs.empty:
        return pd.DataFrame()

    if inputs_df is None:
        stored_inputs = st.session_state.get("vaccine_break_even_inputs")
        if isinstance(stored_inputs, pd.DataFrame) and not stored_inputs.empty:
            inputs_df = stored_inputs
        else:
            inputs_df = base_inputs

    inputs_df = inputs_df.copy()
    if "Vaccine name" in base_inputs.columns:
        inputs_df = inputs_df.merge(
            base_inputs[["Vaccine name", "ID_vaccine"]],
            on="Vaccine name",
            how="left",
            suffixes=("", "_base"),
        )
        if "ID_vaccine_base" in inputs_df.columns:
            inputs_df["ID_vaccine"] = inputs_df["ID_vaccine"].combine_first(inputs_df["ID_vaccine_base"])
            inputs_df = inputs_df.drop(columns=["ID_vaccine_base"], errors="ignore")

    unit_price = _coerce_numeric(inputs_df.get("Unit price (USD)", pd.Series(dtype=float)))
    unit_variable = _coerce_numeric(inputs_df.get("Unit variable cost (USD)", pd.Series(dtype=float)))
    unit_fixed = _coerce_numeric(inputs_df.get("Unit fixed cost (USD/year)", pd.Series(dtype=float)))
    units_per_year = _coerce_numeric(inputs_df.get("Units per year", pd.Series(dtype=float)))

    margin = unit_price - unit_variable
    contribution_pct = np.where(unit_price != 0, margin / unit_price, 0.0)
    break_even_units = np.where(margin > 0, unit_fixed / margin, np.nan)
    break_even_revenue = break_even_units * unit_price
    break_even_unit_cost = np.where(units_per_year > 0, unit_variable + unit_fixed / units_per_year, np.nan)

    results = inputs_df[["ID_vaccine", "Vaccine name"]].copy()
    results["Unit price (USD)"] = unit_price
    results["Unit variable cost (USD)"] = unit_variable
    results["Unit fixed cost (USD/year)"] = unit_fixed
    results["Units per year"] = units_per_year
    results["Unit contribution margin (USD)"] = margin
    results["Contribution margin %"] = contribution_pct
    results["Break-even units"] = break_even_units
    results["Break-even revenue (USD)"] = break_even_revenue
    results["Break-even unit cost (USD)"] = break_even_unit_cost

    if ai_assist is None:
        ai_assist = bool(st.session_state.get("break_even_ai_assist", True))
    if ai_target_years is None:
        ai_target_years = int(st.session_state.get("break_even_ai_target_years", 3))

    if ai_assist:
        required_price = np.where(
            units_per_year > 0,
            unit_variable + unit_fixed / (units_per_year * max(ai_target_years, 1)),
            np.nan,
        )
        results["AI suggested unit price (USD)"] = required_price

    return results


def _evaluate_portfolio_shock(
    portfolio: Portfolio,
    *,
    revenue_multiplier: float = 1.0,
    cost_multiplier: float = 1.0,
    discount_shift: float = 0.0,
    success_prob_multiplier: float = 1.0,
    launch_delay_years: int = 0,
    stage_slippage_years: Optional[Dict[str, int]] = None,
) -> Optional[ValuationResult]:
    """Run a valuation after applying a Scenario-style shock."""

    if portfolio is None:
        return None
    scenario = Scenario(
        name="analytics_scenario",
        revenue_multiplier=revenue_multiplier,
        cost_multiplier=cost_multiplier,
        discount_rate_shift=discount_shift,
        success_prob_multiplier=success_prob_multiplier,
        launch_delay_years=launch_delay_years,
        stage_slippage_years=stage_slippage_years or {},
    )
    scen_engine = ScenarioEngine(portfolio)
    shocked_portfolio = scen_engine._apply_scenario(scenario)
    return ValuationEngine(shocked_portfolio).run()


def _run_sensitivity_matrix(
    portfolio: Portfolio,
    driver_settings: Dict[str, Tuple[float, str]],
) -> pd.DataFrame:
    """Evaluate +/- shocks for each driver and return the resulting rNPVs."""

    rows: List[Dict[str, float]] = []
    if portfolio is None:
        return pd.DataFrame()

    for driver, (delta, driver_type) in driver_settings.items():
        for direction in (-(delta), 0.0, delta):
            rev_mult = 1.0
            cost_mult = 1.0
            if driver_type == "revenue":
                rev_mult = 1.0 + direction
            elif driver_type == "cost":
                cost_mult = 1.0 + direction
            elif driver_type == "productivity":
                rev_mult = 1.0 + direction
                cost_mult = max(0.1, 1.0 - direction / 2)

            result = _evaluate_portfolio_shock(
                portfolio,
                revenue_multiplier=rev_mult,
                cost_multiplier=cost_mult,
            )
            if result is None:
                continue
            rows.append(
                {
                    "Driver": driver,
                    "Change": f"{direction:+.0%}",
                    "rNPV": result.rnpv,
                }
            )

    df = pd.DataFrame(rows)
    if not df.empty:
        df["Delta vs base"] = df.groupby("Driver")["rNPV"].transform(lambda x: x - x.iloc[1])
    return df


def _compute_decomposition(cons: pd.DataFrame) -> Optional[pd.DataFrame]:
    """Run a simple trend/seasonality decomposition on revenue if enough history exists."""

    if len(cons) < 6:
        return None
    series = cons["revenue"].copy()
    idx = pd.PeriodIndex(cons.index, freq="Y").to_timestamp()
    ts = pd.Series(series.values, index=idx)
    period = max(2, min(6, len(ts) // 2))
    try:
        from statsmodels.tsa.seasonal import seasonal_decompose
    except Exception:
        return None

    result = seasonal_decompose(ts, model="additive", period=period, extrapolate_trend="freq")
    return pd.DataFrame(
        {
            "observed": result.observed,
            "trend": result.trend,
            "seasonal": result.seasonal,
            "resid": result.resid,
        }
    )


def _build_segmentation_table(val_result) -> pd.DataFrame:
    rows = []
    if val_result is None:
        return pd.DataFrame()
    per_product = val_result.per_product_prob
    total_rev = sum(df["revenue"].sum() for df in per_product.values()) or 1.0
    for name, df in per_product.items():
        revenue = df["revenue"].sum()
        ebitda = df["ebitda"].sum()
        fcff = df["fcff"].sum()
        margin = ebitda / revenue if revenue else 0.0
        rows.append(
            {
                "Product": name,
                "Revenue share": revenue / total_rev,
                "EBITDA margin": margin,
                "FCFF (PV proxy)": fcff,
            }
        )
    return pd.DataFrame(rows)


def _goal_seek_revenue_multiplier(
    portfolio: Portfolio, target_rnpv: float, tolerance: float = 1e-3, max_iter: int = 20
) -> Tuple[float, Optional[float]]:
    """Binary-search the revenue multiplier needed to hit a target rNPV."""

    if portfolio is None:
        return 1.0, None
    low, high = 0.25, 3.0
    solution = None
    for _ in range(max_iter):
        mid = (low + high) / 2
        result = _evaluate_portfolio_shock(portfolio, revenue_multiplier=mid)
        if result is None:
            break
        diff = result.rnpv - target_rnpv
        if abs(diff) <= tolerance * max(1.0, target_rnpv):
            solution = result.rnpv
            return mid, solution
        if diff < 0:
            low = mid
        else:
            high = mid
    if solution is None:
        result = _evaluate_portfolio_shock(portfolio, revenue_multiplier=high)
        solution = result.rnpv if result else None
    return high, solution


def _tornado_dataframe(portfolio: Portfolio, base_rnpv: float) -> pd.DataFrame:
    """Compute +/- shocks for tornado and spider charts."""

    drivers = [
        ("Revenue", "revenue_multiplier"),
        ("COGS", "cost_multiplier"),
        ("Discount rate", "discount_rate"),
        ("Success probability", "success"),
    ]
    records = []
    for label, driver_type in drivers:
        for change in (-0.2, 0.2):
            kwargs = {
                "revenue_multiplier": 1.0,
                "cost_multiplier": 1.0,
                "discount_shift": 0.0,
                "success_prob_multiplier": 1.0,
            }
            if driver_type == "revenue_multiplier":
                kwargs["revenue_multiplier"] += change
            elif driver_type == "cost_multiplier":
                kwargs["cost_multiplier"] += change
            elif driver_type == "discount_rate":
                kwargs["discount_shift"] = change * 0.5
            else:
                kwargs["success_prob_multiplier"] += change
            result = _evaluate_portfolio_shock(portfolio, **kwargs)
            if result is None:
                continue
            records.append(
                {
                    "Driver": label,
                    "Change": f"{change:+.0%}",
                    "rNPV": result.rnpv,
                    "Delta": result.rnpv - base_rnpv,
                }
            )
    return pd.DataFrame(records)


def _run_linear_regressions(cons: pd.DataFrame) -> Optional[pd.DataFrame]:
    if LinearRegression is None or cons.empty:
        return None
    x = cons[["revenue"]].values
    rows = []
    for target in ["ebitda", "nopat", "fcff_after_wc"]:
        y = cons[target].values
        model = LinearRegression()
        try:
            model.fit(x, y)
        except Exception:
            return None
        rows.append(
            {
                "Target": target.upper(),
                "Intercept": model.intercept_,
                "Revenue beta": model.coef_[0],
                "R^2": model.score(x, y),
            }
        )
    return pd.DataFrame(rows)


def _run_classification_model(seg_df: pd.DataFrame) -> Optional[pd.DataFrame]:
    if LogisticRegression is None or seg_df.empty:
        return None
    df = seg_df.copy()
    df["High margin"] = (df["EBITDA margin"] > 0.3).astype(int)
    X = df[["Revenue share", "EBITDA margin"]].values
    y = df["High margin"].values
    model = LogisticRegression()
    try:
        model.fit(X, y)
    except Exception:
        return None
    probs = model.predict_proba(X)[:, 1]
    df["High-margin probability"] = probs
    return df[["Product", "Revenue share", "EBITDA margin", "High-margin probability"]]


def _optimize_operations(cons: pd.DataFrame) -> Optional[pd.DataFrame]:
    if minimize is None or cons.empty:
        return None

    avg_rev = cons["revenue"].mean()
    avg_cost = (-cons["cogs"].mean()) if not cons["cogs"].empty else 0.0

    def objective(x: np.ndarray) -> float:
        volume, efficiency = x
        revenue = avg_rev * volume * efficiency
        cost = avg_cost * volume * (2 - efficiency)
        return -(revenue - cost)

    cons_list = (
        {"type": "ineq", "fun": lambda x: x[0] - 0.5},
        {"type": "ineq", "fun": lambda x: x[1] - 0.5},
        {"type": "ineq", "fun": lambda x: 2.0 - x[0]},
        {"type": "ineq", "fun": lambda x: 1.5 - x[1]},
    )
    res = minimize(objective, x0=np.array([1.0, 1.0]), constraints=cons_list)
    if not res.success:
        return None
    volume, efficiency = res.x
    opt_profit = -res.fun
    return pd.DataFrame(
        {
            "Metric": ["Optimal volume scale", "Optimal efficiency", "Profit"],
            "Value": [volume, efficiency, opt_profit],
        }
    )


def _mean_variance_portfolio(val_result) -> Optional[pd.DataFrame]:
    if val_result is None:
        return None
    per_product = val_result.per_product_prob
    rows = []
    for name, df in per_product.items():
        returns = df["fcff"].values
        if len(returns) < 2:
            continue
        rows.append(
            {
                "Product": name,
                "Mean": np.mean(returns),
                "Std": np.std(returns),
            }
        )
    if not rows:
        return None
    df = pd.DataFrame(rows)
    inv_var = 1.0 / df["Std"].replace(0, np.nan)
    inv_var = inv_var.fillna(0.0)
    if inv_var.sum() > 0:
        df["Suggested weight"] = inv_var / inv_var.sum()
    else:
        df["Suggested weight"] = 1.0 / len(df)
    return df


def _real_options_value(val_result, volatility: float = 0.35, years: int = 3) -> Optional[float]:
    if val_result is None:
        return None
    underlying = max(val_result.rnpv, 0.0)
    strike = val_result.consolidated["rd_cash"].abs().sum() / years if years else 1.0
    if underlying <= 0 or strike <= 0 or volatility <= 0 or years <= 0:
        return None
    # Black-Scholes call option approximation on project deferral
    from math import log, sqrt
    try:
        from scipy.stats import norm
    except Exception:
        return None

    r = 0.05
    T = max(1e-6, years)
    d1 = (log(underlying / strike) + (r + 0.5 * volatility**2) * T) / (volatility * sqrt(T))
    d2 = d1 - volatility * sqrt(T)
    option_value = underlying * norm.cdf(d1) - strike * np.exp(-r * T) * norm.cdf(d2)
    return option_value


def _copula_simulation(cons: pd.DataFrame, rho: float = 0.4, draws: int = 2000) -> Optional[pd.DataFrame]:
    if cons.empty:
        return None
    mean_vec = np.array([cons["revenue"].mean(), cons["ebitda"].mean()])
    std_vec = np.array([cons["revenue"].std(), cons["ebitda"].std()])
    cov = np.array([[1.0, rho], [rho, 1.0]])
    samples = np.random.multivariate_normal([0, 0], cov, size=draws)
    revenue_sim = mean_vec[0] + std_vec[0] * samples[:, 0]
    ebitda_sim = mean_vec[1] + std_vec[1] * samples[:, 1]
    return pd.DataFrame({"Revenue": revenue_sim, "EBITDA": ebitda_sim})


def _cluster_products(val_result) -> Optional[pd.DataFrame]:
    if val_result is None or KMeans is None:
        return None
    per_product = val_result.per_product_prob
    rows = []
    for name, df in per_product.items():
        revenue = df["revenue"].sum()
        ebitda = df["ebitda"].sum()
        growth = df["revenue"].pct_change().mean()
        rows.append([name, revenue, ebitda, growth if pd.notna(growth) else 0.0])
    if not rows:
        return None
    names, data = zip(*[(r[0], r[1:]) for r in rows])
    scaler = StandardScaler() if StandardScaler else None
    matrix = np.array(data, dtype=float)
    matrix = np.nan_to_num(matrix, nan=0.0, posinf=0.0, neginf=0.0)
    if scaler is not None:
        matrix = scaler.fit_transform(matrix)
    n_clusters = min(3, len(matrix))
    if n_clusters < 1:
        return None
    km = KMeans(n_clusters=n_clusters, n_init=10, random_state=42)
    labels = km.fit_predict(matrix)
    return pd.DataFrame({"Product": names, "Cluster": labels})


def _machine_learning_multiple(cons: pd.DataFrame) -> Optional[pd.DataFrame]:
    if LinearRegression is None or cons.empty or "revenue" not in cons.columns or "ebitda" not in cons.columns:
        return None
    revenue = pd.to_numeric(cons["revenue"], errors="coerce")
    ebitda = pd.to_numeric(cons["ebitda"], errors="coerce")
    if revenue.isna().all() or ebitda.isna().all():
        return None
    growth = revenue.pct_change().replace([np.inf, -np.inf], np.nan).fillna(0.0)
    features = pd.DataFrame(
        {
            "Revenue": revenue,
            "EBITDA": ebitda,
            "Growth": growth,
        }
    )
    features = features.replace([np.inf, -np.inf], np.nan).fillna(0.0)
    multiples = ((ebitda.rolling(3, min_periods=1).mean().bfill()) + 1.0) / 1_000_000
    if multiples.isna().all():
        return None
    model = LinearRegression()
    model.fit(features.to_numpy(dtype=float), multiples.to_numpy(dtype=float))
    pred = model.predict(features.to_numpy(dtype=float))
    return pd.DataFrame({"Year": cons.index, "Predicted multiple": pred})


def _compute_irr(cashflows: List[float]) -> Optional[float]:
    if not cashflows or all(cf >= 0 for cf in cashflows) or all(cf <= 0 for cf in cashflows):
        return None

    def npv(rate: float) -> float:
        return sum(cf / ((1 + rate) ** idx) for idx, cf in enumerate(cashflows))

    low, high = -0.9, 1.0
    npv_low, npv_high = npv(low), npv(high)
    attempts = 0
    while npv_low * npv_high > 0 and attempts < 10:
        high += 1.0
        npv_high = npv(high)
        attempts += 1
    if npv_low * npv_high > 0:
        return None

    for _ in range(60):
        mid = (low + high) / 2
        npv_mid = npv(mid)
        if abs(npv_mid) < 1e-6:
            return mid
        if npv_low * npv_mid <= 0:
            high = mid
            npv_high = npv_mid
        else:
            low = mid
            npv_low = npv_mid
    return (low + high) / 2


def _compute_payback_years(years: List[int], cashflows: List[float]) -> Optional[float]:
    if not years or not cashflows or len(years) != len(cashflows):
        return None
    cumulative = 0.0
    for idx, (year, cf) in enumerate(zip(years, cashflows)):
        prev_cumulative = cumulative
        cumulative += cf
        if cumulative >= 0 and idx > 0:
            prev_year = years[idx - 1]
            if cf == 0:
                return float(year - years[0])
            fraction = (0 - prev_cumulative) / cf
            return (prev_year + fraction * (year - prev_year)) - years[0]
    return None


def _build_snapshot_from_result(
    model_cfg: ModelConfig,
    valuation_result: ValuationResult,
    scenarios: Optional[List[dict]] = None,
    sensitivities: Optional[List[dict]] = None,
) -> dict:
    cons = valuation_result.consolidated
    dcf = valuation_result.dcf_table
    cashflows = dcf["fcff"].tolist()
    if "terminal_value" in dcf.columns:
        cashflows[-1] += float(dcf["terminal_value"].fillna(0.0).iloc[-1])
    if "working_capital_recovery" in dcf.columns:
        cashflows[-1] += float(dcf["working_capital_recovery"].fillna(0.0).iloc[-1])
    irr = _compute_irr(cashflows)
    payback = _compute_payback_years(cons.index.tolist(), cashflows)
    capex_total = -float(cons["capex_cash"].sum()) if "capex_cash" in cons.columns else None
    opex_components = [
        "sales_marketing",
        "gna",
        "royalty",
        "rd_cash",
    ]
    opex_available = [col for col in opex_components if col in cons.columns]
    opex_annual = None
    if opex_available:
        opex_annual = -float(cons[opex_available].sum(axis=1).mean())
    revenue_annual = float(cons["revenue"].mean()) if "revenue" in cons.columns else None
    dscr_min = None
    financing_outputs = _build_financing_outputs(valuation_result, model_cfg)
    lender_metrics = financing_outputs.get("lender_metrics", pd.DataFrame())
    if isinstance(lender_metrics, pd.DataFrame) and not lender_metrics.empty and "DSCR" in lender_metrics.columns:
        finite_dscr = (
            pd.to_numeric(lender_metrics["DSCR"], errors="coerce")
            .replace([np.inf, -np.inf], np.nan)
            .dropna()
        )
        if not finite_dscr.empty:
            dscr_min = float(finite_dscr.min())
    financing_settings = _financing_settings_from_state()
    snapshot = {
        "currency": model_cfg.currency,
        "npv": valuation_result.enterprise_value,
        "irr": irr,
        "dscr_min": dscr_min,
        "payback_years": payback,
        "capex_total": capex_total,
        "opex_annual": opex_annual,
        "revenue_annual": revenue_annual,
        "scenarios": scenarios or [],
        "sensitivities": sensitivities or [],
        "assumptions": {
            "discount_rate": model_cfg.discount_rate,
            "discount_timing": getattr(model_cfg, "discount_timing", "year_end"),
            "tax_rate": model_cfg.tax_rate,
            "working_capital_pct": model_cfg.working_capital_pct_sales,
            "terminal_method": getattr(model_cfg, "terminal_method", "exit_multiple"),
            "perpetuity_growth_rate": getattr(model_cfg, "perpetuity_growth_rate", None),
            "opening_nol_balance": getattr(model_cfg, "opening_nol_balance", 0.0),
            "debt_interest_rate": financing_settings["interest_rate"],
            "debt_repayment_mode": financing_settings["repayment_mode"],
            "debt_target_dscr": financing_settings["target_dscr"],
            "minimum_cash_reserve": financing_settings["minimum_cash_reserve"],
        },
    }
    return snapshot


def _empty_financial_snapshot(currency: str = "USD") -> dict:
    return {
        "currency": currency,
        "npv": None,
        "irr": None,
        "dscr_min": None,
        "payback_years": None,
        "capex_total": None,
        "opex_annual": None,
        "revenue_annual": None,
        "scenarios": [],
        "sensitivities": [],
        "assumptions": {},
    }


def _default_rag_advisory_inputs() -> dict:
    return {
        "scenarios": [],
        "notes": "",
        "workbook_hash": None,
    }


def _build_bankable_snapshot_payload(
    project_id: str,
    model_cfg: Optional[ModelConfig],
    valuation_result: Optional[ValuationResult],
    portfolio: Optional[Portfolio],
    *,
    workbook_hash: Optional[str] = None,
    advisory_inputs: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    snapshot_source = "live_model_only" if model_cfg is not None and valuation_result is not None else "unavailable"
    locked_snapshot = _empty_financial_snapshot(
        currency=getattr(model_cfg, "currency", "USD") if model_cfg is not None else "USD"
    )
    if snapshot_source == "live_model_only":
        locked_snapshot = _build_snapshot_from_result(
            model_cfg,
            valuation_result,
            scenarios=_default_scenario_pack(portfolio),
        )
    return {
        "project_id": project_id,
        "financial_snapshot": locked_snapshot,
        "workbook_hash": workbook_hash,
        "snapshot_source": snapshot_source,
        "advisory_inputs": dict(advisory_inputs or {}),
    }


def _default_scenario_pack(portfolio: Optional[Portfolio]) -> List[dict]:
    if portfolio is None:
        return []
    base = ValuationEngine(portfolio).run()
    upside = _evaluate_portfolio_shock(
        portfolio,
        revenue_multiplier=1.15,
        cost_multiplier=0.95,
        discount_shift=-0.01,
        success_prob_multiplier=1.1,
    )
    downside = _evaluate_portfolio_shock(
        portfolio,
        revenue_multiplier=0.85,
        cost_multiplier=1.05,
        discount_shift=0.02,
        success_prob_multiplier=0.9,
    )
    scenarios = [
        {"name": "Base", "npv": base.rnpv, "irr": None},
    ]
    if upside is not None:
        scenarios.append({"name": "Upside", "npv": upside.rnpv, "irr": None})
    if downside is not None:
        scenarios.append({"name": "Downside", "npv": downside.rnpv, "irr": None})
    return scenarios


def _format_excel_sheet(ws, df: pd.DataFrame, *, freeze_panes: str = "B2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    max_row = df.shape[0] + 1
    max_col = df.shape[1] + 1

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = freeze_panes
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for col_idx, col_name in enumerate(df.columns, start=2):
        col_letter = get_column_letter(col_idx)
        if any(token in col_name.lower() for token in ("pct", "margin", "prob", "%")):
            number_format = "0.0%"
        else:
            number_format = "#,##0.00"
        for row in range(2, max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = number_format

        values = [str(col_name)]
        for row in range(2, max_row + 1):
            values.append(str(ws.cell(row=row, column=col_idx).value or ""))
        width = min(max(len(v) for v in values) + 2, 40)
        ws.column_dimensions[col_letter].width = width

    index_letter = get_column_letter(1)
    index_values = [str(df.index.name or "")] + [str(v) for v in df.index]
    ws.column_dimensions[index_letter].width = min(max(len(v) for v in index_values) + 2, 26)


def _format_excel_table(
    ws,
    df: pd.DataFrame,
    *,
    start_row: int,
    start_col: int = 1,
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    max_row = start_row + df.shape[0]
    max_col = start_col + df.shape[1] - 1

    for col_idx in range(start_col, max_col + 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for offset, col_name in enumerate(df.columns):
        col_idx = start_col + offset
        col_letter = get_column_letter(col_idx)
        if any(token in str(col_name).lower() for token in ("pct", "margin", "prob", "%")):
            number_format = "0.0%"
        else:
            number_format = "#,##0.00"
        for row in range(start_row + 1, max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = number_format

        values = [str(col_name)]
        for row in range(start_row + 1, max_row + 1):
            values.append(str(ws.cell(row=row, column=col_idx).value or ""))
        width = min(max(len(v) for v in values) + 2, 40)
        ws.column_dimensions[col_letter].width = width


def _add_line_chart(
    ws,
    *,
    title: str,
    data_min_col: int,
    data_max_col: int,
    data_max_row: int,
    category_col: int = 1,
    anchor: str = "H2",
) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Year"
    data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=1, max_row=data_max_row)
    categories = Reference(ws, min_col=category_col, min_row=2, max_row=data_max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, anchor)


def _add_bar_chart(
    ws,
    *,
    title: str,
    data_min_col: int,
    data_max_col: int,
    data_max_row: int,
    category_col: int = 1,
    anchor: str = "H2",
) -> None:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Scenario"
    data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=1, max_row=data_max_row)
    categories = Reference(ws, min_col=category_col, min_row=2, max_row=data_max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 12
    chart.width = 22
    ws.add_chart(chart, anchor)


def _build_financial_excel(
    cons: pd.DataFrame,
    perf_df: pd.DataFrame,
    position_df: pd.DataFrame,
    cash_flow_df: pd.DataFrame,
    model_cfg: Optional[ModelConfig] = None,
    lender_metrics: Optional[pd.DataFrame] = None,
    investor_waterfall: Optional[pd.DataFrame] = None,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        cons.to_excel(writer, sheet_name="Consolidated forecast")
        perf_df.to_excel(writer, sheet_name="Financial performance")
        position_df.to_excel(writer, sheet_name="Financial position")
        cash_flow_df.to_excel(writer, sheet_name="Cash flows")
        if lender_metrics is not None and not lender_metrics.empty:
            lender_metrics.to_excel(writer, sheet_name="Debt metrics")
        if investor_waterfall is not None and not investor_waterfall.empty:
            investor_waterfall.to_excel(writer, sheet_name="Investor waterfall", index=False)
        dashboard_cols = [col for col in ["revenue", "ebitda", "fcff_after_wc"] if col in cons.columns]
        dashboard_df = cons[dashboard_cols].copy()
        dashboard_df.to_excel(writer, sheet_name="Dashboard")

        analytics_df = _build_ratio_table(cons)
        if not analytics_df.empty:
            analytics_df.to_excel(writer, sheet_name="Advanced analytics")
        break_even_df = _build_vaccine_break_even_table(model_cfg)
        break_even_start_row = None
        if not break_even_df.empty:
            if analytics_df.empty:
                break_even_start_row = 1
                break_even_df.to_excel(writer, sheet_name="Advanced analytics", index=False)
            else:
                start_row = analytics_df.shape[0] + 3
                break_even_start_row = start_row + 1
                break_even_df.to_excel(
                    writer,
                    sheet_name="Advanced analytics",
                    startrow=start_row,
                    index=False,
                )

        scenario_cols = [col for col in ["revenue", "ebitda", "fcff_after_wc"] if col in cons.columns]
        scenario_df = pd.DataFrame()
        if scenario_cols:
            last_year = cons.index[-1]
            base_values = cons.loc[last_year, scenario_cols]
            scenario_df = pd.DataFrame(
                {
                    "Downside (-10%)": base_values * 0.9,
                    "Base": base_values,
                    "Upside (+10%)": base_values * 1.1,
                }
            ).T
            scenario_df.to_excel(writer, sheet_name="Scenario analysis")

        workbook = writer.book
        for name, df in {
            "Consolidated forecast": cons,
            "Financial performance": perf_df,
            "Financial position": position_df,
            "Cash flows": cash_flow_df,
            "Dashboard": dashboard_df,
        }.items():
            ws = workbook[name]
            _format_excel_sheet(ws, df)
        if lender_metrics is not None and not lender_metrics.empty:
            _format_excel_sheet(workbook["Debt metrics"], lender_metrics)
        if investor_waterfall is not None and not investor_waterfall.empty:
            _format_excel_table(workbook["Investor waterfall"], investor_waterfall, start_row=1)

        if not analytics_df.empty:
            ws = workbook["Advanced analytics"]
            _format_excel_sheet(ws, analytics_df)
        if break_even_start_row is not None:
            ws = workbook["Advanced analytics"]
            title_row = break_even_start_row - 1
            if title_row > 1:
                ws.cell(row=title_row, column=1).value = "Vaccine break-even analysis"
                ws.cell(row=title_row, column=1).font = Font(bold=True)
            _format_excel_table(ws, break_even_df, start_row=break_even_start_row)

        if not scenario_df.empty:
            ws = workbook["Scenario analysis"]
            _format_excel_sheet(ws, scenario_df)

        if not dashboard_df.empty:
            ws = workbook["Dashboard"]
            max_row = dashboard_df.shape[0] + 1
            _add_line_chart(
                ws,
                title="Key Metrics",
                data_min_col=2,
                data_max_col=1 + dashboard_df.shape[1],
                data_max_row=max_row,
            )

        if not analytics_df.empty:
            ws = workbook["Advanced analytics"]
            max_row = analytics_df.shape[0] + 1
            _add_line_chart(
                ws,
                title="Margin Trends",
                data_min_col=2,
                data_max_col=1 + analytics_df.shape[1],
                data_max_row=max_row,
                anchor="H2",
            )

        if lender_metrics is not None and not lender_metrics.empty and "DSCR" in lender_metrics.columns:
            ws = workbook["Debt metrics"]
            max_row = lender_metrics.shape[0] + 1
            _add_line_chart(
                ws,
                title="Debt Coverage",
                data_min_col=2,
                data_max_col=min(4, 1 + lender_metrics.shape[1]),
                data_max_row=max_row,
                anchor="J2",
            )

        if not scenario_df.empty:
            ws = workbook["Scenario analysis"]
            max_row = scenario_df.shape[0] + 1
            _add_bar_chart(
                ws,
                title="Scenario Comparison",
                data_min_col=2,
                data_max_col=1 + scenario_df.shape[1],
                data_max_row=max_row,
                anchor="H2",
            )
    return output.getvalue()


def _rag_section_outline() -> List[str]:
    return [
        "Executive Summary",
        "Project Description & Scope",
        "Market & Demand Analysis",
        "Technical & Operations",
        "Legal, Permitting & Environmental",
        "Implementation Plan",
        "Financial Analysis",
        "Risk Assessment & Mitigations",
        "Conclusion & Recommendation",
        "Appendices",
    ]


def _rag_blueprint_markdown() -> str:
    return (
        "# RAG Feasibility Study Generator\n"
        "\n"
        "A production-ready blueprint (plus reference code) for a Retrieval-Augmented Generation (RAG) "
        "system that ingests up to 1 GB of project materials and automatically drafts a comprehensive "
        "feasibility study grounded in your financial model outputs and accompanying documents.\n"
        "\n"
        "## 0) RAC: Model-Integrated Design (RAG inside the Financial Model)\n"
        "- **What changed**: The Excel workbook is the system of record and orchestrator. The "
        "Retrieval–Aggregation–Composer (RAC) service is triggered from the model to collect results "
        "directly from defined cells/ranges and to ingest up to 1 GB of external evidence.\n"
        "- **Why this pattern**: Single source of truth, fewer manual steps, and repeatable runs tied to "
        "workbook hash + timestamp.\n"
        "\n"
        "## 1) High-level Architecture\n"
        "1. Upload & Ingest: stream large files to disk, parse text, chunk, embed, store in FAISS.\n"
        "2. Financial Model Extraction: load Excel and extract standardized metrics/tables.\n"
        "3. Retrieval: dense + reranker, optional hybrid.\n"
        "4. Planning & Generation: section-by-section prompts grounded by snapshot + retrieved passages.\n"
        "5. Audit: attach provenance and snapshot metadata for reproducibility.\n"
        "\n"
        "## 2) Data Model & Financial Schema\n"
        "Store project artifacts under `projects/<project_id>/` with uploads, parsed text, index, and a "
        "financial snapshot JSON. Snapshot keys include NPV/IRR/DSCR, capex/opex, scenarios, and "
        "sensitivities.\n"
        "\n"
        "## 3) Prompt Strategy & Section Templates\n"
        "Use a strict system prompt that forbids unsupported claims and enforces inline citations. "
        "Each section receives the financial snapshot and top-k contextual passages.\n"
        "\n"
        "## 4) Reference Implementation (FastAPI + FAISS + Sentence-Transformers)\n"
        "The API exposes `/collect`, `/ingest`, and `/generate` endpoints. `/ingest` streams large "
        "uploads, `/collect` stores a validated snapshot, and `/generate` composes the feasibility "
        "study.\n"
        "\n"
        "## 5) Quality, Auditing & Reproducibility\n"
        "- Enforce citations and reject unsupported claims.\n"
        "- Record workbook hash + timestamp.\n"
        "- Run numeric sanity checks (IRR bounds, DSCR thresholds).\n"
        "\n"
        "## 6) Deployment Notes (1 GB uploads)\n"
        "- Stream uploads to disk; avoid in-memory buffers.\n"
        "- Use Nginx `client_max_body_size 1024m` and disable proxy buffering.\n"
        "- Run uvicorn with multiple workers and fast local storage.\n"
        "\n"
        "## 7) Section-specific Retrieval Queries\n"
        "- Executive Summary: decision drivers, showstoppers\n"
        "- Market: market size, demand forecast, price assumptions\n"
        "- Technical: process design, throughput, yield\n"
        "- Legal/Env: permits, EIA/ESIA, land rights\n"
        "- Implementation: schedule, capex phasing\n"
        "- Financial: NPV, IRR, DSCR, sensitivities\n"
        "- Risk/ESG: risk register, mitigations\n"
        "\n"
        "## 8) Appendices & Outputs\n"
        "Include the financial snapshot, sensitivity matrices, scenarios, and an audit trail mapping "
        "sources to citations.\n"
    )


def _build_ai_commentary(
    snapshot_summary: Dict[str, Any],
    perf_df: Optional[pd.DataFrame],
    position_df: Optional[pd.DataFrame],
    cash_flow_df: Optional[pd.DataFrame],
    cons_df: Optional[pd.DataFrame],
    analytics_df: Optional[pd.DataFrame] = None,
) -> List[Dict[str, str]]:
    comments: List[Dict[str, str]] = []

    def _add_comment(section: str, commentary: str, annotation: str = "") -> None:
        comments.append(
            {
                "Section": section,
                "Commentary": commentary,
                "Annotation": annotation,
            }
        )

    def _format_value(value: Any) -> str:
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return "n/a"
        if isinstance(value, (int, float)):
            return f"{value:,.0f}"
        return str(value)

    def _format_pct(value: Any) -> str:
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return "n/a"
        if isinstance(value, (int, float)):
            pct_value = value * 100 if abs(value) <= 1.5 else value
            return f"{pct_value:.1f}%"
        return str(value)

    def _safe_divide(numerator: float, denominator: float) -> float:
        return float(numerator / denominator) if denominator else 0.0

    def _first_last(series: pd.Series) -> tuple[Optional[float], Optional[float]]:
        if series is None or series.empty:
            return None, None
        clean = pd.to_numeric(series, errors="coerce").dropna()
        if clean.empty:
            return None, None
        return float(clean.iloc[0]), float(clean.iloc[-1])

    currency = snapshot_summary.get("currency", "USD") if snapshot_summary else "USD"

    if snapshot_summary:
        npv = snapshot_summary.get("npv")
        irr = snapshot_summary.get("irr")
        dscr_min = snapshot_summary.get("dscr_min")
        payback_years = snapshot_summary.get("payback_years")
        revenue = snapshot_summary.get("revenue_annual")
        opex = snapshot_summary.get("opex_annual")
        capex_total = snapshot_summary.get("capex_total")
        _add_comment(
            "Financial Snapshot",
            (
                f"NPV {_format_value(npv)} {currency}, IRR {_format_pct(irr)}, "
                f"minimum DSCR {_format_value(dscr_min)}, payback {_format_value(payback_years)} years."
            ),
            "Snapshot metrics come from the RAG Assistant inputs.",
        )
        _add_comment(
            "Financial Snapshot",
            (
                f"Annual revenue {_format_value(revenue)} {currency}, annual opex {_format_value(opex)} "
                f"{currency}, total capex {_format_value(capex_total)} {currency}."
            ),
            "Operating spread = annual revenue minus annual opex.",
        )
        if revenue is not None and opex is not None:
            _add_comment(
                "Financial Snapshot",
                f"Estimated operating spread: {_format_value(revenue - opex)} {currency}.",
                "Positive spread indicates operating headroom before financing effects.",
            )
        if npv is not None:
            _add_comment(
                "Financial Snapshot",
                f"NPV implies a {'positive' if npv >= 0 else 'negative'} valuation trend.",
                "NPV sign provides directional value signal.",
            )

    if perf_df is not None and not perf_df.empty:
        revenue_series = perf_df.get("Revenue")
        ebitda_series = perf_df.get("EBITDA")
        cogs_series = perf_df.get("COGS")
        rd_series = perf_df.get("R&D expense")
        if revenue_series is not None:
            avg_revenue = float(revenue_series.mean())
            _add_comment(
                "Statement of Financial Performance",
                f"Average revenue across the plan is {avg_revenue:,.0f} {currency}.",
                "Average computed across modeled forecast years.",
            )
            rev_start, rev_end = _first_last(revenue_series)
            if rev_start is not None and rev_end is not None and rev_start > 0:
                years = max(1, len(revenue_series) - 1)
                cagr = (rev_end / rev_start) ** (1 / years) - 1
                _add_comment(
                    "Statement of Financial Performance",
                    f"Revenue grows from {rev_start:,.0f} to {rev_end:,.0f} {currency} (CAGR {_format_pct(cagr)}).",
                    "CAGR uses first and last modeled revenue values.",
                )
        if revenue_series is not None and ebitda_series is not None:
            total_revenue = float(revenue_series.sum())
            total_ebitda = float(ebitda_series.sum())
            margin = _safe_divide(total_ebitda, total_revenue)
            positive_years = int((ebitda_series > 0).sum())
            _add_comment(
                "Statement of Financial Performance",
                f"Average EBITDA margin is {_format_pct(margin)} with EBITDA positive in {positive_years} year(s).",
                "EBITDA margin = total EBITDA / total revenue.",
            )
            start_margin = _safe_divide(float(ebitda_series.iloc[0]), float(revenue_series.iloc[0]))
            end_margin = _safe_divide(float(ebitda_series.iloc[-1]), float(revenue_series.iloc[-1]))
            _add_comment(
                "Statement of Financial Performance",
                f"EBITDA margin shifts from {_format_pct(start_margin)} to {_format_pct(end_margin)}.",
                "Margin trend compares first and last modeled years.",
            )
        if revenue_series is not None and cogs_series is not None:
            gross_margin = _safe_divide(
                float((revenue_series - cogs_series).sum()),
                float(revenue_series.sum()),
            )
            _add_comment(
                "Statement of Financial Performance",
                f"Average gross margin is {_format_pct(gross_margin)}.",
                "Gross margin = (Revenue - COGS) / Revenue.",
            )
        if revenue_series is not None and rd_series is not None:
            rd_intensity = _safe_divide(float(rd_series.sum()), float(revenue_series.sum()))
            _add_comment(
                "Statement of Financial Performance",
                f"R&D intensity averages {_format_pct(rd_intensity)} of revenue.",
                "R&D intensity = total R&D expense / total revenue.",
            )

    if position_df is not None and not position_df.empty:
        total_assets = position_df.get("Total assets")
        total_equity = position_df.get("Total equity")
        working_capital = position_df.get("Working capital")
        if total_assets is not None:
            end_assets = float(total_assets.iloc[-1])
            _add_comment(
                "Statement of Financial Position",
                f"Ending total assets are {end_assets:,.0f} {currency}.",
                "Ending balances reflect the final forecast year.",
            )
        if total_equity is not None and total_assets is not None:
            end_equity = float(total_equity.iloc[-1])
            equity_ratio = _safe_divide(end_equity, float(total_assets.iloc[-1]))
            _add_comment(
                "Statement of Financial Position",
                f"Ending total equity is {end_equity:,.0f} {currency} (equity ratio {_format_pct(equity_ratio)}).",
                "Equity ratio = total equity / total assets.",
            )
        if working_capital is not None:
            end_wc = float(working_capital.iloc[-1])
            _add_comment(
                "Statement of Financial Position",
                f"Working capital ends at {end_wc:,.0f} {currency}.",
                "Working capital derived from model working capital % assumption.",
            )

    if cash_flow_df is not None and not cash_flow_df.empty:
        net_cash = cash_flow_df.get("Net change in cash")
        cash_ops = cash_flow_df.get("Cash from operations")
        cash_investing = cash_flow_df.get("Cash from investing")
        if net_cash is not None:
            cumulative_cash = float(net_cash.sum())
            positive_years = int((net_cash > 0).sum())
            _add_comment(
                "Statement of Cash Flows",
                f"Cumulative net cash change is {cumulative_cash:,.0f} {currency} with {positive_years} positive year(s).",
                "Net change in cash aggregates operating, investing, and financing flows.",
            )
        if cash_ops is not None:
            avg_ops = float(cash_ops.mean())
            _add_comment(
                "Statement of Cash Flows",
                f"Average operating cash flow is {avg_ops:,.0f} {currency}.",
                "Operating cash flow = NOPAT + depreciation/amortization - working capital change.",
            )
        if cash_investing is not None:
            total_investing = float(cash_investing.sum())
            _add_comment(
                "Statement of Cash Flows",
                f"Total investing cash flow is {total_investing:,.0f} {currency}.",
                "Investing cash flow reflects capex and R&D capitalization.",
            )
        if cash_ops is not None and cash_investing is not None:
            coverage = _safe_divide(float(cash_ops.sum()), abs(float(cash_investing.sum())))
            _add_comment(
                "Statement of Cash Flows",
                f"Operating cash flow covers investing outflows at {_format_pct(coverage)}.",
                "Coverage ratio = total operating cash flow / absolute investing cash flow.",
            )

    if cons_df is not None and not cons_df.empty:
        peak_revenue = float(cons_df["revenue"].max()) if "revenue" in cons_df.columns else None
        total_fcff = float(cons_df["fcff_after_wc"].sum()) if "fcff_after_wc" in cons_df.columns else None
        if peak_revenue is not None:
            _add_comment(
                "Financial Statements Highlights",
                f"Peak revenue reaches {peak_revenue:,.0f} {currency}.",
                "Peak derived from consolidated revenue series.",
            )
        if total_fcff is not None:
            positive_fcff_years = int((cons_df["fcff_after_wc"] > 0).sum())
            _add_comment(
                "Financial Statements Highlights",
                (
                    f"Total FCFF after working capital sums to {total_fcff:,.0f} {currency} "
                    f"with {positive_fcff_years} positive year(s)."
                ),
                "FCFF after WC = free cash flow after working capital change.",
            )

    if analytics_df is not None and not analytics_df.empty:
        narrative = _build_advanced_analytics_narrative(analytics_df)
        for paragraph in narrative:
            _add_comment(
                "Advanced Analytics Narrative",
                paragraph,
                "Derived from the advanced analytics ratio table.",
            )

    scenarios = snapshot_summary.get("scenarios") if snapshot_summary else []
    if scenarios:
        scenario_name = lambda s: s.get("name") or s.get("scenario") or "Scenario"
        scenario_metric = lambda s, key: s.get(key) if isinstance(s, dict) else None
        valid_npvs = [(scenario_name(s), scenario_metric(s, "npv")) for s in scenarios]
        valid_npvs = [(name, value) for name, value in valid_npvs if value is not None]
        if valid_npvs:
            best = max(valid_npvs, key=lambda item: item[1])
            worst = min(valid_npvs, key=lambda item: item[1])
            _add_comment(
                "Scenario Review",
                f"Scenario count {len(scenarios)}; best NPV is {best[0]} at {_format_value(best[1])} {currency}.",
                "Scenario ranking based on reported NPV values.",
            )
            _add_comment(
                "Scenario Review",
                f"Lowest NPV scenario is {worst[0]} at {_format_value(worst[1])} {currency}.",
                "Use scenario deltas to quantify downside exposure.",
            )

    sensitivities = snapshot_summary.get("sensitivities") if snapshot_summary else []
    if sensitivities:
        drivers = []
        for sensitivity in sensitivities:
            if isinstance(sensitivity, dict):
                drivers.append(sensitivity.get("name") or sensitivity.get("driver"))
        drivers = [driver for driver in drivers if driver]
        if drivers:
            _add_comment(
                "Sensitivity Review",
                f"Key sensitivity drivers captured: {', '.join(drivers)}.",
                "Sensitivity drivers sourced from the snapshot table.",
            )

    coverage_notes = []
    if perf_df is not None and not perf_df.empty:
        coverage_notes.append("financial performance")
    if position_df is not None and not position_df.empty:
        coverage_notes.append("financial position")
    if cash_flow_df is not None and not cash_flow_df.empty:
        coverage_notes.append("cash flows")
    if cons_df is not None and not cons_df.empty:
        coverage_notes.append("consolidated statements")
    if coverage_notes:
        _add_comment(
            "Data Coverage",
            f"Report includes {', '.join(coverage_notes)} aligned with the current forecast horizon.",
            "Coverage ensures the business plan narrative reflects model outputs.",
        )

    if not comments:
        _add_comment(
            "Data Coverage",
            "Insufficient data to generate AI commentary. Populate snapshot and financial statements first.",
            "Provide model results to enable narrative generation.",
        )
    return comments


def _group_ai_commentary(ai_commentary: List[Any]) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = {}
    for entry in ai_commentary or []:
        if isinstance(entry, dict):
            section = entry.get("Section", "General")
            grouped.setdefault(section, []).append(entry)
        else:
            grouped.setdefault("General", []).append(
                {"Section": "General", "Commentary": str(entry), "Annotation": ""}
            )
    return grouped


def _format_scenario_prose(scenario: Dict[str, Any], currency: str) -> str:
    name = scenario.get("name") or scenario.get("scenario") or "Scenario"
    npv = scenario.get("npv")
    irr = scenario.get("irr")
    npv_text = f"{npv:,.0f} {currency}" if isinstance(npv, (int, float)) else "n/a"
    irr_text = f"{irr:.1%}" if isinstance(irr, (int, float)) else "n/a"
    return f"{name}: NPV {npv_text}, IRR {irr_text}."


def _format_pct_value(value: Any) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return "n/a"
    if isinstance(value, (int, float)):
        pct_value = value * 100 if abs(value) <= 1.5 else value
        return f"{pct_value:.1f}%"
    return str(value)


def _build_advanced_analytics_narrative(
    analytics_df: Optional[pd.DataFrame],
) -> List[str]:
    if analytics_df is None or analytics_df.empty:
        return []

    clean_df = analytics_df.apply(pd.to_numeric, errors="coerce")
    years = clean_df.index.tolist()
    narrative: List[str] = []

    def _first_last(series: pd.Series) -> tuple[Optional[float], Optional[float]]:
        clean = pd.to_numeric(series, errors="coerce").dropna()
        if clean.empty:
            return None, None
        return float(clean.iloc[0]), float(clean.iloc[-1])

    def _trend_sentence(label: str, series: pd.Series) -> Optional[str]:
        start, end = _first_last(series)
        if start is None or end is None:
            return None
        direction = "improves" if end >= start else "declines"
        return (
            f"{label} {direction} from {_format_pct_value(start)} to "
            f"{_format_pct_value(end)} over the forecast horizon."
        )

    gross_margin = clean_df.get("Gross margin")
    ebitda_margin = clean_df.get("EBITDA margin")
    nopat_margin = clean_df.get("NOPAT margin")
    rd_intensity = clean_df.get("R&D intensity")
    capex_intensity = clean_df.get("Capex intensity")

    if years:
        narrative.append(
            f"The advanced analytics ratios cover {years[0]} through {years[-1]}, "
            "highlighting profitability, efficiency, and reinvestment trends."
        )

    for label, series in [
        ("Gross margin", gross_margin),
        ("EBITDA margin", ebitda_margin),
        ("NOPAT margin", nopat_margin),
    ]:
        if series is not None:
            sentence = _trend_sentence(label, series)
            if sentence:
                narrative.append(sentence)

    if rd_intensity is not None:
        start, end = _first_last(rd_intensity)
        if start is not None and end is not None:
            narrative.append(
                "R&D intensity moderates from "
                f"{_format_pct_value(start)} to {_format_pct_value(end)}, "
                "indicating a tapering of development spend as commercialization matures."
            )
    if capex_intensity is not None:
        start, end = _first_last(capex_intensity)
        if start is not None and end is not None:
            narrative.append(
                "Capex intensity steps down from "
                f"{_format_pct_value(start)} to {_format_pct_value(end)}, "
                "suggesting upfront build-out gives way to steadier maintenance investment."
            )

    if gross_margin is not None and ebitda_margin is not None and nopat_margin is not None:
        peak_year = clean_df[["Gross margin", "EBITDA margin", "NOPAT margin"]].mean(axis=1).idxmax()
        peak_row = clean_df.loc[peak_year]
        narrative.append(
            "Peak profitability occurs around "
            f"{peak_year}, with gross margin {_format_pct_value(peak_row.get('Gross margin'))}, "
            f"EBITDA margin {_format_pct_value(peak_row.get('EBITDA margin'))}, "
            f"and NOPAT margin {_format_pct_value(peak_row.get('NOPAT margin'))}."
        )

    return narrative


def _build_extended_analytics_sections(chart_tables: Dict[str, pd.DataFrame]) -> List[Dict[str, str]]:
    sections: List[Dict[str, str]] = []

    def _add_section(title: str, status: str, details: str) -> None:
        sections.append({"Section": title, "Status": status, "Details": details})

    _add_section(
        "Margin & intensity analysis",
        "Included",
        "Summarizes gross margin, EBITDA margin, NOPAT margin, R&D intensity, and capex intensity trends.",
    )
    _add_section(
        "Vaccine break-even analysis",
        "Included",
        "Highlights unit economics and break-even volumes by vaccine program.",
    )
    _add_section(
        "Scenario stress testing",
        "Included",
        "Compares rNPV outcomes under upside, base, and downside stress scenarios.",
    )
    _add_section(
        "Trend, seasonality & segmentation",
        "Included",
        "Decomposition trends and segmentation splits across revenue drivers.",
    )
    _add_section(
        "Monte Carlo & probabilistic valuation",
        "Included",
        "Monte Carlo simulation outputs provide probabilistic valuation ranges and downside risk bands.",
    )
    _add_section(
        "What-if analysis & goal seek",
        "Not available",
        "Goal seek and what-if sensitivity runs are not available in the current analytics export.",
    )
    _add_section(
        "Tornado & spider diagnostics",
        "Included",
        "Sensitivity drivers ranked by valuation impact.",
    )
    _add_section(
        "Regression & classification models",
        "Not available",
        "ML model outputs are not available in the current analytics export.",
    )
    _add_section(
        "Time-series & ML forecasting",
        "Not available",
        "Forecasting model results are not available in the current analytics export.",
    )
    _add_section(
        "Optimisation, portfolio design & real options",
        "Not available",
        "Optimization and real options outputs are not available in the current analytics export.",
    )
    _add_section(
        "Risk, copulas, macro & ESG linkages",
        "Not available",
        "Macro/ESG linkage analytics are not available in the current analytics export.",
    )
    _add_section(
        "Comparative & ML-based valuation",
        "Not available",
        "Comparable and ML valuation outputs are not available in the current analytics export.",
    )
    _add_section(
        "Scenario analysis",
        "Included",
        "Scenario results compared across key valuation drivers.",
    )
    _add_section(
        "Dashboard snapshot",
        "Included",
        "Snapshot of key dashboard metrics and FCFF trends.",
    )

    return sections


def _build_export_payload(
    bundle_payload: Dict[str, Any],
    analytics_df: Optional[pd.DataFrame] = None,
) -> Dict[str, Any]:
    snapshot_summary = bundle_payload["snapshot"]["financial_snapshot"]
    scenarios = snapshot_summary.get("scenarios") or []
    sensitivities = snapshot_summary.get("sensitivities") or []
    last_report = bundle_payload.get("last_report") or {}
    perf_df = bundle_payload.get("financial_performance")
    position_df = bundle_payload.get("financial_position")
    cash_flow_df = bundle_payload.get("cash_flows")
    cons_df = bundle_payload.get("financial_statements")
    ai_commentary = _build_ai_commentary(
        snapshot_summary,
        perf_df,
        position_df,
        cash_flow_df,
        cons_df,
        analytics_df=analytics_df,
    )
    summary_rows = [
        {"Metric": "Project ID", "Value": bundle_payload["snapshot"]["project_id"]},
        {"Metric": "Currency", "Value": snapshot_summary.get("currency")},
        {"Metric": "NPV", "Value": snapshot_summary.get("npv")},
        {"Metric": "IRR", "Value": snapshot_summary.get("irr")},
        {"Metric": "Min DSCR", "Value": snapshot_summary.get("dscr_min")},
        {"Metric": "Payback (years)", "Value": snapshot_summary.get("payback_years")},
        {"Metric": "Total Capex", "Value": snapshot_summary.get("capex_total")},
        {"Metric": "Annual Opex", "Value": snapshot_summary.get("opex_annual")},
        {"Metric": "Annual Revenue", "Value": snapshot_summary.get("revenue_annual")},
    ]
    return {
        "summary_rows": summary_rows,
        "scenarios": scenarios,
        "sensitivities": sensitivities,
        "last_report": last_report,
        "ai_config": bundle_payload["ai_config"],
        "financial_performance": perf_df,
        "financial_position": position_df,
        "cash_flows": cash_flow_df,
        "financial_statements": cons_df,
        "ai_commentary": ai_commentary,
    }


def _apply_cash_flow_assumptions(
    cash_flow_df: Optional[pd.DataFrame],
    snapshot_summary: Dict[str, Any],
) -> Optional[pd.DataFrame]:
    if cash_flow_df is None or cash_flow_df.empty:
        return cash_flow_df

    updated = cash_flow_df.copy()
    years = updated.index

    beginning_cash = float(snapshot_summary.get("beginning_cash") or 0.0)
    equity_issuance = float(snapshot_summary.get("equity_issuance") or 0.0)
    debt_draw = float(snapshot_summary.get("debt_draw") or 0.0)
    debt_repay = float(snapshot_summary.get("debt_repay") or 0.0)
    interest_paid = float(snapshot_summary.get("interest_paid") or 0.0)

    updated["Equity issuance"] = pd.Series(equity_issuance, index=years)
    updated["Debt drawdowns"] = pd.Series(debt_draw, index=years)
    updated["Debt repayments"] = pd.Series(debt_repay, index=years)
    updated["Interest paid"] = pd.Series(interest_paid, index=years)

    updated["Net cash from financing"] = (
        updated["Equity issuance"]
        + updated["Debt drawdowns"]
        - updated["Debt repayments"]
        - updated["Interest paid"]
    )
    updated["Net change in cash"] = (
        updated["Net cash from operations"]
        + updated["Net cash from investing"]
        + updated["Net cash from financing"]
    )
    return _roll_cash_balances(updated, opening_cash=beginning_cash)


def _financing_settings_from_state() -> Dict[str, object]:
    repayment_mode_value = str(
        st.session_state.get("debt_repayment_mode", DEBT_REPAYMENT_LABELS["straight_line"])
        or DEBT_REPAYMENT_LABELS["straight_line"]
    )
    repayment_mode = DEBT_REPAYMENT_CODES.get(repayment_mode_value, repayment_mode_value)
    if repayment_mode not in DEBT_REPAYMENT_LABELS:
        repayment_mode = "straight_line"
    return {
        "interest_rate": float(st.session_state.get("debt_interest_rate", 0.0)),
        "repayment_mode": repayment_mode,
        "grace_years": int(st.session_state.get("debt_grace_years", 0) or 0),
        "target_dscr": float(st.session_state.get("debt_target_dscr", 1.3) or 1.3),
        "minimum_cash_reserve": float(st.session_state.get("minimum_cash_reserve", 0.0) or 0.0),
    }


def _apply_debt_schedule(
    cash_flow_df: Optional[pd.DataFrame],
    debt_schedule: Optional[pd.DataFrame],
    interest_rate: float,
    repayment_mode: str = "straight_line",
    grace_years: int = 0,
    target_dscr: float = 1.3,
    minimum_cash_reserve: float = 0.0,
) -> Optional[pd.DataFrame]:
    if cash_flow_df is None or cash_flow_df.empty:
        return cash_flow_df
    if debt_schedule is None or debt_schedule.empty:
        return cash_flow_df

    updated = cash_flow_df.copy()
    schedule = debt_schedule.copy()
    if "Year" not in schedule.columns:
        return cash_flow_df

    template = _default_debt_schedule(int(updated.index.min()), len(updated.index))
    schedule = _align_table_to_template(schedule, template)
    schedule["Year"] = pd.to_numeric(schedule["Year"], errors="coerce").astype("Int64")
    schedule = schedule.dropna(subset=["Year"]).set_index("Year")
    if schedule.index.has_duplicates:
        schedule = schedule.groupby(level=0).sum()
    schedule = schedule.reindex(updated.index).fillna(0.0)

    drawdowns = pd.to_numeric(schedule.get("Debt drawdowns", 0.0), errors="coerce").fillna(0.0)
    manual_repayments = (
        pd.to_numeric(schedule.get("Manual debt repayments", 0.0), errors="coerce").fillna(0.0)
    )
    net_ops = _coerce_frame_column(updated, "Net cash from operations")
    net_investing = _coerce_frame_column(updated, "Net cash from investing")
    equity_issuance = _coerce_frame_column(updated, "Equity issuance")
    normalized_mode = str(repayment_mode or "straight_line").strip().lower()
    grace_periods = max(int(grace_years or 0), 0)
    reserve_floor = max(float(minimum_cash_reserve or 0.0), 0.0)
    dscr_target = max(float(target_dscr or 1.3), 0.01)
    opening_cash = 0.0
    if "Beginning cash balance" in updated.columns and not updated.empty:
        opening_cash = float(
            pd.to_numeric(updated["Beginning cash balance"], errors="coerce").fillna(0.0).iloc[0]
        )
    begin_balances = []
    principal_repayments = []
    interest_charges = []
    end_balances = []
    beginning_cash = []
    ending_cash = []
    net_financing = []
    net_change = []
    balance = 0.0
    cash_balance = opening_cash
    years = list(updated.index)
    total_years = len(years)
    for idx, year in enumerate(years):
        draw = float(drawdowns.loc[year]) if year in drawdowns.index else 0.0
        manual_principal = float(manual_repayments.loc[year]) if year in manual_repayments.index else 0.0
        outstanding = max(balance + draw, 0.0)
        interest = balance * float(interest_rate)
        if idx == total_years - 1:
            desired_principal = outstanding
        elif normalized_mode == "manual":
            desired_principal = manual_principal
        elif idx < grace_periods:
            desired_principal = 0.0
        elif normalized_mode == "bullet":
            desired_principal = 0.0
        elif normalized_mode == "sculpted_dscr":
            cfads = float(net_ops.loc[year] + net_investing.loc[year])
            max_service = max(cfads / dscr_target, 0.0)
            desired_principal = max(max_service - interest, 0.0)
        else:
            remaining_periods = max(total_years - max(idx, grace_periods), 1)
            desired_principal = outstanding / remaining_periods

        cash_before_principal = (
            cash_balance
            + float(net_ops.loc[year])
            + float(net_investing.loc[year])
            + float(equity_issuance.loc[year])
            + draw
            - interest
        )
        principal = min(max(desired_principal, 0.0), outstanding)
        if idx < total_years - 1:
            principal = min(principal, max(cash_before_principal - reserve_floor, 0.0))
        end_balance = max(outstanding - principal, 0.0)
        financing_cash = float(equity_issuance.loc[year]) + draw - principal - interest
        total_cash_change = float(net_ops.loc[year] + net_investing.loc[year] + financing_cash)
        ending_cash_balance = cash_balance + total_cash_change

        begin_balances.append(balance)
        principal_repayments.append(principal)
        interest_charges.append(interest)
        end_balances.append(end_balance)
        beginning_cash.append(cash_balance)
        ending_cash.append(ending_cash_balance)
        net_financing.append(financing_cash)
        net_change.append(total_cash_change)

        balance = end_balance
        cash_balance = ending_cash_balance

    updated["Debt drawdowns"] = pd.Series(drawdowns.values, index=updated.index)
    updated["Manual debt repayments"] = pd.Series(manual_repayments.values, index=updated.index)
    updated["Debt opening balance"] = pd.Series(begin_balances, index=updated.index)
    updated["Debt repayments"] = pd.Series(principal_repayments, index=updated.index)
    updated["Interest paid"] = pd.Series(interest_charges, index=updated.index)
    updated["Debt closing balance"] = pd.Series(end_balances, index=updated.index)
    updated["Net cash from financing"] = pd.Series(net_financing, index=updated.index)
    updated["Net change in cash"] = pd.Series(net_change, index=updated.index)
    updated["Beginning cash balance"] = pd.Series(beginning_cash, index=updated.index)
    updated["Ending cash balance"] = pd.Series(ending_cash, index=updated.index)

    return updated


def _build_lender_metrics(
    cash_flow_df: Optional[pd.DataFrame],
    discount_rate: float,
    minimum_cash_reserve: float = 0.0,
    target_dscr: float = 1.3,
) -> pd.DataFrame:
    if cash_flow_df is None or cash_flow_df.empty:
        return pd.DataFrame()

    cfads = _coerce_frame_column(cash_flow_df, "Net cash from operations") + _coerce_frame_column(
        cash_flow_df, "Net cash from investing"
    )
    opening_balance = _coerce_frame_column(cash_flow_df, "Debt opening balance")
    closing_balance = _coerce_frame_column(cash_flow_df, "Debt closing balance")
    principal = _coerce_frame_column(cash_flow_df, "Debt repayments")
    interest = _coerce_frame_column(cash_flow_df, "Interest paid")
    debt_service = principal + interest
    ending_cash = _coerce_frame_column(cash_flow_df, "Ending cash balance")
    reserve_headroom = ending_cash - float(minimum_cash_reserve or 0.0)
    dscr = pd.Series(np.where(debt_service > 0, cfads / debt_service, np.nan), index=cash_flow_df.index)
    discount = max(float(discount_rate or 0.0), 0.0)

    active_debt = (opening_balance > 0) | (closing_balance > 0) | (debt_service > 0)
    active_positions = np.flatnonzero(active_debt.to_numpy())
    last_debt_period = int(active_positions[-1]) if len(active_positions) else -1

    llcr_values: List[float] = []
    plcr_values: List[float] = []
    covenant_status: List[str] = []
    cfads_values = cfads.astype(float).to_numpy()
    opening_values = opening_balance.astype(float).to_numpy()

    for idx, year in enumerate(cash_flow_df.index):
        open_balance = opening_values[idx]
        if open_balance <= 0:
            llcr_values.append(np.nan)
            plcr_values.append(np.nan)
            covenant_status.append("N/A")
            continue

        future_cfads = cfads_values[idx:]
        offsets = np.arange(len(future_cfads), dtype=float)
        pv_project = float((future_cfads / np.power(1.0 + discount, offsets)).sum())
        if last_debt_period >= idx:
            debt_term_cfads = cfads_values[idx : last_debt_period + 1]
            debt_offsets = np.arange(len(debt_term_cfads), dtype=float)
            pv_debt_term = float((debt_term_cfads / np.power(1.0 + discount, debt_offsets)).sum())
        else:
            pv_debt_term = 0.0
        llcr_values.append(pv_debt_term / open_balance)
        plcr_values.append(pv_project / open_balance)

        status_parts: List[str] = []
        if not np.isnan(dscr.loc[year]) and float(dscr.loc[year]) < float(target_dscr):
            status_parts.append("DSCR breach")
        if float(reserve_headroom.loc[year]) < 0:
            status_parts.append("Reserve breach")
        covenant_status.append("Pass" if not status_parts else " + ".join(status_parts))

    return pd.DataFrame(
        {
            "CFADS": cfads,
            "Debt service": debt_service,
            "DSCR": dscr,
            "LLCR": pd.Series(llcr_values, index=cash_flow_df.index),
            "PLCR": pd.Series(plcr_values, index=cash_flow_df.index),
            "Minimum cash reserve": pd.Series(float(minimum_cash_reserve or 0.0), index=cash_flow_df.index),
            "Cash reserve headroom": reserve_headroom,
            "Covenant status": pd.Series(covenant_status, index=cash_flow_df.index),
        }
    )


def _build_enterprise_to_equity_bridge(
    valuation_result: Optional[ValuationResult],
    cash_flow_df: Optional[pd.DataFrame],
    planned_new_equity: float,
) -> pd.DataFrame:
    if valuation_result is None:
        return pd.DataFrame()

    enterprise_value = float(valuation_result.enterprise_value)
    ending_cash = 0.0
    debt_balance = 0.0
    if cash_flow_df is not None and not cash_flow_df.empty:
        if "Ending cash balance" in cash_flow_df.columns:
            ending_cash = float(pd.to_numeric(cash_flow_df["Ending cash balance"], errors="coerce").fillna(0.0).iloc[-1])
        if "Debt closing balance" in cash_flow_df.columns:
            debt_balance = float(pd.to_numeric(cash_flow_df["Debt closing balance"], errors="coerce").fillna(0.0).iloc[-1])

    net_debt = debt_balance - ending_cash
    pre_money_equity = enterprise_value - net_debt
    post_money_equity = pre_money_equity + float(planned_new_equity)

    return pd.DataFrame(
        [
            {"Component": "Enterprise value (DCF)", "Amount": enterprise_value},
            {"Component": "Less: debt outstanding", "Amount": -debt_balance},
            {"Component": "Add: cash / (cash deficit)", "Amount": ending_cash},
            {"Component": "Pre-money equity value", "Amount": pre_money_equity},
            {"Component": "Planned new equity", "Amount": float(planned_new_equity)},
            {"Component": "Post-money equity value", "Amount": post_money_equity},
        ]
    )


def _build_investor_waterfall(
    shareholders_df: Optional[pd.DataFrame],
    exit_equity_value: float,
) -> pd.DataFrame:
    if shareholders_df is None or shareholders_df.empty:
        return pd.DataFrame()

    template = _default_shareholders_table()
    waterfall = _align_table_to_template(shareholders_df, template).copy()
    waterfall["Ownership %"] = _coerce_numeric(waterfall.get("Ownership %", pd.Series(dtype=float)))
    waterfall["Investment"] = _coerce_numeric(waterfall.get("Investment", pd.Series(dtype=float)))
    waterfall["Seniority"] = _coerce_numeric(waterfall.get("Seniority", pd.Series(dtype=float)), default=99.0)
    waterfall["Liquidation preference (x)"] = _coerce_numeric(
        waterfall.get("Liquidation preference (x)", pd.Series(dtype=float))
    )
    waterfall["Participating preferred"] = waterfall.get(
        "Participating preferred", pd.Series(False, index=waterfall.index)
    ).apply(lambda value: bool(value) if isinstance(value, (bool, np.bool_)) else str(value).strip().lower() in {"1", "true", "yes", "y"})
    waterfall["Security"] = waterfall.get("Security", pd.Series("Common", index=waterfall.index)).astype(str)

    exit_value = max(float(exit_equity_value or 0.0), 0.0)
    waterfall["Converted value"] = waterfall["Ownership %"] * exit_value
    waterfall["Preference claim"] = waterfall["Investment"] * waterfall["Liquidation preference (x)"]

    preferred_mask = waterfall["Preference claim"] > 0
    convert_mask = preferred_mask & ~waterfall["Participating preferred"] & (
        waterfall["Converted value"] > waterfall["Preference claim"]
    )
    pref_pool_mask = preferred_mask & ~convert_mask

    waterfall["Decision"] = "Common"
    waterfall.loc[pref_pool_mask, "Decision"] = "Take preference"
    waterfall.loc[convert_mask, "Decision"] = "Convert to common"
    waterfall.loc[waterfall["Participating preferred"] & preferred_mask, "Decision"] = (
        "Participating preferred"
    )
    waterfall["Preference paid"] = 0.0

    remaining_exit = exit_value
    for seniority in sorted(waterfall.loc[pref_pool_mask, "Seniority"].unique()):
        mask = pref_pool_mask & (waterfall["Seniority"] == seniority)
        claim_total = float(waterfall.loc[mask, "Preference claim"].sum())
        if claim_total <= 0 or remaining_exit <= 0:
            continue
        payout = min(remaining_exit, claim_total)
        allocation = waterfall.loc[mask, "Preference claim"] / claim_total
        waterfall.loc[mask, "Preference paid"] = payout * allocation
        remaining_exit -= payout

    common_pool_mask = ~pref_pool_mask | waterfall["Participating preferred"]
    common_pool_ownership = float(waterfall.loc[common_pool_mask, "Ownership %"].sum())
    waterfall["Common pool allocation"] = 0.0
    if remaining_exit > 0 and common_pool_ownership > 0:
        waterfall.loc[common_pool_mask, "Common pool allocation"] = (
            remaining_exit
            * waterfall.loc[common_pool_mask, "Ownership %"]
            / common_pool_ownership
        )

    waterfall["Total proceeds"] = waterfall["Preference paid"] + waterfall["Common pool allocation"]
    waterfall["MOIC"] = np.where(
        waterfall["Investment"] > 0,
        waterfall["Total proceeds"] / waterfall["Investment"],
        np.nan,
    )

    columns = [
        "Shareholder",
        "Security",
        "Seniority",
        "Ownership %",
        "Investment",
        "Decision",
        "Converted value",
        "Preference claim",
        "Preference paid",
        "Common pool allocation",
        "Total proceeds",
        "MOIC",
    ]
    return waterfall[columns].sort_values(["Seniority", "Shareholder"]).reset_index(drop=True)


def _build_financing_outputs(
    valuation_result: Optional[ValuationResult],
    model_cfg: Optional[ModelConfig],
) -> Dict[str, pd.DataFrame]:
    outputs = {
        "financial_performance": pd.DataFrame(),
        "financial_position": pd.DataFrame(),
        "cash_flows": pd.DataFrame(),
        "lender_metrics": pd.DataFrame(),
        "equity_bridge": pd.DataFrame(),
        "investor_waterfall": pd.DataFrame(),
    }
    if valuation_result is None or model_cfg is None:
        return outputs

    perf_df, position_df, cash_flow_df = _compute_financial_statements(
        valuation_result.consolidated,
        model_cfg,
    )
    financing_settings = _financing_settings_from_state()
    cash_flow_df = _apply_debt_schedule(
        cash_flow_df,
        st.session_state.get("debt_schedule_table"),
        float(financing_settings["interest_rate"]),
        repayment_mode=str(financing_settings["repayment_mode"]),
        grace_years=int(financing_settings["grace_years"]),
        target_dscr=float(financing_settings["target_dscr"]),
        minimum_cash_reserve=float(financing_settings["minimum_cash_reserve"]),
    )
    lender_metrics = _build_lender_metrics(
        cash_flow_df,
        model_cfg.discount_rate,
        minimum_cash_reserve=float(financing_settings["minimum_cash_reserve"]),
        target_dscr=float(financing_settings["target_dscr"]),
    )
    equity_bridge = _build_enterprise_to_equity_bridge(
        valuation_result,
        cash_flow_df,
        float(st.session_state.get("planned_new_equity", 0.0)),
    )
    investor_waterfall = pd.DataFrame()
    if not equity_bridge.empty:
        post_money = float(
            equity_bridge.loc[
                equity_bridge["Component"] == "Post-money equity value",
                "Amount",
            ].iloc[0]
        )
        investor_waterfall = _build_investor_waterfall(
            st.session_state.get("shareholders_table"),
            post_money,
        )

    outputs["financial_performance"] = perf_df
    outputs["financial_position"] = position_df
    outputs["cash_flows"] = cash_flow_df if cash_flow_df is not None else pd.DataFrame()
    outputs["lender_metrics"] = lender_metrics
    outputs["equity_bridge"] = equity_bridge
    outputs["investor_waterfall"] = investor_waterfall
    return outputs


def _build_chart_tables(
    valuation_result: Optional[ValuationResult],
    model_cfg: Optional[ModelConfig],
    portfolio: Optional[Portfolio],
) -> Dict[str, pd.DataFrame]:
    tables: Dict[str, pd.DataFrame] = {}
    if valuation_result is None or model_cfg is None:
        return tables

    cons = valuation_result.consolidated.copy()
    cons_display = cons[["revenue", "ebitda", "fcff_after_wc"]].copy()
    cons_display.columns = ["Revenue", "EBITDA", "FCFF after WC"]
    tables["financial_statements_chart"] = cons_display
    tables["dashboard_chart"] = cons[["revenue", "ebitda", "fcff_after_wc"]]
    tables["dashboard_fcff_bar"] = cons[["fcff_after_wc"]]
    ratios = _build_ratio_table(cons)
    if not ratios.empty:
        tables["advanced_analytics_report"] = ratios
    break_even_df = _build_vaccine_break_even_table(model_cfg)
    if not break_even_df.empty:
        tables["vaccine_break_even_report"] = break_even_df

    decomp_df = _compute_decomposition(cons)
    if decomp_df is not None:
        tables["analytics_decomposition"] = decomp_df

    seg_df = _build_segmentation_table(valuation_result)
    if not seg_df.empty:
        tables["analytics_segmentation"] = seg_df

    if portfolio is not None:
        base_rnpv = valuation_result.rnpv
        tornado_df = _tornado_dataframe(portfolio, base_rnpv)
        if not tornado_df.empty:
            tables["analytics_tornado"] = tornado_df

        scenarios = [
            Scenario(
                name="Base case",
                revenue_multiplier=1.0,
                cost_multiplier=1.0,
                discount_rate_shift=0.0,
                success_prob_multiplier=1.0,
            ),
            Scenario(
                name="Upside",
                revenue_multiplier=1.2,
                cost_multiplier=0.9,
                discount_rate_shift=-0.01,
                success_prob_multiplier=1.1,
            ),
            Scenario(
                name="Downside",
                revenue_multiplier=0.8,
                cost_multiplier=1.1,
                discount_rate_shift=0.01,
                success_prob_multiplier=0.9,
            ),
        ]
        scen_results = ScenarioEngine(portfolio).run_scenarios(scenarios)
        tables["scenario_results"] = scen_results

    return tables


def _build_monte_carlo_results(snapshot_summary: Dict[str, Any]) -> pd.DataFrame:
    base_npv = snapshot_summary.get("npv")
    if base_npv is None:
        return pd.DataFrame()
    try:
        base_npv = float(base_npv)
    except (TypeError, ValueError):
        return pd.DataFrame()
    rng = np.random.default_rng(42)
    shocks = rng.normal(loc=0.0, scale=0.2, size=500)
    npv_samples = base_npv * (1 + shocks)
    return pd.DataFrame({"NPV": npv_samples})


def _build_chart_images(chart_tables: Dict[str, pd.DataFrame]) -> Dict[str, BytesIO]:
    images: Dict[str, BytesIO] = {}
    if importlib.util.find_spec("matplotlib") is None:
        return images

    import matplotlib.pyplot as plt
    import numpy as np

    def _save_fig(fig, key: str) -> None:
        buffer = BytesIO()
        fig.savefig(buffer, format="png", bbox_inches="tight")
        buffer.seek(0)
        images[key] = buffer
        plt.close(fig)

    if "financial_statements_chart" in chart_tables:
        fig, ax = plt.subplots()
        chart_tables["financial_statements_chart"].plot(ax=ax)
        ax.set_title("Financial Statements Overview")
        ax.set_xlabel("Year")
        ax.set_ylabel("Value")
        _save_fig(fig, "financial_statements_chart")

    if "dashboard_chart" in chart_tables:
        fig, ax = plt.subplots()
        chart_tables["dashboard_chart"].plot(ax=ax)
        ax.set_title("Dashboard Trends")
        ax.set_xlabel("Year")
        ax.set_ylabel("Value")
        _save_fig(fig, "dashboard_chart")

    if "dashboard_fcff_bar" in chart_tables:
        fig, ax = plt.subplots()
        chart_tables["dashboard_fcff_bar"].plot(kind="bar", ax=ax)
        ax.set_title("FCFF After WC")
        ax.set_xlabel("Year")
        ax.set_ylabel("Value")
        _save_fig(fig, "dashboard_fcff_bar")

    if "analytics_decomposition" in chart_tables:
        fig, ax = plt.subplots()
        chart_tables["analytics_decomposition"].plot(ax=ax)
        ax.set_title("Trend & Seasonality")
        ax.set_xlabel("Year")
        ax.set_ylabel("Value")
        _save_fig(fig, "analytics_decomposition")

    if "analytics_segmentation" in chart_tables:
        fig, ax = plt.subplots()
        seg_df = chart_tables["analytics_segmentation"]
        seg_df.set_index("Product")["Revenue share"].plot(kind="bar", ax=ax)
        ax.set_title("Revenue Share by Product")
        ax.set_xlabel("Product")
        ax.set_ylabel("Revenue Share")
        _save_fig(fig, "analytics_segmentation")

    if "analytics_tornado" in chart_tables:
        fig, ax = plt.subplots()
        tornado_df = chart_tables["analytics_tornado"].sort_values("Delta")
        ax.barh(tornado_df["Driver"], tornado_df["Delta"])
        ax.set_title("Tornado Impact")
        ax.set_xlabel("Delta")
        _save_fig(fig, "analytics_tornado")
        spider_df = chart_tables["analytics_tornado"].copy()
        if not spider_df.empty:
            labels = spider_df["Driver"].astype(str).tolist()
            values = spider_df["Delta"].abs().to_numpy()
            if values.sum() > 0:
                values = values / values.max()
            angles = np.linspace(0, 2 * np.pi, len(labels), endpoint=False).tolist()
            values = np.concatenate([values, values[:1]])
            angles += angles[:1]
            fig, ax = plt.subplots(subplot_kw={"polar": True})
            ax.plot(angles, values, linewidth=2)
            ax.fill(angles, values, alpha=0.25)
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(labels)
            ax.set_title("Spider Diagnostics (Normalized Impact)")
            _save_fig(fig, "spider_diagnostics")

    if "scenario_results" in chart_tables:
        fig, ax = plt.subplots()
        scen_df = chart_tables["scenario_results"]
        ax.bar(scen_df["scenario"], scen_df["rnpv"])
        ax.set_title("Scenario rNPV Comparison")
        ax.set_xlabel("Scenario")
        ax.set_ylabel("rNPV")
        _save_fig(fig, "scenario_results")
    if "scenario_custom" in chart_tables:
        fig, ax = plt.subplots()
        scen_df = chart_tables["scenario_custom"]
        ax.bar(scen_df["scenario"], scen_df["npv"])
        ax.set_title("Custom Scenario NPV Comparison")
        ax.set_xlabel("Scenario")
        ax.set_ylabel("NPV")
        _save_fig(fig, "scenario_custom")

    if "advanced_analytics_report" in chart_tables:
        ratio_df = chart_tables["advanced_analytics_report"].copy()
        fig, ax = plt.subplots()
        ratio_df.plot(ax=ax)
        ax.set_title("Margin & Intensity Analysis")
        ax.set_xlabel("Year")
        ax.set_ylabel("Ratio")
        _save_fig(fig, "margin_intensity_analysis")

    if "vaccine_break_even_report" in chart_tables:
        break_even_df = chart_tables["vaccine_break_even_report"]
        if not break_even_df.empty and "Vaccine name" in break_even_df.columns:
            fig, ax = plt.subplots()
            ax.bar(
                break_even_df["Vaccine name"],
                break_even_df["Break-even units"],
            )
            ax.set_title("Vaccine Break-even Units")
            ax.set_xlabel("Vaccine")
            ax.set_ylabel("Break-even units")
            _save_fig(fig, "vaccine_break_even_chart")

    if "monte_carlo_results" in chart_tables:
        mc_df = chart_tables["monte_carlo_results"]
        if not mc_df.empty and "NPV" in mc_df.columns:
            fig, ax = plt.subplots()
            ax.hist(mc_df["NPV"], bins=30, color="#1F4E78", alpha=0.75)
            ax.set_title("Monte Carlo NPV Distribution")
            ax.set_xlabel("NPV")
            ax.set_ylabel("Frequency")
            _save_fig(fig, "monte_carlo_results")

    return images


def _sync_vaccine_sales_products(
    product_df: pd.DataFrame,
    vaccine_sales_df: pd.DataFrame,
) -> pd.DataFrame:
    if vaccine_sales_df.empty:
        return product_df
    if "Implied revenue" not in vaccine_sales_df.columns:
        return product_df

    updated = product_df.copy()
    if "name" not in updated.columns:
        return updated

    grouped = (
        vaccine_sales_df.groupby(["ID_vaccine", "Vaccine name"], dropna=False)["Implied revenue"]
        .mean()
        .reset_index()
    )
    for _, row in grouped.iterrows():
        vaccine_name = str(row.get("Vaccine name") or row.get("ID_vaccine") or "Vaccine")
        avg_revenue = float(row.get("Implied revenue") or 0.0)
        default_row = _blank_product_row(name=vaccine_name)
        default_row.update(
            {
                "stage": "Commercial",
                "success_prob": 1.0,
                "include_in_consolidation": True,
                "preexisting_market": True,
                "time_to_market": 0,
                "patent_years": 20,
                "patent_revenue_target": avg_revenue,
                "post_patent_revenue_target": avg_revenue,
                "market_growth_patent": 0.0,
                "market_growth_post": 0.0,
            }
        )
        match = updated["name"] == vaccine_name
        if match.any():
            idx = updated.index[match][0]
            for key, value in default_row.items():
                if key in updated.columns:
                    updated.at[idx, key] = value
        else:
            updated = pd.concat([updated, pd.DataFrame([default_row])], ignore_index=True)
    return updated


def _build_excel_export(payload: Dict[str, Any]) -> io.BytesIO:
    xlsx_image = None
    if importlib.util.find_spec("openpyxl") is not None:
        xlsx_image = importlib.import_module("openpyxl.drawing.image").Image

    def _round_table(df: pd.DataFrame) -> pd.DataFrame:
        return df.apply(pd.to_numeric, errors="ignore").round(0)

    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        pd.DataFrame(payload["summary_rows"]).to_excel(writer, index=False, sheet_name="Summary")
        if payload.get("ai_commentary"):
            ai_commentary = payload["ai_commentary"]
            if isinstance(ai_commentary, list) and ai_commentary and isinstance(ai_commentary[0], dict):
                pd.DataFrame(ai_commentary).to_excel(
                    writer,
                    index=False,
                    sheet_name="AI Commentary",
                )
            else:
                pd.DataFrame({"AI commentary": ai_commentary}).to_excel(
                    writer,
                    index=False,
                    sheet_name="AI Commentary",
                )
        if payload["scenarios"]:
            pd.DataFrame(payload["scenarios"]).to_excel(writer, index=False, sheet_name="Scenarios")
        if payload["sensitivities"]:
            pd.DataFrame(payload["sensitivities"]).to_excel(writer, index=False, sheet_name="Sensitivities")
        if payload["last_report"]:
            pd.DataFrame(
                [{"Section": key, "Content": value} for key, value in payload["last_report"].items()]
            ).to_excel(writer, index=False, sheet_name="Last Report")
        if payload.get("advanced_analytics_narrative"):
            pd.DataFrame({"Narrative": payload["advanced_analytics_narrative"]}).to_excel(
                writer,
                index=False,
                sheet_name="Advanced Analytics Narrative",
            )
        if payload.get("extended_analytics_sections"):
            pd.DataFrame(payload["extended_analytics_sections"]).to_excel(
                writer,
                index=False,
                sheet_name="Advanced Analytics Coverage",
            )
        if payload.get("financial_statements") is not None:
            _round_table(payload["financial_statements"]).to_excel(
                writer,
                index=True,
                sheet_name="Financial Statements",
            )
        if payload.get("financial_performance") is not None:
            _round_table(payload["financial_performance"]).to_excel(
                writer,
                index=True,
                sheet_name="Financial Performance",
            )
        if payload.get("financial_position") is not None:
            _round_table(payload["financial_position"]).to_excel(
                writer,
                index=True,
                sheet_name="Financial Position",
            )
        if payload.get("cash_flows") is not None:
            _round_table(payload["cash_flows"]).to_excel(
                writer,
                index=True,
                sheet_name="Cash Flows",
            )
        chart_tables = payload.get("chart_tables", {})
        for sheet_name, table in chart_tables.items():
            if not table.empty:
                safe_name = sheet_name[:31]
                table.to_excel(writer, index=True, sheet_name=safe_name)
        chart_images = payload.get("chart_images", {})
        if chart_images and xlsx_image is not None:
            workbook = writer.book

            def _add_chart_sheet(title: str, image_key: str) -> None:
                image = chart_images.get(image_key)
                if not image:
                    return
                sheet_title = title[:31]
                if sheet_title in workbook.sheetnames:
                    sheet = workbook[sheet_title]
                else:
                    sheet = workbook.create_sheet(sheet_title)
                image.seek(0)
                sheet.add_image(xlsx_image(image), "A1")

            _add_chart_sheet("Financial Statements Charts", "financial_statements_chart")
            _add_chart_sheet("Dashboard Charts", "dashboard_chart")
            _add_chart_sheet("Dashboard Charts", "dashboard_fcff_bar")
            _add_chart_sheet("Advanced Analytics Charts", "analytics_decomposition")
            _add_chart_sheet("Advanced Analytics Charts", "analytics_segmentation")
            _add_chart_sheet("Advanced Analytics Charts", "analytics_tornado")
            _add_chart_sheet("Advanced Analytics Charts", "spider_diagnostics")
            _add_chart_sheet("Advanced Analytics Charts", "margin_intensity_analysis")
            _add_chart_sheet("Advanced Analytics Charts", "vaccine_break_even_chart")
            _add_chart_sheet("Advanced Analytics Charts", "monte_carlo_results")
            _add_chart_sheet("Scenario Analysis Charts", "scenario_results")
            _add_chart_sheet("Scenario Analysis Charts", "scenario_custom")
    excel_buffer.seek(0)
    return excel_buffer


def _build_word_export(payload: Dict[str, Any]) -> io.BytesIO:
    def _round_table(df: pd.DataFrame) -> pd.DataFrame:
        return df.apply(pd.to_numeric, errors="ignore").round(0)

    def _format_value(value: Any) -> str:
        if isinstance(value, (int, float, np.integer, np.floating)):
            if np.isnan(value):
                return ""
            return f"{value:,.0f}"
        return str(value)

    def _set_section_orientation(document, orientation) -> None:
        section = document.sections[-1]
        section.orientation = orientation
        section.page_width, section.page_height = section.page_height, section.page_width

    def _add_docx_table(document, title: str, df: pd.DataFrame) -> None:
        if df is None or df.empty:
            return
        document.add_heading(title, level=2)
        table_df = _round_table(df.copy())
        table_df.insert(0, "Year", table_df.index)
        table = document.add_table(rows=1, cols=len(table_df.columns))
        table.style = "Light Grid"
        hdr_cells = table.rows[0].cells
        for idx, col_name in enumerate(table_df.columns):
            hdr_cells[idx].text = str(col_name)
        for _, row in table_df.iterrows():
            row_cells = table.add_row().cells
            for idx, value in enumerate(row):
                row_cells[idx].text = _format_value(value)
        for row in table.rows:
            for cell in row.cells:
                for paragraph in cell.paragraphs:
                    for run in paragraph.runs:
                        run.font.size = shared.Pt(7)

    def _safe_add_picture(document, image: Any) -> None:
        if image is None:
            return
        try:
            if isinstance(image, (bytes, bytearray)):
                image = BytesIO(image)
            elif hasattr(image, "getvalue"):
                try:
                    image_bytes = image.getvalue()
                except ValueError:
                    image.seek(0)
                    image_bytes = image.read()
                image = BytesIO(image_bytes)
            elif hasattr(image, "save"):
                buffer = BytesIO()
                image.save(buffer, format="PNG")
                buffer.seek(0)
                image = buffer
            elif hasattr(image, "seek"):
                image.seek(0)
            document.add_picture(image)
        except Exception:
            return

    docx_module = importlib.import_module("docx")
    shared = importlib.import_module("docx.shared")
    Document = docx_module.Document
    docx_section = importlib.import_module("docx.enum.section")
    WD_ORIENT = docx_section.WD_ORIENT
    WD_SECTION = docx_section.WD_SECTION
    docx_buffer = io.BytesIO()
    document = Document()
    styles = document.styles
    primary_color = shared.RGBColor(31, 78, 120)
    accent_color = shared.RGBColor(58, 58, 58)
    normal_style = styles["Normal"]
    normal_style.font.name = "Calibri"
    normal_style.font.size = shared.Pt(11)
    normal_style.paragraph_format.space_after = shared.Pt(6)
    title_style = styles["Title"]
    title_style.font.name = "Calibri"
    title_style.font.size = shared.Pt(26)
    title_style.font.color.rgb = primary_color
    subtitle_style = styles["Subtitle"]
    subtitle_style.font.name = "Calibri"
    subtitle_style.font.size = shared.Pt(12)
    subtitle_style.font.color.rgb = accent_color
    for heading_name, size in [("Heading 1", 18), ("Heading 2", 14), ("Heading 3", 12)]:
        heading_style = styles[heading_name]
        heading_style.font.name = "Calibri"
        heading_style.font.size = shared.Pt(size)
        heading_style.font.color.rgb = primary_color

    document.add_paragraph("Business Plan Bundle", style="Title")
    document.add_paragraph(
        "Financial report, analytics, and AI-assisted commentary",
        style="Subtitle",
    )
    document.add_paragraph(
        "This bundle summarizes the financial snapshot and the AI configuration used for the "
        "RAG Assistant report generation."
    )
    document.add_heading("Financial Snapshot", level=2)
    for row in payload["summary_rows"]:
        document.add_paragraph(f"{row['Metric']}: {row['Value']}")
    if payload.get("ai_commentary"):
        document.add_heading("AI Commentary", level=2)
        grouped_comments = _group_ai_commentary(payload["ai_commentary"])
        for section, entries in grouped_comments.items():
            document.add_heading(section, level=3)
            for entry in entries:
                document.add_paragraph(entry.get("Commentary", ""), style="List Bullet")
                annotation = entry.get("Annotation")
                if annotation:
                    document.add_paragraph(f"Annotation: {annotation}", style="List Bullet")
    if payload["scenarios"]:
        document.add_heading("Scenarios", level=2)
        currency = next(
            (row.get("Value") for row in payload.get("summary_rows", []) if row.get("Metric") == "Currency"),
            "USD",
        )
        for scenario in payload["scenarios"]:
            document.add_paragraph(_format_scenario_prose(scenario, currency))
    if payload["sensitivities"]:
        document.add_heading("Sensitivities", level=2)
        for sensitivity in payload["sensitivities"]:
            document.add_paragraph(json.dumps(sensitivity, ensure_ascii=False))
    has_financial_tables = any(
        payload.get(key) is not None
        for key in [
            "financial_statements",
            "financial_performance",
            "financial_position",
            "cash_flows",
        ]
    )
    if has_financial_tables:
        document.add_section(WD_SECTION.NEW_PAGE)
        _set_section_orientation(document, WD_ORIENT.LANDSCAPE)
    if payload.get("financial_statements") is not None:
        _add_docx_table(
            document,
            "Financial Statements",
            payload["financial_statements"],
        )
    if payload.get("financial_performance") is not None:
        _add_docx_table(
            document,
            "Statement of Financial Performance",
            payload["financial_performance"],
        )
    if payload.get("financial_position") is not None:
        _add_docx_table(
            document,
            "Statement of Financial Position",
            payload["financial_position"],
        )
    if payload.get("cash_flows") is not None:
        _add_docx_table(
            document,
            "Statement of Cash Flows",
            payload["cash_flows"],
        )
    if has_financial_tables:
        document.add_section(WD_SECTION.NEW_PAGE)
        _set_section_orientation(document, WD_ORIENT.PORTRAIT)
    if payload.get("chart_tables", {}).get("advanced_analytics_report") is not None:
        document.add_heading("Advanced analytics report", level=2)
        analytics_df = payload["chart_tables"]["advanced_analytics_report"]
        narrative = payload.get("advanced_analytics_narrative") or _build_advanced_analytics_narrative(
            analytics_df
        )
        for paragraph in narrative:
            document.add_paragraph(paragraph)
    if payload.get("extended_analytics_sections"):
        document.add_heading("Advanced analytics coverage", level=2)
        for entry in payload["extended_analytics_sections"]:
            document.add_paragraph(
                f"{entry.get('Section')}: {entry.get('Status')}",
                style="List Bullet",
            )
            details = entry.get("Details")
            if details:
                document.add_paragraph(details, style="List Bullet")
    if payload.get("chart_tables", {}).get("vaccine_break_even_report") is not None:
        document.add_heading("Vaccine break-even analysis", level=2)
        break_even_df = payload["chart_tables"]["vaccine_break_even_report"]
        for _, row in break_even_df.iterrows():
            document.add_paragraph(
                f"{row.get('Vaccine name', '')}: unit price {row.get('Unit price (USD)')}, "
                f"unit variable cost {row.get('Unit variable cost (USD)')}, "
                f"unit fixed cost {row.get('Unit fixed cost (USD/year)')}, "
                f"unit margin {row.get('Unit contribution margin (USD)')}, "
                f"break-even units {row.get('Break-even units')}"
            )
    document.add_heading("AI Configuration", level=2)
    for key, value in payload["ai_config"].items():
        document.add_paragraph(f"{key}: {value}")
    if payload["last_report"]:
        document.add_heading("Last Report", level=2)
        for key, value in payload["last_report"].items():
            document.add_paragraph(f"{key}: {value}")
    if payload.get("chart_images"):
        document.add_heading("Financial Statements Charts", level=2)
        if payload["chart_images"].get("financial_statements_chart"):
            _safe_add_picture(document, payload["chart_images"]["financial_statements_chart"])
        document.add_heading("Dashboard Charts", level=2)
        if payload["chart_images"].get("dashboard_chart"):
            _safe_add_picture(document, payload["chart_images"]["dashboard_chart"])
        if payload["chart_images"].get("dashboard_fcff_bar"):
            _safe_add_picture(document, payload["chart_images"]["dashboard_fcff_bar"])
        document.add_heading("Advanced Analytics Charts", level=2)
        if payload["chart_images"].get("analytics_decomposition"):
            _safe_add_picture(document, payload["chart_images"]["analytics_decomposition"])
        if payload["chart_images"].get("analytics_segmentation"):
            _safe_add_picture(document, payload["chart_images"]["analytics_segmentation"])
        if payload["chart_images"].get("analytics_tornado"):
            _safe_add_picture(document, payload["chart_images"]["analytics_tornado"])
        if payload["chart_images"].get("spider_diagnostics"):
            _safe_add_picture(document, payload["chart_images"]["spider_diagnostics"])
        if payload["chart_images"].get("margin_intensity_analysis"):
            _safe_add_picture(document, payload["chart_images"]["margin_intensity_analysis"])
        if payload["chart_images"].get("vaccine_break_even_chart"):
            _safe_add_picture(document, payload["chart_images"]["vaccine_break_even_chart"])
        if payload["chart_images"].get("monte_carlo_results"):
            _safe_add_picture(document, payload["chart_images"]["monte_carlo_results"])
        document.add_heading("Scenario Analysis Charts", level=2)
        if payload["chart_images"].get("scenario_results"):
            _safe_add_picture(document, payload["chart_images"]["scenario_results"])
        if payload["chart_images"].get("scenario_custom"):
            _safe_add_picture(document, payload["chart_images"]["scenario_custom"])
    document.save(docx_buffer)
    docx_buffer.seek(0)
    return docx_buffer


def _build_pdf_export(payload: Dict[str, Any]) -> io.BytesIO:
    canvas = importlib.import_module("reportlab.pdfgen.canvas")
    image_reader = importlib.import_module("reportlab.lib.utils").ImageReader
    tables = importlib.import_module("reportlab.platypus.tables")
    pagesizes = importlib.import_module("reportlab.lib.pagesizes")
    colors = importlib.import_module("reportlab.lib.colors")
    import textwrap
    pdf_buffer = io.BytesIO()
    pdf_canvas = canvas.Canvas(pdf_buffer)
    portrait_size = pagesizes.letter
    landscape_size = pagesizes.landscape(portrait_size)
    left_margin = 72
    primary_color = colors.HexColor("#1F4E78")
    accent_color = colors.HexColor("#3A3A3A")
    page_width = portrait_size[0]
    top_margin = 72
    bottom_margin = 72

    def _reset_page(page_size) -> None:
        nonlocal y_position, page_width
        pdf_canvas.setPageSize(page_size)
        page_width = page_size[0]
        y_position = page_size[1] - top_margin
        pdf_canvas.setFillColor(colors.black)

    def _draw_cover() -> None:
        nonlocal y_position
        pdf_canvas.setFont("Helvetica-Bold", 18)
        pdf_canvas.setFillColor(primary_color)
        pdf_canvas.drawCentredString(page_width / 2, y_position, "Business Plan Bundle")
        pdf_canvas.setFont("Helvetica", 12)
        pdf_canvas.setFillColor(accent_color)
        pdf_canvas.drawCentredString(
            page_width / 2,
            y_position - 20,
            "Financial report, analytics, and AI commentary",
        )
        pdf_canvas.setFillColor(colors.black)
        pdf_canvas.setFont("Helvetica", 11)
        y_position -= 44

    def _ensure_space(required: float, page_size) -> None:
        nonlocal y_position
        if y_position - required <= bottom_margin:
            pdf_canvas.showPage()
            _reset_page(page_size)
            _draw_cover()

    _reset_page(portrait_size)
    _draw_cover()

    def _draw_section_title(title: str) -> None:
        nonlocal y_position
        _ensure_space(24, portrait_size)
        pdf_canvas.setFont("Helvetica-Bold", 12)
        pdf_canvas.setFillColor(primary_color)
        pdf_canvas.drawString(left_margin, y_position, title)
        y_position -= 6
        pdf_canvas.setStrokeColor(primary_color)
        pdf_canvas.line(left_margin, y_position, page_width - left_margin, y_position)
        y_position -= 14
        pdf_canvas.setFillColor(colors.black)
        pdf_canvas.setFont("Helvetica", 11)

    _draw_section_title("Financial Snapshot")
    for row in payload["summary_rows"]:
        _ensure_space(16, portrait_size)
        pdf_canvas.drawString(left_margin, y_position, f"{row['Metric']}: {row['Value']}")
        y_position -= 16
    if payload.get("ai_commentary"):
        y_position -= 6
        _ensure_space(18, portrait_size)
        _draw_section_title("AI Commentary")
        grouped_comments = _group_ai_commentary(payload["ai_commentary"])
        for section, entries in grouped_comments.items():
            _ensure_space(16, portrait_size)
            pdf_canvas.drawString(left_margin, y_position, section)
            y_position -= 16
            for entry in entries:
                lines = textwrap.wrap(f"- {entry.get('Commentary', '')}", width=92)
                for line in lines:
                    _ensure_space(16, portrait_size)
                    pdf_canvas.drawString(left_margin, y_position, line)
                    y_position -= 16
                annotation = entry.get("Annotation")
                if annotation:
                    for line in textwrap.wrap(f"Annotation: {annotation}", width=92):
                        _ensure_space(16, portrait_size)
                        pdf_canvas.drawString(left_margin + 18, y_position, line)
                        y_position -= 16
    if payload["scenarios"]:
        _ensure_space(18, portrait_size)
        _draw_section_title("Scenarios")
        currency = next(
            (row.get("Value") for row in payload.get("summary_rows", []) if row.get("Metric") == "Currency"),
            "USD",
        )
        for scenario in payload["scenarios"]:
            for line in textwrap.wrap(_format_scenario_prose(scenario, currency), width=92):
                _ensure_space(16, portrait_size)
                pdf_canvas.drawString(left_margin, y_position, line)
                y_position -= 16
            _ensure_space(6, portrait_size)
            y_position -= 4
    if payload["sensitivities"]:
        _ensure_space(18, portrait_size)
        _draw_section_title("Sensitivities")
        for sensitivity in payload["sensitivities"]:
            lines = textwrap.wrap(json.dumps(sensitivity, ensure_ascii=False), width=92)
            for line in lines:
                _ensure_space(16, portrait_size)
                pdf_canvas.drawString(left_margin, y_position, line)
                y_position -= 16
    def _round_table(df: pd.DataFrame) -> pd.DataFrame:
        return df.apply(pd.to_numeric, errors="ignore").round(0)

    def _format_value(value: Any) -> str:
        if isinstance(value, (int, float, np.integer, np.floating)):
            if np.isnan(value):
                return ""
            return f"{value:,.0f}"
        return str(value)

    def _switch_orientation(page_size) -> None:
        pdf_canvas.showPage()
        _reset_page(page_size)

    def _draw_pdf_table(title: str, df: pd.DataFrame, page_size) -> None:
        nonlocal y_position, pdf_canvas
        if df is None or df.empty:
            return
        _ensure_space(32, page_size)
        pdf_canvas.setFont("Helvetica-Bold", 12)
        pdf_canvas.setFillColor(primary_color)
        pdf_canvas.drawString(left_margin, y_position, title)
        pdf_canvas.setFillColor(colors.black)
        pdf_canvas.setFont("Helvetica", 11)
        y_position -= 12

        table_df = _round_table(df.copy())
        table_df.insert(0, "Year", table_df.index)
        data = [list(table_df.columns)] + table_df.reset_index(drop=True).values.tolist()
        data = [[_format_value(value) for value in row] for row in data]
        table = tables.Table(data, repeatRows=1)
        style = tables.TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), "#1F4E78"),
                ("TEXTCOLOR", (0, 0), (-1, 0), "#FFFFFF"),
                ("GRID", (0, 0), (-1, -1), 0.25, "#CCCCCC"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]
        )
        table.setStyle(style)
        available_width = page_size[0] - (left_margin * 2)
        width, height = table.wrap(available_width, y_position - bottom_margin)
        if y_position - height <= bottom_margin:
            _switch_orientation(page_size)
            pdf_canvas.setFont("Helvetica-Bold", 12)
            pdf_canvas.setFillColor(primary_color)
            pdf_canvas.drawString(left_margin, y_position, title)
            pdf_canvas.setFillColor(colors.black)
            pdf_canvas.setFont("Helvetica", 11)
            y_position -= 12
            width, height = table.wrap(available_width, y_position - bottom_margin)
        table.drawOn(pdf_canvas, left_margin, y_position - height)
        y_position -= height + 18

    perf_df = payload.get("financial_performance")
    cons_df = payload.get("financial_statements")
    position_df = payload.get("financial_position")
    cash_flow_df = payload.get("cash_flows")
    has_financial_tables = any(
        table is not None for table in [perf_df, cons_df, position_df, cash_flow_df]
    )
    if has_financial_tables:
        _switch_orientation(landscape_size)
    if perf_df is not None:
        _draw_pdf_table("Statement of Financial Performance", perf_df, landscape_size)
    if cons_df is not None:
        _draw_pdf_table("Financial Statements", cons_df, landscape_size)
    if position_df is not None:
        _draw_pdf_table("Statement of Financial Position", position_df, landscape_size)
    if cash_flow_df is not None:
        _draw_pdf_table("Statement of Cash Flows", cash_flow_df, landscape_size)
    if has_financial_tables:
        _switch_orientation(portrait_size)
    analytics_df = payload.get("chart_tables", {}).get("advanced_analytics_report")
    if analytics_df is not None:
        _ensure_space(18, portrait_size)
        _draw_section_title("Advanced analytics report")
        narrative = payload.get("advanced_analytics_narrative") or _build_advanced_analytics_narrative(
            analytics_df
        )
        for paragraph in narrative:
            for line in textwrap.wrap(paragraph, width=92):
                _ensure_space(16, portrait_size)
                pdf_canvas.drawString(left_margin, y_position, line)
                y_position -= 16
    if payload.get("extended_analytics_sections"):
        _ensure_space(18, portrait_size)
        _draw_section_title("Advanced analytics coverage")
        for entry in payload["extended_analytics_sections"]:
            _ensure_space(16, portrait_size)
            pdf_canvas.drawString(
                left_margin,
                y_position,
                f"{entry.get('Section')}: {entry.get('Status')}",
            )
            y_position -= 16
            details = entry.get("Details")
            if details:
                for line in textwrap.wrap(details, width=92):
                    _ensure_space(16, portrait_size)
                    pdf_canvas.drawString(left_margin + 14, y_position, line)
                    y_position -= 16
    break_even_df = payload.get("chart_tables", {}).get("vaccine_break_even_report")
    if break_even_df is not None:
        _ensure_space(18, portrait_size)
        _draw_section_title("Vaccine break-even analysis")
        for _, row in break_even_df.iterrows():
            line = (
                f"{row.get('Vaccine name', '')}: unit price {row.get('Unit price (USD)')}, "
                f"unit variable cost {row.get('Unit variable cost (USD)')}, "
                f"unit fixed cost {row.get('Unit fixed cost (USD/year)')}, "
                f"unit margin {row.get('Unit contribution margin (USD)')}, "
                f"break-even units {row.get('Break-even units')}"
            )
            for wrapped in textwrap.wrap(line, width=92):
                _ensure_space(16, portrait_size)
                pdf_canvas.drawString(left_margin, y_position, wrapped)
                y_position -= 16
    _ensure_space(18, portrait_size)
    _draw_section_title("AI Configuration")
    for key, value in payload["ai_config"].items():
        _ensure_space(16, portrait_size)
        pdf_canvas.drawString(left_margin, y_position, f"{key}: {value}")
        y_position -= 16
    if payload["last_report"]:
        _ensure_space(18, portrait_size)
        _draw_section_title("Last Report")
        for key, value in payload["last_report"].items():
            for line in textwrap.wrap(f"{key}: {value}", width=92):
                _ensure_space(16, portrait_size)
                pdf_canvas.drawString(left_margin, y_position, line)
                y_position -= 16
    chart_images = payload.get("chart_images", {})
    if chart_images:
        pdf_canvas.showPage()
        pdf_canvas.setFont("Helvetica-Bold", 14)
        pdf_canvas.drawString(72, 770, "Charts & Graphs")
        y_position = 740
        pdf_canvas.setFont("Helvetica", 11)

        def _draw_image(image_key: str, title: str) -> None:
            nonlocal y_position
            image = chart_images.get(image_key)
            if not image:
                return
            try:
                if isinstance(image, (bytes, bytearray)):
                    image = BytesIO(image)
                elif hasattr(image, "getvalue"):
                    try:
                        image_bytes = image.getvalue()
                    except ValueError:
                        image.seek(0)
                        image_bytes = image.read()
                    image = BytesIO(image_bytes)
                elif hasattr(image, "save"):
                    buffer = BytesIO()
                    image.save(buffer, format="PNG")
                    buffer.seek(0)
                    image = buffer
                elif hasattr(image, "seek"):
                    image.seek(0)
            except Exception:
                return
            if y_position <= 180:
                pdf_canvas.showPage()
                _reset_page(portrait_size)
                pdf_canvas.setFont("Helvetica-Bold", 14)
                pdf_canvas.setFillColor(primary_color)
                pdf_canvas.drawString(left_margin, y_position, "Charts & Graphs (cont.)")
                pdf_canvas.setFillColor(colors.black)
                pdf_canvas.setFont("Helvetica", 11)
                y_position -= 20
            pdf_canvas.drawString(left_margin, y_position, title)
            y_position -= 14
            pdf_canvas.drawImage(image_reader(image), left_margin, y_position - 120, width=450, height=120)
            y_position -= 140

        _draw_image("financial_statements_chart", "Financial Statements")
        _draw_image("dashboard_chart", "Dashboard Trends")
        _draw_image("dashboard_fcff_bar", "Dashboard FCFF")
        _draw_image("analytics_decomposition", "Analytics Decomposition")
        _draw_image("analytics_segmentation", "Analytics Segmentation")
        _draw_image("analytics_tornado", "Analytics Tornado")
        _draw_image("spider_diagnostics", "Spider Diagnostics")
        _draw_image("margin_intensity_analysis", "Margin & Intensity Analysis")
        _draw_image("vaccine_break_even_chart", "Vaccine Break-even Analysis")
        _draw_image("monte_carlo_results", "Monte Carlo NPV Distribution")
        _draw_image("scenario_results", "Scenario Analysis")
        _draw_image("scenario_custom", "Custom Scenario Analysis")
    pdf_canvas.save()
    pdf_buffer.seek(0)
    return pdf_buffer


def _build_export_buffers(payload: Dict[str, Any]) -> Tuple[Dict[str, io.BytesIO], List[str]]:
    buffers: Dict[str, io.BytesIO] = {}
    warnings: List[str] = []
    if importlib.util.find_spec("openpyxl") is not None:
        buffers["excel"] = _build_excel_export(payload)
    else:
        warnings.append("Excel export unavailable: install openpyxl.")

    if importlib.util.find_spec("docx") is not None:
        buffers["docx"] = _build_word_export(payload)
    else:
        warnings.append("Word export unavailable: install python-docx.")

    if importlib.util.find_spec("reportlab") is not None:
        buffers["pdf"] = _build_pdf_export(payload)
    else:
        warnings.append("PDF export unavailable: install reportlab.")

    if payload.get("chart_tables") and importlib.util.find_spec("matplotlib") is None:
        warnings.append("Charts export unavailable: install matplotlib to embed plots.")

    return buffers, warnings


def _render_export_downloads(
    buffers: Dict[str, io.BytesIO],
    *,
    project_id: str,
    rag_key_prefix: str,
) -> None:
    if "excel" in buffers:
        st.download_button(
            "Download business plan (Excel)",
            data=buffers["excel"],
            file_name=f"{project_id}_business_plan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{rag_key_prefix}_bundle_download_excel",
        )
    if "docx" in buffers:
        st.download_button(
            "Download business plan (Word)",
            data=buffers["docx"],
            file_name=f"{project_id}_business_plan.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            use_container_width=True,
            key=f"{rag_key_prefix}_bundle_download_docx",
        )
    if "pdf" in buffers:
        st.download_button(
            "Download business plan (PDF)",
            data=buffers["pdf"],
            file_name=f"{project_id}_business_plan.pdf",
            mime="application/pdf",
            use_container_width=True,
            key=f"{rag_key_prefix}_bundle_download_pdf",
        )


def _render_rag_assistant_page() -> None:
    st.subheader("RAG Assistant")

    rag_key_prefix = "rag_assistant"
    st.markdown("## Upload reference documents")
    project_id = st.session_state.get(
        f"{rag_key_prefix}_project_id",
        os.environ.get("RAG_PROJECT_ID", "default-project"),
    )
    rag_host = st.session_state.get(
        f"{rag_key_prefix}_rag_host",
        os.environ.get("RAG_HOST", "http://localhost:8000"),
    )
    st.caption("Set RAG_HOST and RAG_PROJECT_ID env vars to override defaults.")

    uploads = st.file_uploader(
        "Upload reference documents",
        accept_multiple_files=True,
        key=f"{rag_key_prefix}_uploads",
    )
    if uploads:
        st.caption(f"{len(uploads)} document(s) ready for indexing.")

    st.markdown("## AI & Machine Learning Configuration")
    enable_ai = st.checkbox("Enable AI enhancements", value=True, key=f"{rag_key_prefix}_enable_ai")
    provider_options_key = f"{rag_key_prefix}_provider_options"
    if provider_options_key not in st.session_state:
        st.session_state[provider_options_key] = [
            "OpenAI",
            "Azure OpenAI",
            "Anthropic",
            "Vertex",
            "Custom",
        ]
    provider_options = st.session_state[provider_options_key]
    provider = st.selectbox(
        "Provider",
        provider_options,
        key=f"{rag_key_prefix}_provider",
    )
    custom_provider = ""
    if provider == "Custom":
        custom_provider = st.text_input(
            "Custom provider name",
            placeholder="Enter a provider name (e.g., Cohere, Mistral)",
            key=f"{rag_key_prefix}_custom_provider",
        )
        add_provider = st.button("Add provider", key=f"{rag_key_prefix}_add_provider")
        if add_provider and custom_provider:
            updated_providers = [*provider_options]
            if custom_provider not in updated_providers:
                updated_providers.insert(-1, custom_provider)
                st.session_state[provider_options_key] = updated_providers
                st.success(f"Added provider: {custom_provider}")
            else:
                st.info("That provider is already available.")
    model_name = st.text_input(
        "Model",
        value="gpt-4o-mini",
        key=f"{rag_key_prefix}_model",
    )
    forecast_horizon = st.number_input(
        "Forecast horizon (years)",
        min_value=1,
        max_value=50,
        value=10,
        key=f"{rag_key_prefix}_forecast_horizon",
    )
    ml_methods = st.multiselect(
        "Machine learning method",
        ["Linear regression", "Compound annual growth", "ARIMA", "Prophet", "LSTM"],
        default=["Linear regression"],
        key=f"{rag_key_prefix}_ml_methods",
    )
    generative_features = st.multiselect(
        "Generative features",
        ["Executive summary", "Risk review", "Cash flow highlights", "ESG review", "Market overview"],
        default=["Executive summary", "Risk review", "Cash flow highlights"],
        key=f"{rag_key_prefix}_gen_features",
    )
    api_key = st.text_input(
        "API key",
        type="password",
        key=f"{rag_key_prefix}_api_key",
    )
    if st.button("Save AI configuration", key=f"{rag_key_prefix}_save_config"):
        st.session_state["rag_ai_config"] = {
            "enable_ai": enable_ai,
            "provider": custom_provider or provider,
            "model": model_name,
            "forecast_horizon": forecast_horizon,
            "ml_methods": ml_methods,
            "generative_features": generative_features,
            "api_key_set": bool(api_key),
        }
        st.success("AI configuration saved.")

    st.markdown("## AI Insights")
    model_cfg = st.session_state.get("model_config")
    valuation_result = st.session_state.get("valuation_result")
    portfolio = st.session_state.get("portfolio")

    if "rag_snapshot" not in st.session_state:
        st.session_state["rag_snapshot"] = _default_rag_advisory_inputs()

    snapshot_state = st.session_state["rag_snapshot"]
    if model_cfg is not None and valuation_result is not None and not snapshot_state.get("scenarios"):
        snapshot_state["scenarios"] = _default_scenario_pack(portfolio)

    if st.button("Refresh advisory scenarios from latest model", key=f"{rag_key_prefix}_refresh_snapshot"):
        if model_cfg is None or valuation_result is None:
            st.warning("Run the model workspace to seed advisory scenarios from the model.")
        else:
            snapshot_state["scenarios"] = _default_scenario_pack(portfolio)

    snapshot_payload = _build_bankable_snapshot_payload(
        project_id,
        model_cfg,
        valuation_result,
        portfolio,
        workbook_hash=snapshot_state.get("workbook_hash"),
        advisory_inputs=snapshot_state,
    )
    locked_snapshot = snapshot_payload["financial_snapshot"]
    has_live_model = snapshot_payload["snapshot_source"] == "live_model_only"

    with st.expander("Locked export snapshot", expanded=False):
        st.caption(
            "Bankable exports use only these model-derived values. Manual overrides are disabled in the export path."
        )
        snapshot_rows = pd.DataFrame(
            [
                {"Metric": "Currency", "Value": locked_snapshot.get("currency")},
                {"Metric": "NPV", "Value": locked_snapshot.get("npv")},
                {"Metric": "IRR", "Value": locked_snapshot.get("irr")},
                {"Metric": "Minimum DSCR", "Value": locked_snapshot.get("dscr_min")},
                {"Metric": "Payback (years)", "Value": locked_snapshot.get("payback_years")},
                {"Metric": "Total capex", "Value": locked_snapshot.get("capex_total")},
                {"Metric": "Annual opex", "Value": locked_snapshot.get("opex_annual")},
                {"Metric": "Annual revenue", "Value": locked_snapshot.get("revenue_annual")},
            ]
        )
        st.dataframe(snapshot_rows, hide_index=True, use_container_width=True)
        if not has_live_model:
            st.warning("Run the model workspace to generate a locked snapshot for lender/investor exports.")

    with st.expander("Advisory scenario inputs (not exported)", expanded=False):
        st.caption(
            "These notes can support AI drafting and internal discussion, but they are excluded from bankable exports."
        )
        scenarios_df = pd.DataFrame(snapshot_state.get("scenarios") or [])
        scenarios_df = st.data_editor(
            scenarios_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "name": st.column_config.TextColumn("Scenario"),
                "npv": st.column_config.NumberColumn("NPV"),
                "irr": st.column_config.NumberColumn("IRR"),
            },
            key=f"{rag_key_prefix}_scenarios_editor",
        )
        snapshot_state["scenarios"] = scenarios_df.to_dict(orient="records")
        snapshot_state["notes"] = st.text_area(
            "Advisory notes",
            value=str(snapshot_state.get("notes") or ""),
            key=f"{rag_key_prefix}_advisory_notes",
        )

    has_uploads = bool(uploads)
    has_indexed = bool(st.session_state.get("rag_last_ingest"))
    insight_cols = st.columns(3)
    if insight_cols[0].button(
        "Index documents",
        key=f"{rag_key_prefix}_index_docs",
        disabled=not has_uploads,
    ):
        files = [("files", (u.name, u.getvalue(), u.type or "application/octet-stream")) for u in uploads]
        try:
            response = requests.post(
                f"{rag_host.rstrip('/')}/ingest",
                params={"project_id": project_id},
                files=files,
                timeout=120,
            )
            response.raise_for_status()
            st.session_state["rag_last_ingest"] = response.json()
            st.success(response.json())
        except requests.RequestException as exc:
            if isinstance(exc, requests.ConnectionError):
                st.warning(
                    "RAG service unreachable. Start the service or update RAG_HOST to a reachable URL."
                )
            else:
                st.error(f"Failed to ingest files: {exc}")
    if not has_uploads:
        st.caption("Upload reference documents to enable indexing.")

    if insight_cols[1].button("Clear indexed documents", key=f"{rag_key_prefix}_clear_index"):
        st.session_state.pop("rag_last_ingest", None)
        st.session_state.pop("rag_last_report", None)
        st.info("Local index metadata cleared. Clear the backend index from the service if needed.")

    if insight_cols[2].button(
        "Run AI insights",
        key=f"{rag_key_prefix}_run_ai",
        disabled=not has_indexed,
    ):
        outline = _rag_section_outline()
        try:
            response = requests.post(
                f"{rag_host.rstrip('/')}/generate",
                json={"project_id": project_id, "section_outline": outline},
                timeout=180,
            )
            response.raise_for_status()
            st.session_state["rag_last_report"] = response.json()
            st.success(response.json())
        except requests.RequestException as exc:
            if isinstance(exc, requests.ConnectionError):
                st.warning(
                    "RAG service unreachable. Start the service or update RAG_HOST to a reachable URL."
                )
            else:
                st.error(f"Failed to run AI insights: {exc}")
    if not has_indexed:
        st.caption("Index documents before running AI insights.")

    question = st.text_input(
        "Ask a question",
        key=f"{rag_key_prefix}_question",
    )
    if st.button("Search", key=f"{rag_key_prefix}_search"):
        if not question:
            st.warning("Enter a question to search.")
        else:
            st.info("Search requires a backend endpoint (e.g. /search). Configure it to enable results.")

    st.markdown("## Business Plan Downloads")
    st.caption(
        "Generate a consolidated business plan bundle that includes the full financial report and snapshot."
    )
    if st.button(
        "Prepare business plan bundle",
        key=f"{rag_key_prefix}_bundle",
        disabled=not has_live_model,
    ):
        st.session_state["rag_bundle_ready"] = True
        st.success("Bundle ready. Download below.")
    if not has_live_model:
        st.caption("Run the model workspace first. Business plan exports now require locked model outputs.")

    if st.session_state.get("rag_bundle_ready"):
        if not has_live_model:
            st.warning("Locked export data is unavailable. Re-run the model before preparing a business plan bundle.")
            st.session_state["rag_bundle_ready"] = False
            return
        valuation_result = st.session_state.get("valuation_result")
        model_cfg = st.session_state.get("model_config")
        perf_df = None
        position_df = None
        cash_flow_df = None
        cons_df = None
        if valuation_result is not None and model_cfg is not None:
            cons = valuation_result.consolidated
            cons_df = cons.copy()
            perf_df, position_df, cash_flow_df = _compute_financial_statements(cons, model_cfg)
        cash_flow_df = _apply_debt_schedule(
            cash_flow_df,
            st.session_state.get("debt_schedule_table"),
            float(st.session_state.get("debt_interest_rate", 0.0)),
        )
        bundle_payload = {
            "snapshot": snapshot_payload,
            "ai_config": st.session_state.get("rag_ai_config", {}),
            "last_report": st.session_state.get("rag_last_report", {}),
            "financial_performance": perf_df,
            "financial_position": position_df,
            "cash_flows": cash_flow_df,
            "financial_statements": cons_df,
        }
        chart_tables = _build_chart_tables(
            st.session_state.get("valuation_result"),
            st.session_state.get("model_config"),
            st.session_state.get("portfolio"),
        )
        monte_carlo_df = _build_monte_carlo_results(snapshot_payload["financial_snapshot"])
        if not monte_carlo_df.empty:
            chart_tables["monte_carlo_results"] = monte_carlo_df
        export_payload = _build_export_payload(
            bundle_payload,
            analytics_df=chart_tables.get("advanced_analytics_report"),
        )
        export_payload["chart_tables"] = chart_tables
        export_payload["advanced_analytics_narrative"] = _build_advanced_analytics_narrative(
            chart_tables.get("advanced_analytics_report")
        )
        export_payload["extended_analytics_sections"] = _build_extended_analytics_sections(chart_tables)
        export_payload["chart_images"] = _build_chart_images(chart_tables)
        export_buffers, export_warnings = _build_export_buffers(export_payload)

        for warning in export_warnings:
            st.warning(warning)

        _render_export_downloads(
            export_buffers,
            project_id=project_id,
            rag_key_prefix=rag_key_prefix,
        )

def main() -> None:
    st.set_page_config(
        page_title="Biotech Financial Model",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    _inject_app_theme()
    _render_model_hero()

    model_cfg: ModelConfig | None = None
    portfolio: Portfolio | None = None
    valuation_result = None

    (
        config_tab,
        financial_tab,
        analytics_tab,
        dashboard_scenario_tab,
        vc_tab,
        rag_tab,
    ) = st.tabs(
        [
            "Model configuration",
            "Financial statements",
            "Advanced analytics",
            "Dashboard & Scenarios",
            "VC helper",
            "RAG Assistant",
        ]
    )

    with config_tab:
        with _section_block(
            "Model assumptions",
            heading_level=2,
            caption="Configure stage templates, core model settings, financing assumptions, and governance inputs.",
        ):

            with st.expander("Start here: guided setup", expanded=True):
                st.markdown(
                    "\n".join(
                        [
                            "1. Confirm the pipeline stage taxonomy.",
                            "2. Set general assumptions (years, tax, working capital).",
                            "3. Load a template or enter product assumptions.",
                            "4. Run the model and review dashboard + scenarios.",
                        ]
                    )
                )
                st.caption("Use this checklist to keep inputs consistent and audit-ready.")

            with st.expander("Pipeline stage templates", expanded=True):
                st.markdown(f"**Use {' → '.join(STAGE_SEQUENCE)}**")
                selected_stage = st.selectbox(
                    "Select a stage template to apply",
                    options=STAGE_SEQUENCE,
                    key="pipeline_stage_template_select",
                )
                st.markdown("**Selected template**")
                st.markdown(f"- {selected_stage}")
                st.caption("Select a stage to align asset setup and scenario inputs.")
                visibility = _stage_visibility_flags(selected_stage)
                show_forecast_ramp = visibility["show_forecast_ramp"]
                show_vaccine_sales = visibility["show_vaccine_sales"]
                show_uses_sources = visibility["show_uses_sources"]
                show_relevant_market_sizes = visibility["show_relevant_market_sizes"]
                show_market_size_estimation = visibility["show_market_size_estimation"]
                show_revenue_estimation = visibility["show_revenue_estimation"]
                show_cost_assumptions = visibility["show_cost_assumptions"]
                show_royalties = visibility["show_royalties"]
                show_market_share = visibility["show_market_share"]
                show_rd = visibility["show_rd"]
                show_capex = visibility["show_capex"]

            with _section_block("Stage-to-schedule mapping", heading_level=3):
                st.caption(
                    "Define default schedule assumptions per stage. These defaults can automatically "
                    "populate product assumptions when the stage changes. Stage durations are used to "
                    "derive time-to-market and to build annual transition probability curves."
                )
                audit_owner = st.text_input(
                    "Mapping updated by",
                    value=st.session_state.get("stage_mapping_audit_owner", "Finance"),
                    key="stage_mapping_audit_owner",
                )
                if "stage_mapping_audit_log" not in st.session_state:
                    st.session_state["stage_mapping_audit_log"] = []
                auto_apply_defaults = st.checkbox(
                    "Auto-apply stage defaults to product assumptions",
                    value=st.session_state.get("stage_mapping_auto_apply", True),
                    key="stage_mapping_auto_apply",
                )
                overwrite_defaults = st.checkbox(
                    "Override existing values when applying defaults",
                    value=st.session_state.get("stage_mapping_overwrite", False),
                    key="stage_mapping_overwrite",
                )
                mapping_df = _ensure_table_state(
                    "stage_schedule_mapping",
                    _default_stage_schedule_mapping,
                )
                previous_mapping = mapping_df.copy()
                with st.expander("Edit stage mapping", expanded=True):
                    st.caption(
                        "Use a structured editor for one stage at a time. "
                        "Time to market is derived from the stage-duration inputs before save."
                    )
                    editor_stage_key = "stage_mapping_editor_stage"
                    editor_revision_key = "stage_mapping_editor_revision"
                    editor_flash_key = "stage_mapping_editor_flash"
                    flash_message = st.session_state.pop(editor_flash_key, None)
                    if flash_message:
                        level, message = flash_message
                        if level == "success":
                            st.success(message)
                        elif level == "warning":
                            st.warning(message)
                        else:
                            st.info(message)

                    st.markdown('<div class="stage-mapping-editor-controls"></div>', unsafe_allow_html=True)
                    control_cols = st.columns([2.2, 1.15, 1.0, 3.65])
                    selected_stage = control_cols[0].selectbox(
                        "Select stage",
                        options=STAGE_OPTIONS,
                        key="stage_mapping_selected_stage",
                    )
                    with control_cols[1]:
                        st.caption("Open structured editor")
                        edit_clicked = st.button(
                            "Edit",
                            key="stage_mapping_edit_open",
                            type="primary",
                            use_container_width=True,
                        )
                    active_stage = st.session_state.get(editor_stage_key)
                    if edit_clicked:
                        st.session_state[editor_stage_key] = selected_stage
                        st.session_state[editor_revision_key] = int(
                            st.session_state.get(editor_revision_key, 0)
                        ) + 1
                        st.rerun()
                    if control_cols[2].button(
                        "Close",
                        key="stage_mapping_edit_close",
                        disabled=not active_stage,
                    ):
                        st.session_state.pop(editor_stage_key, None)
                        st.session_state[editor_revision_key] = int(
                            st.session_state.get(editor_revision_key, 0)
                        ) + 1
                        st.session_state[editor_flash_key] = ("info", "Edit stage mapping closed.")
                        st.rerun()
                    if active_stage:
                        control_cols[3].caption(
                            f"Editing {active_stage}. Save commits the row; discard reverts the staged changes."
                        )
                    else:
                        control_cols[3].caption(
                            "Select a stage and click Edit to open the structured editor."
                        )

                    active_stage = st.session_state.get(editor_stage_key)
                    if active_stage:
                        row_mask = (
                            mapping_df["Stage"].astype(str).map(normalize_stage_label)
                            == normalize_stage_label(active_stage)
                        )
                        if not row_mask.any():
                            st.warning("Selected stage not found in the mapping table.")
                        else:
                            row_idx = mapping_df.index[row_mask][0]
                            base_row = mapping_df.loc[row_idx].copy()
                            stage_label = str(base_row.get("Stage") or active_stage)
                            revision = int(st.session_state.get(editor_revision_key, 0))
                            updates: Dict[str, float | int | str] = {"Stage": stage_label}

                            st.markdown(f"**Editing {stage_label}**")
                            core_cols = st.columns(3)
                            with core_cols[0]:
                                updates["Success Probability %"] = st.number_input(
                                    "Success Probability %",
                                    min_value=0.0,
                                    max_value=100.0,
                                    value=float(base_row.get("Success Probability %", 0.0) or 0.0),
                                    step=1.0,
                                    key=_stage_mapping_input_key(stage_label, revision, "Success Probability %"),
                                )
                            with core_cols[1]:
                                updates["Sales ramp length (years)"] = st.number_input(
                                    "Sales ramp length (years)",
                                    min_value=0,
                                    value=int(base_row.get("Sales ramp length (years)", 0) or 0),
                                    step=1,
                                    key=_stage_mapping_input_key(stage_label, revision, "Sales ramp length (years)"),
                                )
                            with core_cols[2]:
                                current_shape = base_row.get("Ramp shape", RAMP_SHAPE_OPTIONS[0])
                                if current_shape not in RAMP_SHAPE_OPTIONS:
                                    current_shape = RAMP_SHAPE_OPTIONS[0]
                                updates["Ramp shape"] = st.selectbox(
                                    "Ramp shape",
                                    options=RAMP_SHAPE_OPTIONS,
                                    index=RAMP_SHAPE_OPTIONS.index(current_shape),
                                    key=_stage_mapping_input_key(stage_label, revision, "Ramp shape"),
                                )

                            funding_cols = st.columns(2)
                            with funding_cols[0]:
                                updates["R&D remaining pre-launch (USD)"] = st.number_input(
                                    "R&D remaining pre-launch (USD)",
                                    min_value=0.0,
                                    value=float(base_row.get("R&D remaining pre-launch (USD)", 0.0) or 0.0),
                                    step=1_000_000.0,
                                    key=_stage_mapping_input_key(
                                        stage_label,
                                        revision,
                                        "R&D remaining pre-launch (USD)",
                                    ),
                                )
                            with funding_cols[1]:
                                updates["R&D annual post-launch (USD/year)"] = st.number_input(
                                    "R&D annual post-launch (USD/year)",
                                    min_value=0.0,
                                    value=float(base_row.get("R&D annual post-launch (USD/year)", 0.0) or 0.0),
                                    step=1_000_000.0,
                                    key=_stage_mapping_input_key(
                                        stage_label,
                                        revision,
                                        "R&D annual post-launch (USD/year)",
                                    ),
                                )

                            with st.expander("Stage durations", expanded=True):
                                st.caption(
                                    "These durations drive the derived time to market for the selected stage."
                                )
                                duration_cols = st.columns(3)
                                for idx, col in enumerate(STAGE_DURATION_COLUMNS):
                                    with duration_cols[idx % 3]:
                                        updates[col] = st.number_input(
                                            col,
                                            min_value=0,
                                            value=int(base_row.get(col, 0) or 0),
                                            step=1,
                                            key=_stage_mapping_input_key(stage_label, revision, col),
                                        )

                            with st.expander("Transition probabilities", expanded=False):
                                trans_cols = st.columns(3)
                                for idx, col in enumerate(STAGE_TRANSITION_COLUMNS):
                                    with trans_cols[idx % 3]:
                                        updates[col] = st.number_input(
                                            col,
                                            min_value=0.0,
                                            max_value=100.0,
                                            value=float(base_row.get(col, 0.0) or 0.0),
                                            step=1.0,
                                            key=_stage_mapping_input_key(stage_label, revision, col),
                                        )
                                annual_cols = st.columns(3)
                                for idx, col in enumerate(STAGE_TRANSITION_ANNUAL_COLUMNS):
                                    with annual_cols[idx % 3]:
                                        updates[col] = st.number_input(
                                            col,
                                            min_value=0.0,
                                            max_value=100.0,
                                            value=float(base_row.get(col, 0.0) or 0.0),
                                            step=1.0,
                                            key=_stage_mapping_input_key(stage_label, revision, col),
                                        )

                            with st.expander("R&D and CAPEX allocation", expanded=False):
                                rd_cols = st.columns(3)
                                for idx, col in enumerate(STAGE_COST_WEIGHT_COLUMNS):
                                    with rd_cols[idx % 3]:
                                        updates[col] = st.number_input(
                                            col,
                                            min_value=0.0,
                                            max_value=100.0,
                                            value=float(base_row.get(col, 0.0) or 0.0),
                                            step=1.0,
                                            key=_stage_mapping_input_key(stage_label, revision, col),
                                        )
                                capex_cols = st.columns(3)
                                for idx, col in enumerate(STAGE_CAPEX_WEIGHT_COLUMNS):
                                    with capex_cols[idx % 3]:
                                        updates[col] = st.number_input(
                                            col,
                                            min_value=0.0,
                                            max_value=100.0,
                                            value=float(base_row.get(col, 0.0) or 0.0),
                                            step=1.0,
                                            key=_stage_mapping_input_key(stage_label, revision, col),
                                        )

                            with st.expander("Milestones", expanded=False):
                                milestone_cols = st.columns(2)
                                for idx, col in enumerate(STAGE_MILESTONE_COLUMNS):
                                    with milestone_cols[idx % 2]:
                                        updates[col] = st.number_input(
                                            col,
                                            min_value=0.0,
                                            value=float(base_row.get(col, 0.0) or 0.0),
                                            step=1_000_000.0,
                                            key=_stage_mapping_input_key(stage_label, revision, col),
                                        )

                            candidate_row = _build_stage_mapping_candidate_row(base_row, updates)
                            row_warnings = _stage_mapping_row_warnings(mapping_df, row_idx, candidate_row)

                            derived_cols = st.columns([1.2, 1.2, 3.6])
                            derived_cols[0].metric(
                                "Derived time to market (years)",
                                int(candidate_row.get("Time to market (years)", 0) or 0),
                            )
                            derived_cols[1].metric(
                                "Saved value",
                                int(base_row.get("Time to market (years)", 0) or 0),
                            )
                            derived_cols[2].caption(
                                "The saved time-to-market value is computed from the stage-duration inputs above."
                            )

                            if row_warnings:
                                st.warning(
                                    "Scientific/commercial check for this stage: review the items below before saving."
                                )
                                for warning in row_warnings:
                                    st.write(f"- {warning}")
                            else:
                                st.success("No row-level scientific/commercial warnings for this stage.")

                            preview_cols = [
                                "Stage",
                                "Success Probability %",
                                "Time to market (years)",
                                "Sales ramp length (years)",
                                "Ramp shape",
                                "R&D remaining pre-launch (USD)",
                                "R&D annual post-launch (USD/year)",
                            ]
                            st.markdown("**Save preview**")
                            st.dataframe(
                                pd.DataFrame([candidate_row.reindex(preview_cols)]),
                                hide_index=True,
                                use_container_width=True,
                            )

                            action_cols = st.columns([1.2, 1.2, 4.6])
                            if action_cols[0].button("Save stage", key="stage_mapping_edit_save"):
                                updated_mapping = mapping_df.copy()
                                for col in updated_mapping.columns:
                                    updated_mapping.at[row_idx, col] = candidate_row.get(
                                        col,
                                        updated_mapping.at[row_idx, col],
                                    )
                                mapping_df = updated_mapping
                                st.session_state["stage_schedule_mapping"] = mapping_df
                                if not mapping_df.equals(previous_mapping):
                                    st.session_state["stage_mapping_audit_log"].append(
                                        {
                                            "timestamp": pd.Timestamp.utcnow().isoformat(),
                                            "updated_by": audit_owner,
                                            "note": f"Stage mapping updated: {stage_label}",
                                        }
                                    )
                                st.session_state.pop(editor_stage_key, None)
                                st.session_state[editor_revision_key] = revision + 1
                                st.session_state[editor_flash_key] = (
                                    "success",
                                    f"Saved {stage_label} stage assumptions.",
                                )
                                st.rerun()
                            if action_cols[1].button("Discard edits", key="stage_mapping_edit_discard"):
                                st.session_state.pop(editor_stage_key, None)
                                st.session_state[editor_revision_key] = revision + 1
                                st.session_state[editor_flash_key] = (
                                    "info",
                                    f"Discarded edits for {stage_label}.",
                                )
                                st.rerun()

                st.info("Editing happens above. Use Edit stage mapping to make changes before reviewing the summary.")
                with st.expander("Full mapping table (summary)", expanded=False):
                    st.caption("Read-only summary. Use Edit stage mapping above to make changes.")
                    st.dataframe(
                        mapping_df,
                        hide_index=True,
                        use_container_width=True,
                    )
                mapping_warnings = _stage_mapping_sanity_checks(mapping_df)
                if mapping_warnings:
                    st.warning(
                        "Scientific/commercial check: please review the items below so stage inputs "
                        "remain realistic and internally consistent."
                    )
                    for warning in mapping_warnings:
                        st.write(f"- {warning}")
                st.session_state["stage_schedule_mapping"] = mapping_df
                with st.expander("Mapping audit trail", expanded=False):
                    audit_log = st.session_state.get("stage_mapping_audit_log", [])
                    if audit_log:
                        st.dataframe(pd.DataFrame(audit_log), use_container_width=True)
                    else:
                        st.caption("No mapping changes recorded yet.")

            with st.expander("General assumptions", expanded=True):
                col1, col2, col3 = st.columns(3)
                with col1:
                    first_year = st.number_input("First forecast year", value=2024)
                    n_years = st.number_input("Number of years", min_value=5, max_value=40, value=25)
                    currency = st.text_input("Currency", value="USD")
                with col2:
                    tax_rate = st.slider("Tax rate", min_value=0.0, max_value=0.35, value=0.25)
                    wc_pct = st.slider("Working capital (% sales)", 0.0, 0.3, 0.08)
                with col3:
                    inflation = st.number_input("Inflation assumption", value=0.02, min_value=0.0, max_value=0.25, step=0.005)
                    base_fx = st.text_input("Reporting FX pair", value="USD/EUR")
                auto_sync_vaccine_sales = st.checkbox(
                    "Rebuild Vaccine Sales table when assumptions change",
                    value=True,
                )
                st.caption("Set the macro baseline for the consolidated forecast and disclosures.")

            if show_forecast_ramp:
                with st.expander("Forecast assumptions", expanded=True):
                    ramp_df = _render_schedule_editor("Sales ramp schedule", "sales_ramp_schedule")
                    ramp_df = ramp_df.sort_values("Year offset")
                    if ramp_df.empty:
                        st.warning("Ramp schedule empty. Reverting to default values.")
                        ramp = _default_ramp_schedule()["Ramp factor"].tolist()
                    else:
                        ramp = ramp_df["Ramp factor"].astype(float).tolist()
                    st.caption("Ramp factors feed revenue build-ups across every product.")
            else:
                ramp = _default_ramp_schedule()["Ramp factor"].tolist()

            if show_vaccine_sales:
                with st.expander("Vaccine sales"):
                    assumptions_changed = (
                        st.session_state.get("vaccine_sales_first_year") != int(first_year)
                        or st.session_state.get("vaccine_sales_n_years") != int(n_years)
                    )
                    if auto_sync_vaccine_sales and assumptions_changed:
                        st.session_state["vaccine_sales_table"] = _default_vaccine_sales_table(
                            int(first_year),
                            int(n_years),
                        )
                    st.session_state["vaccine_sales_first_year"] = int(first_year)
                    st.session_state["vaccine_sales_n_years"] = int(n_years)
                    vaccine_df = _render_product_assumption_table(
                        session_key="vaccine_sales_table",
                        default_factory=lambda: _default_vaccine_sales_table(int(first_year), int(n_years)),
                        blank_row_factory=lambda df: _blank_vaccine_sales_row(df, int(first_year)),
                        id_column=None,
                        name_column="Vaccine name",
                        column_config={
                            "ID_vaccine": st.column_config.TextColumn("ID", help="Vaccine ID"),
                            "Vaccine name": st.column_config.TextColumn("Vaccine name"),
                            "Year": st.column_config.NumberColumn("Year", step=1),
                            "Doses (M)": st.column_config.NumberColumn("Doses (M)", min_value=0.0, step=0.5),
                            "Price per dose": st.column_config.NumberColumn(
                                "Price per dose", min_value=0.0, step=1.0
                            ),
                        },
                    )
                    vaccine_df = _recompute_vaccine_sales_implied_revenue(vaccine_df)
                    st.session_state["vaccine_sales_table"] = vaccine_df
                    with _section_block("Yearly Increment Helper", heading_level=4):
                        def _filter(df: pd.DataFrame, selected_id: Optional[str], start_year: int) -> pd.Series:
                            if selected_id is None:
                                return pd.Series([False] * len(df), index=df.index)
                            year_values = pd.to_numeric(df["Year"], errors="coerce").fillna(0).astype(int)
                            return (df["ID_vaccine"].astype(str) == str(selected_id)) & (year_values >= int(start_year))

                        if {"ID_vaccine", "Year"}.issubset(vaccine_df.columns):
                            vaccine_df = _render_yearly_increment_helper(
                                section_key="vaccine_sales",
                                df=vaccine_df,
                                year_column="Year",
                                target_columns=["Doses (M)", "Price per dose"],
                                filter_builder=_filter,
                                id_column="ID_vaccine",
                                id_label="Vaccine ID",
                                start_year_label="Start year",
                                start_year_default=int(first_year),
                                periods_default=5,
                                increment_default=1.0,
                                allow_compound=True,
                                create_missing_rows=False,
                                base_value_mode="first_row",
                            )
                            vaccine_df = _recompute_vaccine_sales_implied_revenue(vaccine_df)
                            st.session_state["vaccine_sales_table"] = vaccine_df
                        else:
                            st.caption("Add vaccine IDs and years to use the helper.")
                    sync_sales_to_revenue = st.checkbox(
                        "Sync vaccine sales to revenue estimation",
                        value=True,
                        key="sync_vaccine_sales_to_revenue",
                    )
                    if sync_sales_to_revenue and not vaccine_df.empty:
                        revenue_table = st.session_state.get(
                            "vaccine_revenue_table",
                            _default_vaccine_revenue_table(),
                        ).copy()
                        if {
                            "Patent customers per year",
                            "Patent price (USD/customer)",
                            "ID_vaccine",
                        }.issubset(revenue_table.columns):
                            price_series = _coerce_numeric(
                                revenue_table["Patent price (USD/customer)"], 0.0
                            ).replace(0, np.nan)
                            revenue_table["ID_vaccine"] = revenue_table["ID_vaccine"].astype(str)
                            sales_by_vaccine = (
                                vaccine_df.groupby("ID_vaccine")["Implied revenue"].mean().to_dict()
                            )
                            desired_targets = revenue_table["ID_vaccine"].map(sales_by_vaccine)
                            if "Vaccine name" in vaccine_df.columns and "Vaccine name" in revenue_table.columns:
                                sales_by_name = (
                                    vaccine_df.groupby("Vaccine name")["Implied revenue"].mean().to_dict()
                                )
                                name_targets = revenue_table["Vaccine name"].map(sales_by_name)
                                desired_targets = desired_targets.fillna(name_targets)
                            desired_targets = desired_targets.fillna(0.0)
                            revenue_table["Patent customers per year"] = (
                                desired_targets / price_series
                            ).fillna(0.0)
                            st.session_state["vaccine_revenue_table"] = revenue_table
                    st.metric(f"{int(n_years)}-year vaccine sales", f"{vaccine_df['Implied revenue'].sum():,.0f}")
                    base_products = st.session_state.get("product_table", _default_products())
                    st.session_state["product_table"] = _sync_vaccine_sales_products(
                        base_products,
                        vaccine_df,
                    )

            funding_required = float(st.session_state.get("funding_required", 250_000_000.0))
            planned_new_equity = float(st.session_state.get("planned_new_equity", 200_000_000.0))
            uses_total = float(st.session_state.get("uses_total", 0.0))
            sources_total = float(st.session_state.get("sources_total", 0.0))
            burn_total = float(st.session_state.get("burn_total", 0.0))
            wc_total = float(st.session_state.get("wc_total", 0.0))

            if show_uses_sources:
                with st.expander("Uses and sources of funds"):
                    auto_funding_required = st.checkbox(
                        "Auto-calculate funding required from model outputs",
                        value=True,
                    )
                    uses_col, sources_col = st.columns(2)
                    with uses_col:
                        st.markdown("**Uses**")
                        uses_df = _render_product_assumption_table(
                            session_key="uses_table",
                            default_factory=_default_uses_table,
                            blank_row_factory=_blank_use_row,
                            id_column=None,
                            name_column="Item",
                            column_config={
                                "ID_vaccine": st.column_config.TextColumn("ID", help="Vaccine ID"),
                                "Vaccine name": st.column_config.TextColumn("Vaccine name"),
                                "Amount": st.column_config.NumberColumn("Amount", step=1_000_000.0),
                            },
                        )
                        uses_total = float(uses_df.get("Amount", pd.Series(dtype=float)).sum())
                        st.session_state["uses_total"] = uses_total
                        st.metric("Total uses", f"{uses_total:,.0f}")
                        uses_warnings = []
                        if (pd.to_numeric(uses_df.get("Amount", pd.Series(dtype=float)), errors="coerce") < 0).any():
                            uses_warnings.append("Uses contain negative amounts; use positive values.")
                        _render_section_warnings("Uses", uses_warnings)
                        if {"ID_vaccine", "Vaccine name", "Amount"}.issubset(uses_df.columns):
                            uses_by_vaccine = (
                                uses_df.groupby(["ID_vaccine", "Vaccine name"], dropna=False)["Amount"]
                                .sum()
                                .reset_index()
                            )
                            st.dataframe(
                                uses_by_vaccine.style.format({"Amount": "{:,.0f}"}),
                                use_container_width=True,
                            )
                    with sources_col:
                        st.markdown("**Sources**")
                        sources_df = _render_product_assumption_table(
                            session_key="sources_table",
                            default_factory=_default_sources_table,
                            blank_row_factory=_blank_source_row,
                            id_column=None,
                            name_column="Item",
                            column_config={
                                "Amount": st.column_config.NumberColumn("Amount", step=1_000_000.0),
                            },
                        )
                        sources_other_total = 0.0
                        if {"Item", "Amount"}.issubset(sources_df.columns):
                            source_items = sources_df["Item"].astype(str).str.strip().str.lower()
                            sources_other_total = float(
                                sources_df.loc[source_items != "new equity", "Amount"]
                                .apply(pd.to_numeric, errors="coerce")
                                .fillna(0.0)
                                .sum()
                            )
                        debt_draw_total = 0.0
                        debt_schedule_df = st.session_state.get("debt_schedule_table")
                        if debt_schedule_df is not None and "Debt drawdowns" in debt_schedule_df.columns:
                            debt_draw_total = float(
                                pd.to_numeric(debt_schedule_df["Debt drawdowns"], errors="coerce")
                                .fillna(0.0)
                                .sum()
                            )
                        sources_other_total += debt_draw_total
                        valuation_result = st.session_state.get("valuation_result")
                        burn_total = 0.0
                        wc_total = 0.0
                        if valuation_result is not None:
                            cons = valuation_result.consolidated
                            if "fcff_after_wc" in cons.columns:
                                burn_total = float((-cons["fcff_after_wc"].clip(upper=0)).sum())
                            if "delta_wc" in cons.columns:
                                wc_total = float((-cons["delta_wc"].clip(upper=0)).sum())
                        st.session_state["burn_total"] = burn_total
                        st.session_state["wc_total"] = wc_total
                        derived_funding_required = uses_total + burn_total + wc_total
                        if auto_funding_required:
                            funding_required = float(derived_funding_required)
                            st.session_state["funding_required"] = funding_required
                        planned_new_equity = max(funding_required - sources_other_total, 0.0)
                        st.session_state["planned_new_equity"] = planned_new_equity
                        if {"Item", "Amount"}.issubset(sources_df.columns):
                            mask = sources_df["Item"].astype(str).str.strip().str.lower() == "new equity"
                            if mask.any():
                                sources_df.loc[mask, "Amount"] = planned_new_equity
                                st.session_state["sources_table"] = sources_df
                            elif planned_new_equity > 0:
                                sources_df.loc[len(sources_df)] = {
                                    "Item": "New equity",
                                    "Amount": planned_new_equity,
                                }
                                st.session_state["sources_table"] = sources_df
                        sources_table_total = float(sources_df.get("Amount", pd.Series(dtype=float)).sum())
                        sources_total = sources_table_total + debt_draw_total
                        st.session_state["sources_total"] = sources_total
                        st.metric("Total sources", f"{sources_total:,.0f}")
                        if debt_draw_total:
                            st.caption(
                                f"Includes scheduled debt drawdowns of {debt_draw_total:,.0f} from the debt schedule."
                            )
                        sources_warnings = []
                        if (pd.to_numeric(sources_df.get("Amount", pd.Series(dtype=float)), errors="coerce") < 0).any():
                            sources_warnings.append("Sources contain negative amounts; use positive values.")
                        _render_section_warnings("Sources", sources_warnings)
                    delta = sources_total - funding_required
                    st.info(f"Funding gap (sources - total funding required): {delta:,.0f}")

            if show_uses_sources:
                with st.expander("Debt schedule inputs", expanded=False):
                    debt_template = _default_debt_schedule(int(first_year), int(n_years))
                    debt_table_changed = (
                        st.session_state.get("debt_schedule_first_year") != int(first_year)
                        or st.session_state.get("debt_schedule_n_years") != int(n_years)
                    )
                    if debt_table_changed or "debt_schedule_table" not in st.session_state:
                        st.session_state["debt_schedule_table"] = debt_template
                    else:
                        st.session_state["debt_schedule_table"] = _align_table_to_template(
                            st.session_state.get("debt_schedule_table"),
                            debt_template,
                        )
                    st.session_state["debt_schedule_first_year"] = int(first_year)
                    st.session_state["debt_schedule_n_years"] = int(n_years)
                    debt_cols = st.columns(5)
                    with debt_cols[0]:
                        debt_interest_rate = st.number_input(
                            "Debt interest rate",
                            min_value=0.0,
                            max_value=1.0,
                            value=float(st.session_state.get("debt_interest_rate", 0.08)),
                            step=0.005,
                            format="%.3f",
                            key="debt_interest_rate",
                        )
                    with debt_cols[1]:
                        current_repayment_mode = str(
                            st.session_state.get(
                                "debt_repayment_mode",
                                DEBT_REPAYMENT_LABELS["straight_line"],
                            )
                            or DEBT_REPAYMENT_LABELS["straight_line"]
                        )
                        if current_repayment_mode in DEBT_REPAYMENT_LABELS:
                            current_repayment_mode = DEBT_REPAYMENT_LABELS[current_repayment_mode]
                        if current_repayment_mode not in DEBT_REPAYMENT_CODES:
                            current_repayment_mode = DEBT_REPAYMENT_LABELS["straight_line"]
                        if st.session_state.get("debt_repayment_mode") != current_repayment_mode:
                            st.session_state["debt_repayment_mode"] = current_repayment_mode
                        debt_repayment_mode_label = st.selectbox(
                            "Repayment mode",
                            options=list(DEBT_REPAYMENT_CODES.keys()),
                            index=list(DEBT_REPAYMENT_CODES.keys()).index(current_repayment_mode),
                            key="debt_repayment_mode",
                        )
                        debt_repayment_mode = DEBT_REPAYMENT_CODES[debt_repayment_mode_label]
                    with debt_cols[2]:
                        debt_grace_years = st.number_input(
                            "Grace years",
                            min_value=0,
                            max_value=int(n_years),
                            value=int(st.session_state.get("debt_grace_years", 0) or 0),
                            step=1,
                            key="debt_grace_years",
                        )
                    with debt_cols[3]:
                        debt_target_dscr = st.number_input(
                            "Target DSCR",
                            min_value=0.5,
                            max_value=5.0,
                            value=float(st.session_state.get("debt_target_dscr", 1.3) or 1.3),
                            step=0.05,
                            format="%.2f",
                            key="debt_target_dscr",
                        )
                    with debt_cols[4]:
                        minimum_cash_reserve = st.number_input(
                            "Minimum cash reserve",
                            min_value=0.0,
                            value=float(st.session_state.get("minimum_cash_reserve", 0.0) or 0.0),
                            step=1_000_000.0,
                            format="%0.0f",
                            key="minimum_cash_reserve",
                        )
                    debt_schedule_df = _render_product_assumption_table(
                        session_key="debt_schedule_table",
                        default_factory=lambda: _default_debt_schedule(int(first_year), int(n_years)),
                        blank_row_factory=lambda df: _blank_debt_schedule_row(
                            df,
                            int(first_year),
                            int(n_years),
                        ),
                        id_column=None,
                        name_column="Year",
                        column_config={
                            "Year": st.column_config.NumberColumn("Year", step=1),
                            "Debt drawdowns": st.column_config.NumberColumn(
                                "Debt drawdowns", step=1_000_000.0
                            ),
                            "Manual debt repayments": st.column_config.NumberColumn(
                                "Manual debt repayments",
                                step=1_000_000.0,
                            ),
                        },
                    )
                    st.session_state["debt_schedule_table"] = debt_schedule_df
                    st.caption(
                        "Drawdowns stay manual. Repayments follow the selected mode; the manual repayment column is used only when Manual schedule is selected."
                    )
                    debt_warnings = []
                    if (pd.to_numeric(debt_schedule_df.get("Debt drawdowns", pd.Series(dtype=float)), errors="coerce") < 0).any():
                        debt_warnings.append("Debt drawdowns should be zero or positive.")
                    if (
                        pd.to_numeric(
                            debt_schedule_df.get("Manual debt repayments", pd.Series(dtype=float)),
                            errors="coerce",
                        )
                        < 0
                    ).any():
                        debt_warnings.append("Manual debt repayments should be zero or positive.")
                    if debt_interest_rate < 0 or debt_interest_rate > 1:
                        debt_warnings.append("Debt interest rate should be between 0% and 100%.")
                    if debt_repayment_mode == "manual" and (
                        pd.to_numeric(
                            debt_schedule_df.get("Manual debt repayments", pd.Series(dtype=float)),
                            errors="coerce",
                        )
                        .fillna(0.0)
                        .sum()
                        <= 0
                    ):
                        debt_warnings.append("Manual repayment mode requires at least one positive manual repayment.")
                    _render_section_warnings("Debt schedule", debt_warnings)
                    funding_gap = sources_total - funding_required
                    st.metric("Sources less funding required", f"{funding_gap:,.0f}")
                    if abs(funding_gap) > 1.0:
                        st.warning("Sources and total funding required are not yet reconciled.")
                    reconciliation = pd.DataFrame(
                        [
                            {"Component": "Uses total", "Amount": uses_total},
                            {"Component": "Cash burn (FCFF < 0)", "Amount": burn_total},
                            {"Component": "Working capital draw", "Amount": wc_total},
                            {"Component": "Funding required", "Amount": funding_required},
                            {"Component": "Total sources", "Amount": sources_total},
                            {"Component": "Minimum cash reserve", "Amount": minimum_cash_reserve},
                        ]
                    )
                    st.dataframe(reconciliation.style.format({"Amount": "{:,.0f}"}))

            with st.expander("Risk-adjusted DCF valuation method - assumptions"):
                col_a, col_b, col_c = st.columns(3)
                with col_a:
                    discount_rate = st.slider("Discount rate", min_value=0.02, max_value=0.30, value=0.10)
                with col_b:
                    discount_timing_label = st.selectbox(
                        "Discount timing",
                        options=list(DISCOUNT_TIMING_CODES.keys()),
                    )
                    discount_timing = DISCOUNT_TIMING_CODES[discount_timing_label]
                with col_c:
                    risk_buffer = st.number_input(
                        "Additional risk premium", min_value=0.0, max_value=0.20, value=0.0, step=0.01
                    )
                dcf_cols = st.columns(3)
                with dcf_cols[0]:
                    terminal_method_label = st.selectbox(
                        "Terminal method",
                        options=list(TERMINAL_METHOD_CODES.keys()),
                    )
                    terminal_method = TERMINAL_METHOD_CODES[terminal_method_label]
                with dcf_cols[1]:
                    ev_multiple = st.slider("Terminal EV/EBITDA multiple", 2.0, 30.0, 8.0)
                with dcf_cols[2]:
                    perpetuity_growth = st.slider(
                        "Perpetuity growth rate",
                        min_value=0.0,
                        max_value=0.08,
                        value=0.02,
                        step=0.005,
                    )
                opening_nol_balance = st.number_input(
                    "Opening NOL balance",
                    min_value=0.0,
                    value=float(st.session_state.get("opening_nol_balance", 0.0)),
                    step=5_000_000.0,
                    format="%0.0f",
                    key="opening_nol_balance",
                )
                st.caption(
                    "Discount rate + premium governs enterprise value. Terminal method, discount timing, "
                    "working capital unwind, and NOL usage are now explicit model mechanics."
                )

            with st.expander("Funding required"):
                funding_required = st.number_input(
                    "Total funding required",
                    value=float(st.session_state.get("funding_required", 250_000_000.0)),
                    step=5_000_000.0,
                    format="%0.0f",
                    key="funding_required",
                )

            with st.expander("Shareholders / Investors"):
                shareholder_template = _default_shareholders_table()
                if "shareholders_table" in st.session_state:
                    st.session_state["shareholders_table"] = _align_table_to_template(
                        st.session_state.get("shareholders_table"),
                        shareholder_template,
                    )
                shareholders_df = _render_product_assumption_table(
                    session_key="shareholders_table",
                    default_factory=_default_shareholders_table,
                    blank_row_factory=_blank_shareholder_row,
                    id_column=None,
                    name_column="Shareholder",
                    column_config={
                        "Security": st.column_config.SelectboxColumn(
                            "Security",
                            options=["Common", "Preferred", "Convertible note"],
                        ),
                        "Seniority": st.column_config.NumberColumn("Seniority", min_value=1, step=1),
                        "Ownership %": st.column_config.NumberColumn(
                            "Ownership %", min_value=0.0, max_value=1.0, step=0.01
                        ),
                        "Investment": st.column_config.NumberColumn("Investment", step=1_000_000.0),
                        "Liquidation preference (x)": st.column_config.NumberColumn(
                            "Liquidation preference (x)",
                            min_value=0.0,
                            step=0.1,
                        ),
                        "Participating preferred": st.column_config.CheckboxColumn(
                            "Participating preferred"
                        ),
                    },
                )
                investment = pd.to_numeric(
                    shareholders_df.get("Investment", pd.Series(dtype=float)), errors="coerce"
                ).fillna(0.0)
                planned_new_equity = float(st.session_state.get("planned_new_equity", 0.0))
                pre_money = float(investment.sum())
                post_money = max(pre_money + planned_new_equity, 1.0)
                if "Shareholder" in shareholders_df.columns:
                    trimmed = shareholders_df["Shareholder"].astype(str).str.strip().str.lower()
                    new_equity_mask = trimmed == "new equity round"
                    if new_equity_mask.any():
                        shareholders_df.loc[new_equity_mask, "Investment"] = planned_new_equity
                        shareholders_df.loc[new_equity_mask, "Security"] = "Preferred"
                        shareholders_df.loc[new_equity_mask, "Seniority"] = 1
                        shareholders_df.loc[new_equity_mask, "Liquidation preference (x)"] = 1.0
                        shareholders_df.loc[new_equity_mask, "Participating preferred"] = False
                    elif planned_new_equity > 0:
                        shareholders_df.loc[len(shareholders_df)] = {
                            "Shareholder": "New equity round",
                            "Security": "Preferred",
                            "Seniority": 1,
                            "Ownership %": planned_new_equity / post_money,
                            "Investment": planned_new_equity,
                            "Liquidation preference (x)": 1.0,
                            "Participating preferred": False,
                        }

                ownership = pd.to_numeric(
                    shareholders_df.get("Investment", pd.Series(dtype=float)), errors="coerce"
                ).fillna(0.0) / post_money
                shareholders_df["Ownership %"] = ownership
                st.session_state["shareholders_table"] = shareholders_df
                st.metric("Total ownership (post-money)", f"{shareholders_df['Ownership %'].sum():.0%}")
                st.dataframe(
                    shareholders_df.style.format(
                        {
                            "Ownership %": "{:.1%}",
                            "Investment": "{:,.0f}",
                            "Liquidation preference (x)": "{:.1f}",
                        }
                    )
                )
                st.caption(
                    "Diluted ownership is recalculated from invested capital and planned new equity; liquidation preference and seniority feed the exit waterfall after a run."
                )

            if show_relevant_market_sizes:
                with st.expander("Relevant market sizes"):
                    market_df = _render_product_assumption_table(
                        session_key="market_sizes_table",
                        default_factory=_default_market_sizes_table,
                        blank_row_factory=_blank_relevant_market_row,
                        id_column=None,
                        name_column="Segment",
                        column_config={
                            "Value": st.column_config.NumberColumn("Value", step=1_000_000.0),
                        },
                    )
                    market_warnings = []
                    if (pd.to_numeric(market_df.get("Value", pd.Series(dtype=float)), errors="coerce") <= 0).any():
                        market_warnings.append("Relevant market sizes should be greater than zero.")
                    _render_section_warnings("Relevant market sizes", market_warnings)

            with st.expander("New equity issued"):
                new_equity = st.number_input(
                    "Planned new equity",
                    value=planned_new_equity,
                    step=5_000_000.0,
                    format="%0.0f",
                    key="planned_new_equity",
                )

            with st.expander("Selectors"):
                selector_choices = st.multiselect(
                    "Tag this run with selectors", options=SELECTOR_OPTIONS, default=["Base case"]
                )
                st.write("Active selectors:", ", ".join(selector_choices) or "None")

            effective_discount_rate = float(min(0.40, discount_rate + risk_buffer))
            model_cfg = ModelConfig(
                first_year=int(first_year),
                n_years=int(n_years),
                currency=currency,
                discount_rate=effective_discount_rate,
                tax_rate=float(tax_rate),
                working_capital_pct_sales=float(wc_pct),
                ev_ebitda_multiple=float(ev_multiple),
                opening_nol_balance=float(opening_nol_balance),
                sales_ramp_factors=ramp,
                discount_timing=str(discount_timing),
                terminal_method=str(terminal_method),
                perpetuity_growth_rate=float(perpetuity_growth),
                unwind_working_capital=True,
            )

        with _section_block(
            "Product assumptions",
            heading_level=2,
            caption="Define development, commercial, cost, R&D, CAPEX, royalty, and market-share inputs by asset.",
        ):

            dev_df = _render_product_assumption_table(
                session_key="vaccine_development_table",
                default_factory=lambda: _default_vaccine_development_table(int(first_year)),
                blank_row_factory=lambda df: _blank_vaccine_development_row(df, int(first_year)),
                column_config={
                    "Stage": st.column_config.SelectboxColumn("Stage", options=STAGE_OPTIONS),
                    "Consolidation": st.column_config.CheckboxColumn("Consolidate", default=True),
                    "Success Probability %": st.column_config.NumberColumn(
                        "Success Probability %", min_value=0.0, max_value=100.0, step=1.0
                    ),
                },
            )
            entry_calc = _coerce_numeric(dev_df.get("First year forecast", pd.Series(dtype=float))) + _coerce_numeric(
                dev_df.get("Time to market", pd.Series(dtype=float))
            )
            if "Market entry year" not in dev_df.columns:
                dev_df["Market entry year"] = entry_calc
            else:
                missing_entry = dev_df["Market entry year"].isna()
                dev_df.loc[missing_entry, "Market entry year"] = entry_calc[missing_entry]
            if "End patent year" not in dev_df.columns:
                dev_df["End patent year"] = dev_df["Market entry year"] + _coerce_numeric(
                    dev_df.get("Patent duration years", pd.Series(dtype=float)), default=0
                ) - 1
            else:
                mask_patent = dev_df["End patent year"].isna()
                dev_df.loc[mask_patent, "End patent year"] = (
                    dev_df.loc[mask_patent, "Market entry year"]
                    + _coerce_numeric(
                        dev_df.loc[mask_patent, "Patent duration years"],
                        default=0,
                    )
                    - 1
                )
            st.session_state["vaccine_development_table"] = dev_df
            st.caption("Track each vaccine's readiness, probability of success, and patent end year.")

            if show_market_size_estimation:
                with st.expander("Vaccine market size estimation", expanded=True):
                    market_size_df = _render_product_assumption_table(
                        session_key="market_size_estimation",
                        default_factory=_default_market_size_estimation_table,
                        blank_row_factory=_blank_market_size_row,
                    )
                    market_size = _coerce_numeric(
                        market_size_df.get("Market size (# customers)", pd.Series(dtype=float))
                    )
                    avg_spend = _coerce_numeric(
                        market_size_df.get("Average spend (USD/customer)", pd.Series(dtype=float))
                    )
                    tam = market_size * avg_spend
                    market_size_df["Total Addressable Market Size (USD)"] = tam
                    sam_pct = _coerce_numeric(
                        market_size_df.get("Serviceable Available Market (% TAM)", pd.Series(dtype=float))
                    )
                    market_size_df["Serviceable Available Market (USD)"] = tam * sam_pct.div(100)
                    som_pct = _coerce_numeric(
                        market_size_df.get("Serviceable Obtainable Market (%)", pd.Series(dtype=float))
                    )
                    market_size_df["Serviceable Obtainable Market (USD)"] = tam * som_pct.div(100)
                    st.session_state["market_size_estimation"] = market_size_df
                    market_size_display = market_size_df[
                        [
                            "ID_vaccine",
                            "Vaccine name",
                            "Total Addressable Market Size (USD)",
                            "Serviceable Available Market (USD)",
                            "Serviceable Obtainable Market (USD)",
                        ]
                    ]
                    st.dataframe(
                        market_size_display.style.format(
                            {
                                "Total Addressable Market Size (USD)": "{:.0f}",
                                "Serviceable Available Market (USD)": "{:.0f}",
                                "Serviceable Obtainable Market (USD)": "{:.0f}",
                            }
                        )
                    )
                    market_size_warnings = []
                    if (market_size <= 0).any():
                        market_size_warnings.append("Market size (# customers) should be greater than zero.")
                    if (avg_spend <= 0).any():
                        market_size_warnings.append("Average spend should be greater than zero.")
                    if (sam_pct > 100).any() or (sam_pct < 0).any():
                        market_size_warnings.append("Serviceable Available Market % should be 0–100%.")
                    if (som_pct > 100).any() or (som_pct < 0).any():
                        market_size_warnings.append("Serviceable Obtainable Market % should be 0–100%.")
                    _render_section_warnings("Market size estimation", market_size_warnings)

            if show_revenue_estimation:
                with st.expander("Vaccines revenue estimation", expanded=True):
                    revenue_df = _render_product_assumption_table(
                        session_key="vaccine_revenue_table",
                        default_factory=_default_vaccine_revenue_table,
                        blank_row_factory=_blank_vaccine_revenue_row,
                    )
                    patent_customers = _coerce_numeric(
                        revenue_df.get("Patent customers per year", pd.Series(dtype=float))
                    )
                    patent_price = _coerce_numeric(
                        revenue_df.get("Patent price (USD/customer)", pd.Series(dtype=float))
                    )
                    revenue_df["Patent revenue target (USD)"] = patent_customers * patent_price
                    cust_adj = _coerce_numeric(
                        revenue_df.get("Post patent customer adj. %", pd.Series(dtype=float))
                    ).div(100).replace(0, np.nan)
                    price_adj = _coerce_numeric(
                        revenue_df.get("Post patent price adj. %", pd.Series(dtype=float))
                    ).div(100).replace(0, np.nan)
                    if "Post patent customers per year" not in revenue_df.columns:
                        revenue_df["Post patent customers per year"] = patent_customers * cust_adj.fillna(1.0)
                    else:
                        post_patent_customers = _coerce_numeric(
                            revenue_df["Post patent customers per year"], 0.0
                        )
                        mask_missing = post_patent_customers.isna() | (post_patent_customers == 0)
                        revenue_df.loc[mask_missing, "Post patent customers per year"] = (
                            patent_customers[mask_missing] * cust_adj.fillna(1.0)[mask_missing]
                        )
                    if "Post patent price (USD/customer)" not in revenue_df.columns:
                        revenue_df["Post patent price (USD/customer)"] = patent_price * price_adj.fillna(1.0)
                    else:
                        post_patent_price = _coerce_numeric(
                            revenue_df["Post patent price (USD/customer)"], 0.0
                        )
                        mask_price = post_patent_price.isna() | (post_patent_price == 0)
                        revenue_df.loc[mask_price, "Post patent price (USD/customer)"] = (
                            patent_price[mask_price] * price_adj.fillna(1.0)[mask_price]
                        )
                    revenue_df["Post patent revenue target (USD)"] = (
                        _coerce_numeric(revenue_df["Post patent customers per year"], 0)
                        * _coerce_numeric(revenue_df["Post patent price (USD/customer)"], 0)
                    )
                    st.session_state["vaccine_revenue_table"] = revenue_df
                    revenue_display = revenue_df[
                        [
                            "ID_vaccine",
                            "Vaccine name",
                            "Patent revenue target (USD)",
                            "Post patent revenue target (USD)",
                        ]
                    ]
                    st.dataframe(
                        revenue_display.style.format(
                            {
                                "Patent revenue target (USD)": "{:.0f}",
                                "Post patent revenue target (USD)": "{:.0f}",
                            }
                        )
                    )
                    revenue_warnings = []
                    if (patent_customers < 0).any():
                        revenue_warnings.append("Patent customers per year should be zero or positive.")
                    if (patent_price < 0).any():
                        revenue_warnings.append("Patent price should be zero or positive.")
                    if (revenue_df["Patent revenue target (USD)"] < 0).any():
                        revenue_warnings.append("Patent revenue targets should be zero or positive.")
                    if (revenue_df["Post patent revenue target (USD)"] < 0).any():
                        revenue_warnings.append("Post-patent revenue targets should be zero or positive.")
                    _render_section_warnings("Revenue estimation", revenue_warnings)

            if show_cost_assumptions:
                with st.expander("Vaccine cost assumptions", expanded=True):
                    cost_df = _render_product_assumption_table(
                        session_key="vaccine_cost_table",
                        default_factory=_default_vaccine_cost_table,
                        blank_row_factory=_blank_vaccine_cost_row,
                    )
                    cogs_patent = _coerce_numeric(cost_df.get("COGS patent % of sales", pd.Series(dtype=float)))
                    cogs_post = _coerce_numeric(cost_df.get("COGS post % of sales", pd.Series(dtype=float)))
                    marketing_pct = _coerce_numeric(cost_df.get("Marketing annual % of sales", pd.Series(dtype=float)))
                    royalty_pct = _coerce_numeric(cost_df.get("Royalties cost % of sales", pd.Series(dtype=float)))
                    gna_cols = [
                        "Indirect staff cost (USD)",
                        "Electricity (USD)",
                        "Depreciation (USD)",
                        "Interest & amortization (USD)",
                    ]
                    cost_df["G&A total (USD)"] = cost_df[gna_cols].sum(axis=1)
                    cost_df["Patent operating cost %"] = cogs_patent + marketing_pct + royalty_pct
                    cost_df["Post operating cost %"] = cogs_post + marketing_pct + royalty_pct
                    st.session_state["vaccine_cost_table"] = cost_df
                    cost_display = cost_df[
                        [
                            "ID_vaccine",
                            "Vaccine name",
                            "COGS patent % of sales",
                            "COGS post % of sales",
                            "Marketing annual % of sales",
                            "Marketing launch cost (USD)",
                            "Royalties cost % of sales",
                            "G&A total (USD)",
                            "Patent operating cost %",
                            "Post operating cost %",
                        ]
                    ]
                    percent_cols = [
                        "COGS patent % of sales",
                        "COGS post % of sales",
                        "Marketing annual % of sales",
                        "Royalties cost % of sales",
                        "Patent operating cost %",
                        "Post operating cost %",
                    ]
                    percent_fmt = {col: "{:.1f}%" for col in percent_cols if col in cost_display.columns}
                    currency_fmt = {
                        col: "{:.0f}"
                        for col in ["Marketing launch cost (USD)", "G&A total (USD)"]
                        if col in cost_display.columns
                    }
                    st.dataframe(cost_display.style.format({**percent_fmt, **currency_fmt}))
                    cost_warnings = []
                    for label, series in [
                        ("COGS patent % of sales", cogs_patent),
                        ("COGS post % of sales", cogs_post),
                        ("Marketing annual % of sales", marketing_pct),
                        ("Royalties cost % of sales", royalty_pct),
                    ]:
                        if (series < 0).any() or (series > 1).any():
                            cost_warnings.append(f"{label} should be between 0% and 100%.")
                    if (cost_df["G&A total (USD)"] < 0).any():
                        cost_warnings.append("G&A total should be zero or positive.")
                    _render_section_warnings("Cost assumptions", cost_warnings)

            if show_rd:
                with st.expander("Vaccines research & development (R&D)", expanded=True):
                    rd_df = _render_product_assumption_table(
                        session_key="vaccine_rd_table",
                        default_factory=_default_vaccine_rd_table,
                        blank_row_factory=_blank_vaccine_rd_row,
                    )
                    rd_df["Pre-GTM total (USD)"] = _coerce_numeric(
                        rd_df.get("Pre-GTM spent to date (USD)", pd.Series(dtype=float))
                    ) + _coerce_numeric(rd_df.get("Pre-GTM remaining (USD)", pd.Series(dtype=float)))
                    st.session_state["vaccine_rd_table"] = rd_df
                    rd_display = rd_df[
                        [
                            "ID_vaccine",
                            "Vaccine name",
                            "Cost accounting (capitalisation)",
                            "Pre-GTM spent to date (USD)",
                            "Pre-GTM remaining (USD)",
                            "Pre-GTM total (USD)",
                            "Post-GTM annual cost (USD/year)",
                        ]
                    ]
                    rd_fmt = {
                        col: "{:.0f}"
                        for col in rd_display.columns
                        if col not in ["ID_vaccine", "Vaccine name", "Cost accounting (capitalisation)"]
                    }
                    st.dataframe(rd_display.style.format(rd_fmt))
                    rd_warnings = []
                    for col in [
                        "Pre-GTM spent to date (USD)",
                        "Pre-GTM remaining (USD)",
                        "Post-GTM annual cost (USD/year)",
                    ]:
                        if (pd.to_numeric(rd_df.get(col, pd.Series(dtype=float)), errors="coerce") < 0).any():
                            rd_warnings.append(f"{col} should be zero or positive.")
                    _render_section_warnings("R&D assumptions", rd_warnings)

            if show_capex:
                with st.expander("Vaccine CAPEX assumptions", expanded=True):
                    with _section_block("Shared CAPEX pools", heading_level=4):
                        shared_pools_df = _render_product_assumption_table(
                            session_key="shared_capex_pools_table",
                            default_factory=_default_shared_capex_pools_table,
                            blank_row_factory=lambda df: {
                                "Pool name": "New shared pool",
                                "Applies to (IDs or ALL)": "ALL",
                                "Allocation method": "Equal",
                            },
                            column_config={
                                "Allocation method": st.column_config.SelectboxColumn(
                                    "Allocation method", options=["Equal", "By Weight"]
                                )
                            },
                        )
                        st.session_state["shared_capex_pools_table"] = shared_pools_df
                    with _section_block("Shared CAPEX allocation weights", heading_level=4):
                        shared_allocations_df = _render_product_assumption_table(
                            session_key="shared_capex_allocations_table",
                            default_factory=_default_shared_capex_allocations_table,
                            blank_row_factory=lambda df: {
                                "Pool name": "Core manufacturing facility",
                                "ID_vaccine": _next_vaccine_id(df),
                                "Weight": 1.0,
                            },
                        )
                        st.session_state["shared_capex_allocations_table"] = shared_allocations_df

                    capex_df = _render_product_assumption_table(
                        session_key="vaccine_capex_table",
                        default_factory=_default_vaccine_capex_table,
                        blank_row_factory=_blank_vaccine_capex_row,
                    )
                    capex_pre_cols = [
                        "Manufacturing & Scale-up Assets (Pre-GTM, USD)",
                        "Quality & Compliance Infrastructure (Pre-GTM, USD)",
                        "Cold-chain / Distribution Assets (Pre-GTM, USD)",
                        "IT / Data / Digital Infrastructure (Pre-GTM, USD)",
                        "Facility Build-out / Leasehold Improvements (Pre-GTM, USD)",
                        "Process Development & Tech-Transfer Assets (Pre-GTM, USD)",
                    ]
                    capex_post_cols = [
                        "Manufacturing & Scale-up Assets (Post-GTM, USD/year)",
                        "Quality & Compliance Infrastructure (Post-GTM, USD/year)",
                        "Cold-chain / Distribution Assets (Post-GTM, USD/year)",
                        "IT / Data / Digital Infrastructure (Post-GTM, USD/year)",
                        "Facility Build-out / Leasehold Improvements (Post-GTM, USD/year)",
                        "Process Development & Tech-Transfer Assets (Post-GTM, USD/year)",
                    ]
                    capex_pre = capex_df.get(capex_pre_cols, pd.DataFrame()).apply(
                        pd.to_numeric, errors="coerce"
                    )
                    capex_post = capex_df.get(capex_post_cols, pd.DataFrame()).apply(
                        pd.to_numeric, errors="coerce"
                    )
                    capex_warnings = []
                    if (capex_pre < 0).any().any():
                        capex_warnings.append("Pre-GTM CAPEX entries should be zero or positive.")
                    if (capex_post < 0).any().any():
                        capex_warnings.append("Post-GTM CAPEX entries should be zero or positive.")
                    _render_section_warnings("CAPEX assumptions", capex_warnings)
                    capex_df["Total Pre-GTM capex (USD)"] = capex_pre.fillna(0.0).sum(axis=1)
                    capex_df["Total Post-GTM capex (USD/year)"] = capex_post.fillna(0.0).sum(axis=1)
                    if not shared_pools_df.empty:
                        shared_allocations = _build_shared_capex_allocations(
                            st.session_state.get("vaccine_development_table", pd.DataFrame()),
                            shared_pools_df,
                            shared_allocations_df,
                        )
                        if not shared_allocations.empty:
                            pool_values = shared_pools_df.copy()
                            pool_values["Pool name"] = pool_values.get("Pool name", "").astype(str)
                            pool_values["Pre-GTM total (USD)"] = pool_values.get(
                                capex_pre_cols, pd.DataFrame()
                            ).apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)
                            pool_values["Post-GTM total (USD/year)"] = pool_values.get(
                                capex_post_cols, pd.DataFrame()
                            ).apply(pd.to_numeric, errors="coerce").fillna(0.0).sum(axis=1)
                            shared_totals = shared_allocations.merge(
                                pool_values[
                                    ["Pool name", "Pre-GTM total (USD)", "Post-GTM total (USD/year)"]
                                ],
                                on="Pool name",
                                how="left",
                            )
                            shared_totals["Shared Pre-GTM capex (USD)"] = (
                                shared_totals["Share"]
                                * shared_totals["Pre-GTM total (USD)"].fillna(0.0)
                            )
                            shared_totals["Shared Post-GTM capex (USD/year)"] = (
                                shared_totals["Share"]
                                * shared_totals["Post-GTM total (USD/year)"].fillna(0.0)
                            )
                            shared_summary = (
                                shared_totals.groupby("ID_vaccine", as_index=False)[
                                    ["Shared Pre-GTM capex (USD)", "Shared Post-GTM capex (USD/year)"]
                                ]
                                .sum()
                            )
                            capex_df = capex_df.drop(
                                columns=[
                                    "Shared Pre-GTM capex (USD)",
                                    "Shared Post-GTM capex (USD/year)",
                                ],
                                errors="ignore",
                            )
                            capex_df = capex_df.merge(shared_summary, on="ID_vaccine", how="left")
                            capex_df["Shared Pre-GTM capex (USD)"] = capex_df.get(
                                "Shared Pre-GTM capex (USD)", pd.Series(0.0, index=capex_df.index)
                            ).fillna(0.0)
                            capex_df["Shared Post-GTM capex (USD/year)"] = capex_df.get(
                                "Shared Post-GTM capex (USD/year)", pd.Series(0.0, index=capex_df.index)
                            ).fillna(0.0)
                            capex_df["Total Pre-GTM capex (USD)"] = (
                                capex_df["Total Pre-GTM capex (USD)"]
                                + capex_df["Shared Pre-GTM capex (USD)"]
                            )
                            capex_df["Total Post-GTM capex (USD/year)"] = (
                                capex_df["Total Post-GTM capex (USD/year)"]
                                + capex_df["Shared Post-GTM capex (USD/year)"]
                            )
                    st.session_state["vaccine_capex_table"] = capex_df
                    capex_display = capex_df[
                        [
                            "ID_vaccine",
                            "Vaccine name",
                            "Total Pre-GTM capex (USD)",
                            "Total Post-GTM capex (USD/year)",
                        ]
                    ]
                    capex_fmt = {
                        col: "{:.0f}"
                        for col in capex_display.columns
                        if col not in ["ID_vaccine", "Vaccine name"]
                    }
                    st.dataframe(capex_display.style.format(capex_fmt))

            if show_royalties:
                with st.expander("Vaccines royalty revenues", expanded=True):
                    royalty_df = _render_product_assumption_table(
                        session_key="vaccine_royalty_table",
                        default_factory=_default_royalty_table,
                        blank_row_factory=_blank_vaccine_royalty_row,
                        column_config={
                            "Monetization model": st.column_config.SelectboxColumn(
                                "Monetization model", options=["Product Sale", "Licensing"]
                            )
                        },
                    )
                    revenue_lookup = st.session_state.get("vaccine_revenue_table", pd.DataFrame())
                    if "ID_vaccine" in revenue_lookup.columns:
                        revenue_lookup = revenue_lookup.drop_duplicates("ID_vaccine", keep="last")
                        patent_lookup = revenue_lookup.set_index("ID_vaccine").get(
                            "Patent revenue target (USD)", pd.Series(dtype=float)
                        )
                        post_lookup = revenue_lookup.set_index("ID_vaccine").get(
                            "Post patent revenue target (USD)", pd.Series(dtype=float)
                        )
                    else:
                        patent_lookup = pd.Series(dtype=float)
                        post_lookup = pd.Series(dtype=float)
                    royalty_rate = _coerce_numeric(royalty_df.get("Royalty rate (%)", pd.Series(dtype=float))).div(100)
                    royalty_df["Patent revenue (USD)"] = royalty_df["ID_vaccine"].map(patent_lookup)
                    royalty_df["Post patent revenue (USD)"] = royalty_df["ID_vaccine"].map(post_lookup)
                    royalty_df["Royalty income (USD)"] = royalty_df["Patent revenue (USD)"] * royalty_rate
                    st.session_state["vaccine_royalty_table"] = royalty_df
                    st.dataframe(
                        royalty_df[
                            [
                                "ID_vaccine",
                                "Vaccine name",
                                "Royalty rate (%)",
                                "Royalty income (USD)",
                                "Patent revenue (USD)",
                                "Post patent revenue (USD)",
                            ]
                        ].style.format(
                            {
                                "Royalty rate (%)": "{:.1f}",
                                "Royalty income (USD)": "{:.0f}",
                                "Patent revenue (USD)": "{:.0f}",
                                "Post patent revenue (USD)": "{:.0f}",
                            }
                        )
                    )

            if show_market_share:
                with st.expander("Vaccines market share", expanded=True):
                    market_share_df = _render_product_assumption_table(
                        session_key="vaccine_market_share_table",
                        default_factory=_default_market_share_table,
                        blank_row_factory=_blank_vaccine_market_share_row,
                    )
                    relevant_market = _coerce_numeric(
                        market_share_df.get("Relevant market size (USD)", pd.Series(dtype=float))
                    )
                    patent_target_pct = _coerce_numeric(
                        market_share_df.get("Revenue target - patent %", pd.Series(dtype=float))
                    ).div(100)
                    post_target_pct = _coerce_numeric(
                        market_share_df.get("Revenue target - post %", pd.Series(dtype=float))
                    ).div(100)
                    market_share_df["Revenue target patent (USD)"] = relevant_market * patent_target_pct
                    market_share_df["Revenue target post (USD)"] = relevant_market * post_target_pct
                    st.session_state["vaccine_market_share_table"] = market_share_df
                    st.dataframe(
                        market_share_df[
                            [
                                "ID_vaccine",
                                "Vaccine name",
                                "Relevant market type",
                                "Relevant market size (USD)",
                                "Revenue target patent (USD)",
                                "Revenue target post (USD)",
                                "Market share patent %",
                                "Market share post %",
                                "Market growth %",
                                "Sales growth %",
                            ]
                        ].style.format(
                            {
                                "Relevant market size (USD)": "{:.0f}",
                                "Revenue target patent (USD)": "{:.0f}",
                                "Revenue target post (USD)": "{:.0f}",
                                "Market share patent %": "{:.1f}",
                                "Market share post %": "{:.1f}",
                                "Market growth %": "{:.1f}",
                                "Sales growth %": "{:.1f}",
                            }
                        )
                    )

            with st.expander("Template library", expanded=False):
                templates = _template_library()
                template_name = st.selectbox("Choose a template", options=list(templates.keys()))
                if st.button("Load template into product table"):
                    st.session_state["product_table"] = templates[template_name].copy()
                    st.success(f"Loaded template: {template_name}")
                st.caption("Templates provide starting points for common biotech asset profiles.")

            product_df = _render_product_assumption_table(
                session_key="product_table",
                default_factory=_default_products,
                blank_row_factory=lambda df: _blank_product_row(f"Product {len(df) + 1}"),
                column_config={
                    "stage": st.column_config.SelectboxColumn("Stage", options=STAGE_OPTIONS),
                    "sales_ramp_length": st.column_config.NumberColumn(
                        "Sales ramp length (years)", min_value=0, step=1
                    ),
                    "sales_ramp_shape": st.column_config.SelectboxColumn(
                        "Ramp shape", options=RAMP_SHAPE_OPTIONS
                    ),
                    "include_in_consolidation": st.column_config.CheckboxColumn("Include", default=True),
                    "success_prob": st.column_config.NumberColumn(
                        "Success probability", min_value=0.0, max_value=1.0, step=0.05
                    ),
                    "labor_pct": st.column_config.NumberColumn(
                        "Labor %", min_value=0.0, max_value=1.0, step=0.01
                    ),
                    "overhead_pct": st.column_config.NumberColumn(
                        "Overhead %", min_value=0.0, max_value=1.0, step=0.01
                    ),
                    "material_pct": st.column_config.NumberColumn(
                        "Material %", min_value=0.0, max_value=1.0, step=0.01
                    ),
                },
                id_column=None,
                name_column="name",
            )
            stage_mapping = st.session_state.get(
                "stage_schedule_mapping",
                _default_stage_schedule_mapping(),
            )
            if st.session_state.get("stage_mapping_auto_apply", True):
                product_df = _apply_stage_schedule_defaults(
                    product_df,
                    stage_mapping,
                    stage_column="stage",
                    overwrite=st.session_state.get("stage_mapping_overwrite", False),
                )
            detail_tables = _detail_tables_from_state()
            product_df = _validate_product_df(product_df)
            st.session_state["product_table"] = product_df
            probability_preview = _build_probability_preview(
                product_df,
                model_cfg,
                stage_mapping,
                overwrite_defaults=st.session_state.get("stage_mapping_overwrite", False),
                detail_tables=detail_tables,
            )
            if not probability_preview.empty:
                st.markdown("**Probability basis before valuation**")
                st.dataframe(
                    probability_preview.style.format(
                        {
                            "Input success probability": "{:.1%}",
                            "Effective cumulative success probability": "{:.1%}",
                        }
                    )
                )
                st.caption(
                    "Stage-transition curves are authoritative when present; the single success probability is a fallback only."
                )

            portfolio = _build_portfolio(
                product_df,
                model_cfg,
                stage_mapping=stage_mapping,
                overwrite_defaults=st.session_state.get("stage_mapping_overwrite", False),
                detail_tables=detail_tables,
            )
            if portfolio is None:
                st.info("Add at least one product with a name to run valuations.")
            else:
                validation_issues = validate_portfolio(portfolio)
                if validation_issues:
                    st.error("Validation issues detected:")
                    for issue in validation_issues:
                        st.write(f"- {issue}")
                    st.session_state.pop("model_config", None)
                    st.session_state.pop("portfolio", None)
                    st.session_state.pop("valuation_result", None)
                    portfolio = None
                    valuation_result = None
                else:
                    valuation_result = ValuationEngine(portfolio).run()
                    st.session_state["model_config"] = model_cfg
                    st.session_state["portfolio"] = portfolio
                    st.session_state["valuation_result"] = valuation_result
                    st.success(
                        f"Run complete: enterprise value = {valuation_result.enterprise_value:,.0f} {model_cfg.currency}."
                    )
                    financing_outputs = _build_financing_outputs(valuation_result, model_cfg)
                    equity_bridge = financing_outputs["equity_bridge"]
                    lender_metrics = financing_outputs["lender_metrics"]
                    investor_waterfall = financing_outputs["investor_waterfall"]
                    st.markdown("**Enterprise-to-equity bridge**")
                    st.dataframe(equity_bridge.style.format({"Amount": "{:,.0f}"}), use_container_width=True)
                    if not lender_metrics.empty:
                        st.markdown("**Lender metrics**")
                        st.dataframe(
                            lender_metrics.style.format(
                                {
                                    "CFADS": "{:,.0f}",
                                    "Debt service": "{:,.0f}",
                                    "DSCR": "{:.2f}",
                                    "LLCR": "{:.2f}",
                                    "PLCR": "{:.2f}",
                                    "Minimum cash reserve": "{:,.0f}",
                                    "Cash reserve headroom": "{:,.0f}",
                                }
                            ),
                            use_container_width=True,
                        )
                    if not investor_waterfall.empty:
                        st.markdown("**Investor waterfall**")
                        st.dataframe(
                            investor_waterfall.style.format(
                                {
                                    "Ownership %": "{:.1%}",
                                    "Investment": "{:,.0f}",
                                    "Converted value": "{:,.0f}",
                                    "Preference claim": "{:,.0f}",
                                    "Preference paid": "{:,.0f}",
                                    "Common pool allocation": "{:,.0f}",
                                    "Total proceeds": "{:,.0f}",
                                    "MOIC": "{:.2f}",
                                }
                            ),
                            use_container_width=True,
                        )

    with financial_tab:
        st.subheader("Financial statements")
        if valuation_result is None or model_cfg is None:
            st.info("Run the model configuration tab to populate the statements.")
        else:
            cons = valuation_result.consolidated
            financing_outputs = _build_financing_outputs(valuation_result, model_cfg)
            perf_df = financing_outputs["financial_performance"]
            position_df = financing_outputs["financial_position"]
            cash_flow_df = financing_outputs["cash_flows"]
            lender_metrics = financing_outputs["lender_metrics"]
            investor_waterfall = financing_outputs["investor_waterfall"]
            with st.expander("Consolidated forecast", expanded=True):
                cons_display = cons[["revenue", "ebitda", "fcff_after_wc"]].copy()
                cons_display.columns = ["Revenue", "EBITDA", "FCFF after WC"]
                st.dataframe(
                    cons_display.style.format(
                        {
                            "Revenue": "{:.0f}",
                            "EBITDA": "{:.0f}",
                            "FCFF after WC": "{:.0f}",
                        }
                    )
                )
                st.line_chart(cons_display)
            st.markdown("**Statement of Financial Performance**")
            st.dataframe(
                perf_df.style.format({col: "{:.0f}" for col in perf_df.columns})
            )
            st.markdown("**Statement of Financial Position**")
            st.dataframe(
                position_df.style.format({col: "{:.0f}" for col in position_df.columns})
            )
            st.markdown("**Statement of Cash Flows**")
            st.dataframe(
                cash_flow_df.style.format({col: "{:.0f}" for col in cash_flow_df.columns})
            )
            debt_draw = cash_flow_df.get("Debt drawdowns")
            debt_repay = cash_flow_df.get("Debt repayments")
            debt_open = cash_flow_df.get("Debt opening balance")
            debt_close = cash_flow_df.get("Debt closing balance")
            if debt_draw is not None and debt_repay is not None:
                debt_schedule = pd.DataFrame(
                    {
                        "Beginning balance": debt_open.fillna(0.0) if debt_open is not None else 0.0,
                        "Debt drawdowns": debt_draw.fillna(0.0),
                        "Debt repayments": debt_repay.fillna(0.0),
                        "Ending balance": debt_close.fillna(0.0) if debt_close is not None else 0.0,
                    },
                    index=cash_flow_df.index,
                )
                st.markdown("**Debt schedule**")
                st.dataframe(
                    debt_schedule.style.format({col: "{:.0f}" for col in debt_schedule.columns})
                )
            else:
                st.info("Debt schedule unavailable: cash flow inputs are missing debt columns.")
            if not lender_metrics.empty:
                st.markdown("**Lender metrics**")
                st.dataframe(
                    lender_metrics.style.format(
                        {
                            "CFADS": "{:,.0f}",
                            "Debt service": "{:,.0f}",
                            "DSCR": "{:.2f}",
                            "LLCR": "{:.2f}",
                            "PLCR": "{:.2f}",
                            "Minimum cash reserve": "{:,.0f}",
                            "Cash reserve headroom": "{:,.0f}",
                        }
                    ),
                    use_container_width=True,
                )
            if not investor_waterfall.empty:
                st.markdown("**Investor waterfall**")
                st.dataframe(
                    investor_waterfall.style.format(
                        {
                            "Ownership %": "{:.1%}",
                            "Investment": "{:,.0f}",
                            "Converted value": "{:,.0f}",
                            "Preference claim": "{:,.0f}",
                            "Preference paid": "{:,.0f}",
                            "Common pool allocation": "{:,.0f}",
                            "Total proceeds": "{:,.0f}",
                            "MOIC": "{:.2f}",
                        }
                    ),
                    use_container_width=True,
                )
            st.markdown("**Excel Model Download**")
            excel_bytes = st.session_state.get("financial_excel_bytes")
            download_container = st.container()
            with download_container:
                if not excel_bytes:
                    if st.button("Prepare Excel Model", key="prepare_financial_excel"):
                        with st.spinner("Preparing Excel workbook..."):
                            excel_bytes = _build_financial_excel(
                                cons,
                                perf_df,
                                position_df,
                                cash_flow_df,
                                model_cfg,
                                lender_metrics=lender_metrics,
                                investor_waterfall=investor_waterfall,
                            )
                        st.session_state["financial_excel_bytes"] = excel_bytes
                if excel_bytes:
                    st.download_button(
                        "Download Excel Model",
                        data=excel_bytes,
                        file_name="Financial_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_financial_excel",
                    )
                    if st.button("Clear Prepared Excel", key="clear_financial_excel"):
                        st.session_state.pop("financial_excel_bytes", None)
                        excel_bytes = None
                if not excel_bytes:
                    st.info("Click 'Prepare Excel Model' to generate the workbook for download.")

    with dashboard_scenario_tab:
        st.subheader("Dashboard & scenarios")
        if valuation_result is None or model_cfg is None:
            st.info("Configure and run the model to see dashboard metrics.")
        else:
            st.markdown("**Dashboard snapshot**")
            cons = valuation_result.consolidated
            kpi_cols = st.columns(5)
            kpi_cols[0].metric("Portfolio rNPV", f"{valuation_result.rnpv:,.0f} {model_cfg.currency}")
            peak_year = int(cons["revenue"].idxmax())
            kpi_cols[1].metric("Peak revenue year", f"{peak_year}")
            kpi_cols[2].metric("Peak revenue", f"{cons['revenue'].max():,.0f}")
            avg_margin = cons["ebitda"].sum() / cons["revenue"].sum() if cons["revenue"].sum() else 0.0
            kpi_cols[3].metric("Avg EBITDA margin", f"{avg_margin:.1%}")
            break_even_candidates = cons.index[cons["fcff_after_wc"] > 0]
            break_even_year = int(break_even_candidates[0]) if len(break_even_candidates) else None
            kpi_cols[4].metric("Break-even year", f"{break_even_year}" if break_even_year else "N/A")

            chart_data = cons[["revenue", "ebitda", "fcff_after_wc"]]
            st.area_chart(chart_data)
            st.bar_chart(cons["fcff_after_wc"], use_container_width=True)

            with st.expander("Comparable multiples (EV/EBITDA or EV/Sales)", expanded=False):
                comps_df = st.session_state.get(
                    "comps_table",
                    pd.DataFrame(
                        [
                            {"Peer": "Peer A", "Multiple": 8.0, "Metric": "EV/EBITDA"},
                            {"Peer": "Peer B", "Multiple": 10.0, "Metric": "EV/EBITDA"},
                            {"Peer": "Peer C", "Multiple": 12.0, "Metric": "EV/EBITDA"},
                        ]
                    ),
                )
                comps_df = st.data_editor(comps_df, num_rows="dynamic", key="comps_table_editor")
                st.session_state["comps_table"] = comps_df
                valid_mult = pd.to_numeric(comps_df.get("Multiple"), errors="coerce").dropna()
                if not valid_mult.empty:
                    min_mult = float(valid_mult.min())
                    med_mult = float(valid_mult.median())
                    max_mult = float(valid_mult.max())
                    last_year = cons.index.max()
                    base_ebitda = float(cons.loc[last_year, "ebitda"])
                    st.markdown(
                        f"Implied EV range (using last-year EBITDA {base_ebitda:,.0f}): "
                        f"{min_mult * base_ebitda:,.0f} - {max_mult * base_ebitda:,.0f}"
                    )
                    st.caption(f"Median multiple: {med_mult:.1f}x")
                else:
                    st.info("Add comparable multiples to see implied valuation ranges.")

        st.markdown("**Scenario analysis**")
        if portfolio is None:
            st.info("Configure the model in the first tab to enable scenarios.")
        else:
            preset_col, name_col = st.columns([3, 2])
            with preset_col:
                st.markdown("**Scenario presets**")
                preset_buttons = st.columns(4)

                def _apply_preset(
                    *,
                    rev: float,
                    cost: float,
                    dr: float,
                    prob: float,
                    delay: int = 0,
                    stage_slippage: Optional[Dict[str, int]] = None,
                ) -> None:
                    st.session_state["scenario_rev_mult"] = rev
                    st.session_state["scenario_cost_mult"] = cost
                    st.session_state["scenario_dr_shift"] = dr
                    st.session_state["scenario_prob_mult"] = prob
                    st.session_state["scenario_delay"] = delay
                    st.session_state["scenario_stage_slippage"] = stage_slippage or {}

                if preset_buttons[0].button("Base", key="scenario_preset_base"):
                    _apply_preset(rev=1.0, cost=1.0, dr=0.0, prob=1.0, delay=0)
                if preset_buttons[1].button("Upside", key="scenario_preset_upside"):
                    _apply_preset(rev=1.2, cost=0.9, dr=-0.01, prob=1.1, delay=0)
                if preset_buttons[2].button("Downside", key="scenario_preset_downside"):
                    _apply_preset(rev=0.8, cost=1.1, dr=0.01, prob=0.9, delay=1)
                if preset_buttons[3].button("Trial failure", key="scenario_preset_failure"):
                    _apply_preset(rev=0.6, cost=1.3, dr=0.03, prob=0.75, delay=2)

            with name_col:
                scenario_name = st.text_input("Scenario name", value="Custom scenario", key="scenario_name")

            col1, col2, col3, col4, col5 = st.columns(5)
            rev_mult = col1.slider(
                "Revenue multiplier",
                0.25,
                2.5,
                st.session_state.get("scenario_rev_mult", 1.0),
                key="scenario_rev_mult",
            )
            cost_mult = col2.slider(
                "Cost multiplier",
                0.5,
                2.0,
                st.session_state.get("scenario_cost_mult", 1.0),
                key="scenario_cost_mult",
            )
            dr_shift = col3.slider(
                "Discount rate shift",
                -0.05,
                0.1,
                st.session_state.get("scenario_dr_shift", 0.0),
                key="scenario_dr_shift",
            )
            prob_mult = col4.slider(
                "Success prob multiplier",
                0.5,
                1.5,
                st.session_state.get("scenario_prob_mult", 1.0),
                key="scenario_prob_mult",
            )
            launch_delay = col5.slider(
                "Launch delay (years)",
                0,
                5,
                int(st.session_state.get("scenario_delay", 0)),
                key="scenario_delay",
            )
            st.markdown("**Stage slippage (years)**")
            slip_col1, slip_col2 = st.columns(2)
            slip_phase_ii = slip_col1.slider(
                "Phase II delay",
                0,
                3,
                int(st.session_state.get("scenario_slip_phase_ii", 0)),
                key="scenario_slip_phase_ii",
            )
            slip_phase_iii = slip_col2.slider(
                "Phase III delay",
                0,
                3,
                int(st.session_state.get("scenario_slip_phase_iii", 0)),
                key="scenario_slip_phase_iii",
            )
            stage_slippage = {
                "Phase II": int(slip_phase_ii),
                "Phase III": int(slip_phase_iii),
            }
            scenario = Scenario(
                name=scenario_name or "Custom scenario",
                revenue_multiplier=float(rev_mult),
                cost_multiplier=float(cost_mult),
                discount_rate_shift=float(dr_shift),
                success_prob_multiplier=float(prob_mult),
                launch_delay_years=int(launch_delay),
                stage_slippage_years=stage_slippage,
            )
            scen_results = ScenarioEngine(portfolio).run_scenarios([scenario])

            scenario_result = _evaluate_portfolio_shock(
                portfolio,
                revenue_multiplier=float(rev_mult),
                cost_multiplier=float(cost_mult),
                discount_shift=float(dr_shift),
                success_prob_multiplier=float(prob_mult),
                launch_delay_years=int(launch_delay),
                stage_slippage_years=stage_slippage,
            )
            if scenario_result is not None and valuation_result is not None:
                base_cons = valuation_result.consolidated
                base_rnpv = valuation_result.rnpv
                base_ebitda = base_cons["ebitda"].sum()
                scen_cons = scenario_result.consolidated
                scen_rnpv = scenario_result.rnpv
                scen_ebitda = scen_cons["ebitda"].sum()
                delta_cols = st.columns(4)
                delta_cols[0].metric("Scenario rNPV", f"{scen_rnpv:,.0f}", f"{scen_rnpv - base_rnpv:+,.0f}")
                delta_cols[1].metric(
                    "Scenario EBITDA",
                    f"{scen_ebitda:,.0f}",
                    f"{scen_ebitda - base_ebitda:+,.0f}",
                )
                delta_cols[2].metric(
                    "Revenue delta",
                    f"{scen_cons['revenue'].sum():,.0f}",
                    f"{scen_cons['revenue'].sum() - base_cons['revenue'].sum():+,.0f}",
                )
                delta_cols[3].metric(
                    "FCFF delta",
                    f"{scen_cons['fcff_after_wc'].sum():,.0f}",
                    f"{scen_cons['fcff_after_wc'].sum() - base_cons['fcff_after_wc'].sum():+,.0f}",
                )

                def _funding_required_from_cons(cons_df: pd.DataFrame) -> float:
                    uses_total = float(st.session_state.get("uses_total", 0.0))
                    burn_total = 0.0
                    wc_total = 0.0
                    if "fcff_after_wc" in cons_df.columns:
                        burn_total = float((-cons_df["fcff_after_wc"].clip(upper=0)).sum())
                    if "delta_wc" in cons_df.columns:
                        wc_total = float((-cons_df["delta_wc"].clip(upper=0)).sum())
                    return uses_total + burn_total + wc_total

                component_rows = [
                    {
                        "Component": "Revenue",
                        "Base": float(base_cons["revenue"].sum()),
                        "Scenario": float(scen_cons["revenue"].sum()),
                    },
                    {
                        "Component": "R&D cash burn",
                        "Base": float((-base_cons.get("rd_cash", pd.Series(0.0, index=base_cons.index))).sum()),
                        "Scenario": float((-scen_cons.get("rd_cash", pd.Series(0.0, index=scen_cons.index))).sum()),
                    },
                    {
                        "Component": "CAPEX cash",
                        "Base": float((-base_cons.get("capex_cash", pd.Series(0.0, index=base_cons.index))).sum()),
                        "Scenario": float((-scen_cons.get("capex_cash", pd.Series(0.0, index=scen_cons.index))).sum()),
                    },
                    {
                        "Component": "Equity required (uses + burn + WC)",
                        "Base": _funding_required_from_cons(base_cons),
                        "Scenario": _funding_required_from_cons(scen_cons),
                    },
                ]
                component_df = pd.DataFrame(component_rows)
                component_df["Delta"] = component_df["Scenario"] - component_df["Base"]
                st.markdown("**Scenario deltas by component**")
                st.dataframe(
                    component_df.style.format(
                        {"Base": "{:,.0f}", "Scenario": "{:,.0f}", "Delta": "{:+,.0f}"}
                    ),
                    use_container_width=True,
                )

                overlay_df = pd.DataFrame(
                    {
                        "Base revenue": base_cons["revenue"],
                        "Scenario revenue": scen_cons["revenue"],
                        "Base EBITDA": base_cons["ebitda"],
                        "Scenario EBITDA": scen_cons["ebitda"],
                        "Base FCFF": base_cons["fcff_after_wc"],
                        "Scenario FCFF": scen_cons["fcff_after_wc"],
                    }
                )
                st.markdown("**Scenario overlay vs base**")
                st.line_chart(overlay_df)

            st.markdown("**Scenario result**")
            st.dataframe(scen_results.style.format({"rnpv": "{:.0f}", "ebitda_value": "{:.0f}"}))

            st.markdown("**Multi-scenario comparison**")
            if "scenario_basket" not in st.session_state:
                st.session_state["scenario_basket"] = []
            basket_col1, basket_col2 = st.columns([1, 1])
            if basket_col1.button("Add to comparison", key="scenario_add_to_basket"):
                st.session_state["scenario_basket"].append(
                    {
                        "name": scenario.name,
                        "revenue_multiplier": float(rev_mult),
                        "cost_multiplier": float(cost_mult),
                        "discount_rate_shift": float(dr_shift),
                        "success_prob_multiplier": float(prob_mult),
                    }
                )
            if basket_col2.button("Clear comparison", key="scenario_clear_basket"):
                st.session_state["scenario_basket"] = []

            basket = st.session_state.get("scenario_basket", [])
            if basket:
                scenario_list = [Scenario(**entry) for entry in basket]
                basket_results = ScenarioEngine(portfolio).run_scenarios(scenario_list)
                st.dataframe(
                    basket_results.style.format({"rnpv": "{:.0f}", "ebitda_value": "{:.0f}"})
                )
            else:
                st.caption("Add scenarios to compare multiple cases side-by-side.")

            st.markdown("**Tornado sensitivity (interactive)**")
            if valuation_result is not None:
                tornado_df = _tornado_dataframe(portfolio, valuation_result.rnpv)
                if tornado_df.empty:
                    st.info("Unable to compute tornado deltas.")
                else:
                    st.dataframe(tornado_df.style.format({"rnpv": "{:.0f}", "Delta": "{:+,.0f}"}))
            else:
                st.info("Run a valuation to unlock tornado sensitivities.")

            st.markdown("**Goal seek (scenario)**")
            target_rnpv = st.number_input(
                "Target rNPV",
                value=float(valuation_result.rnpv) if valuation_result is not None else 0.0,
                key="scenario_goal_seek_target",
            )
            if st.button("Solve revenue multiplier", key="scenario_goal_seek"):
                multiplier, achieved = _goal_seek_revenue_multiplier(portfolio, float(target_rnpv))
                if achieved is not None:
                    st.success(
                        f"Revenue multiplier {multiplier:.2f} approximates the goal (achieved rNPV {achieved:,.0f})."
                    )
                else:
                    st.warning("Goal seek failed—try adjusting the target or assumptions.")

    with analytics_tab:
        st.subheader("Advanced financial analytics")
        if valuation_result is None or model_cfg is None or portfolio is None:
            st.info("Configure the model to unlock analytics.")
        else:
            cons = valuation_result.consolidated
            base_rnpv = valuation_result.rnpv
            ratios = _build_ratio_table(cons)
            st.markdown("**Margin & intensity analysis**")
            st.dataframe(ratios.style.format("{:.1%}"))

            with st.expander("Probability-weighted cost burden", expanded=False):
                per_product_prob = valuation_result.per_product_prob
                cost_rows: List[Dict[str, float]] = []
                cost_columns = [
                    "cogs",
                    "labor",
                    "overhead",
                    "materials",
                    "sales_marketing",
                    "gna",
                    "royalty",
                ]
                for name, df in per_product_prob.items():
                    opex = df[cost_columns].sum().sum() if all(col in df.columns for col in cost_columns) else 0.0
                    rd_cash = df["rd_cash"].sum() if "rd_cash" in df.columns else 0.0
                    capex_cash = df["capex_cash"].sum() if "capex_cash" in df.columns else 0.0
                    total_cost = -(opex + rd_cash + capex_cash)
                    cost_rows.append(
                        {
                            "Product": name,
                            "Probability-weighted opex": -opex,
                            "Probability-weighted R&D cash": -rd_cash,
                            "Probability-weighted CAPEX": -capex_cash,
                            "Total cost burden": total_cost,
                        }
                    )
                if cost_rows:
                    cost_df = pd.DataFrame(cost_rows).sort_values("Total cost burden", ascending=False)
                    st.dataframe(
                        cost_df.style.format(
                            {
                                "Probability-weighted opex": "{:,.0f}",
                                "Probability-weighted R&D cash": "{:,.0f}",
                                "Probability-weighted CAPEX": "{:,.0f}",
                                "Total cost burden": "{:,.0f}",
                            }
                        ),
                        use_container_width=True,
                    )
                    st.caption(
                        "Costs are weighted by the annual success schedule so early-stage programs show "
                        "risk-adjusted cash burn rather than raw spend."
                    )
                else:
                    st.info("Run the model to view probability-weighted cost burdens.")
            st.markdown("**Vaccine break-even analysis (interactive)**")
            base_inputs = _build_vaccine_break_even_inputs(model_cfg)
            if base_inputs.empty:
                st.info("Add vaccine assumptions to unlock break-even analytics.")
            else:
                if "vaccine_break_even_inputs" not in st.session_state:
                    st.session_state["vaccine_break_even_inputs"] = base_inputs
                else:
                    current_inputs = st.session_state["vaccine_break_even_inputs"]
                    if isinstance(current_inputs, pd.DataFrame):
                        missing = set(base_inputs["Vaccine name"]) - set(current_inputs.get("Vaccine name", []))
                        if missing:
                            st.session_state["vaccine_break_even_inputs"] = pd.concat(
                                [
                                    current_inputs,
                                    base_inputs[base_inputs["Vaccine name"].isin(missing)],
                                ],
                                ignore_index=True,
                            )
                ai_cols = st.columns(2)
                ai_assist = ai_cols[0].toggle(
                    "AI/ML assist: suggest unit prices for target break-even",
                    value=st.session_state.get("break_even_ai_assist", True),
                    key="break_even_ai_assist",
                )
                ai_target_years = ai_cols[1].slider(
                    "Target break-even horizon (years)",
                    1,
                    10,
                    st.session_state.get("break_even_ai_target_years", 3),
                    key="break_even_ai_target_years",
                )
                st.caption(
                    "Adjust unit price and cost inputs to see contribution margin, break-even units, "
                    "and AI-assisted price suggestions based on the target horizon."
                )
                edited_inputs = st.data_editor(
                    st.session_state["vaccine_break_even_inputs"],
                    use_container_width=True,
                    num_rows="dynamic",
                    column_config={
                        "Unit price (USD)": st.column_config.NumberColumn(format="$%0.2f", step=1.0),
                        "Unit variable cost (USD)": st.column_config.NumberColumn(format="$%0.2f", step=1.0),
                        "Unit fixed cost (USD/year)": st.column_config.NumberColumn(format="$%0.0f", step=1000.0),
                        "Units per year": st.column_config.NumberColumn(format="%0.0f", step=1.0),
                    },
                    key="vaccine_break_even_editor",
                )
                st.session_state["vaccine_break_even_inputs"] = edited_inputs
                break_even_df = _build_vaccine_break_even_table(
                    model_cfg,
                    inputs_df=edited_inputs,
                    ai_assist=ai_assist,
                    ai_target_years=ai_target_years,
                )
                st.markdown("**Break-even outputs**")
                st.dataframe(
                    break_even_df.style.format(
                        {
                            "Unit price (USD)": "{:,.2f}",
                            "Unit variable cost (USD)": "{:,.2f}",
                            "Unit fixed cost (USD/year)": "{:,.0f}",
                            "Units per year": "{:,.0f}",
                            "Unit contribution margin (USD)": "{:,.2f}",
                            "Contribution margin %": "{:.1%}",
                            "Break-even units": "{:,.0f}",
                            "Break-even revenue (USD)": "{:,.0f}",
                            "Break-even unit cost (USD)": "{:,.2f}",
                            "AI suggested unit price (USD)": "{:,.2f}",
                        }
                    )
                )

            with st.expander("Sensitivity & stress testing", expanded=True):
                sens_cols = st.columns(3)
                pricing_delta = sens_cols[0].slider(
                    "Pricing pressure swing",
                    0.0,
                    0.5,
                    0.15,
                    help="Revenue-linked driver",
                )
                manufacturing_delta = sens_cols[1].slider("Manufacturing cost swing", 0.0, 0.5, 0.2)
                clinical_delta = sens_cols[2].slider("Clinical success swing", 0.0, 0.5, 0.1)
                drivers = {
                    "Pricing pressure": (pricing_delta, "revenue"),
                    "Manufacturing costs": (manufacturing_delta, "cost"),
                    "Clinical success": (clinical_delta, "productivity"),
                }
                sens_df = _run_sensitivity_matrix(portfolio, drivers)
                if sens_df.empty:
                    st.info("Not enough data to compute sensitivities.")
                else:
                    st.dataframe(sens_df.style.format({"rNPV": "{:.0f}", "Delta vs base": "{:+,.0f}"}))

                st.markdown("**Scenario stress testing**")
                severe_cases = [
                    ("Regulatory delay", 0.7, 1.2, 0.03, 0.9),
                    ("Trial failure", 0.6, 1.3, 0.04, 0.75),
                    ("Pricing squeeze", 0.5, 1.05, 0.02, 0.95),
                ]
                stress_rows = []
                for name, rev_mult, cost_mult, dr_shift, prob_mult in severe_cases:
                    result = _evaluate_portfolio_shock(
                        portfolio,
                        revenue_multiplier=rev_mult,
                        cost_multiplier=cost_mult,
                        discount_shift=dr_shift,
                        success_prob_multiplier=prob_mult,
                    )
                    if result is None:
                        continue
                    stress_rows.append(
                        {
                            "Scenario": name,
                            "rNPV": result.rnpv,
                            "EBITDA impact": result.consolidated["ebitda"].sum(),
                        }
                    )
                if stress_rows:
                    stress_df = pd.DataFrame(stress_rows)
                    numeric_cols = stress_df.select_dtypes(include="number").columns
                    formatter = {col: "{:+,.0f}" if "impact" in col.lower() else "{:,}" for col in numeric_cols}
                    st.dataframe(stress_df.style.format(formatter))

            with st.expander("Trend, seasonality & segmentation", expanded=False):
                decomp_df = _compute_decomposition(cons)
                if decomp_df is not None:
                    st.line_chart(decomp_df)
                else:
                    st.info("Need more history to decompose trend/seasonality.")

                seg_df = _build_segmentation_table(valuation_result)
                if not seg_df.empty:
                    st.dataframe(
                        seg_df.style.format({
                            "Revenue share": "{:.1%}",
                            "EBITDA margin": "{:.1%}",
                            "FCFF (PV proxy)": "{:.0f}",
                        })
                    )
                    st.bar_chart(seg_df.set_index("Product")["Revenue share"])
                else:
                    st.info("Add probability-weighted products to see segmentation insights.")

            with st.expander("Monte Carlo & probabilistic valuation", expanded=False):
                mc_cols = st.columns(4)
                n_sims = mc_cols[0].number_input("Simulations", min_value=100, max_value=5000, value=1000, step=100)
                rev_dist = mc_cols[1].selectbox("Revenue distribution", ["Normal", "Lognormal", "Uniform"])
                cost_dist = mc_cols[2].selectbox("Cost distribution", ["Normal", "Lognormal", "Uniform"])
                seed = mc_cols[3].number_input("Random seed", min_value=0, value=42)

                sigma_cols = st.columns(3)
                rev_sigma = sigma_cols[0].number_input(
                    "Revenue sigma", min_value=0.01, max_value=0.5, value=0.15, step=0.01
                )
                cost_sigma = sigma_cols[1].number_input(
                    "Cost sigma", min_value=0.01, max_value=0.5, value=0.1, step=0.01
                )
                launch_delay_sigma = sigma_cols[2].number_input(
                    "Launch delay sigma (years)", min_value=0.0, max_value=5.0, value=0.5, step=0.1
                )
                rev_bounds = st.columns(2)
                rev_min = rev_bounds[0].number_input("Revenue min (uniform)", value=0.8, step=0.05)
                rev_max = rev_bounds[1].number_input("Revenue max (uniform)", value=1.2, step=0.05)
                cost_bounds = st.columns(2)
                cost_min = cost_bounds[0].number_input("Cost min (uniform)", value=0.8, step=0.05)
                cost_max = cost_bounds[1].number_input("Cost max (uniform)", value=1.2, step=0.05)

                if st.button("Run Monte Carlo simulation"):
                    sims = MonteCarloEngine(portfolio).simulate(
                        n_sims=int(n_sims),
                        revenue_sigma=float(rev_sigma),
                        cost_sigma=float(cost_sigma),
                        revenue_dist=str(rev_dist).lower(),
                        cost_dist=str(cost_dist).lower(),
                        revenue_min=float(rev_min),
                        revenue_max=float(rev_max),
                        cost_min=float(cost_min),
                        cost_max=float(cost_max),
                        launch_delay_sigma=float(launch_delay_sigma),
                        random_seed=int(seed),
                    )
                    st.session_state["mc_results"] = sims

                sims = st.session_state.get("mc_results")
                if sims is not None:
                    st.line_chart(sims.reset_index(drop=True))
                    hist = np.histogram(sims, bins=20)
                    st.bar_chart(pd.DataFrame({"rNPV": hist[0]}, index=hist[1][:-1]))
                    var = MonteCarloEngine.value_at_risk(sims)
                    cvar = MonteCarloEngine.conditional_value_at_risk(sims)
                    st.write(
                        f"Mean rNPV: {sims.mean():,.0f} | Std: {sims.std():,.0f} | VaR95: {var:,.0f} | CVaR95: {cvar:,.0f}"
                    )
                    st.write(
                        "Probabilistic valuation percentiles:",
                        sims.quantile([0.1, 0.25, 0.5, 0.75, 0.9]).to_dict(),
                    )

                else:
                    st.info("Run the simulation to unlock probabilistic metrics.")

            with st.expander("What-if analysis & goal seek", expanded=False):
                what_cols = st.columns(3)
                what_rev = what_cols[0].slider("Revenue multiplier", 0.4, 2.0, 1.0)
                what_cost = what_cols[1].slider("Cost multiplier", 0.5, 2.5, 1.0)
                what_dr = what_cols[2].slider("Discount shift", -0.05, 0.1, 0.0)
                if st.button("Evaluate what-if case"):
                    result = _evaluate_portfolio_shock(
                        portfolio,
                        revenue_multiplier=float(what_rev),
                        cost_multiplier=float(what_cost),
                        discount_shift=float(what_dr),
                    )
                    if result is not None:
                        st.success(f"What-if rNPV: {result.rnpv:,.0f}")

                target_rnpv = st.number_input("Target rNPV for goal seek", value=base_rnpv)
                if st.button("Solve revenue multiplier"):
                    multiplier, achieved = _goal_seek_revenue_multiplier(portfolio, float(target_rnpv))
                    if achieved is not None:
                        st.write(
                            f"Revenue multiplier {multiplier:.2f} approximates the goal (achieved rNPV {achieved:,.0f})."
                        )
                    else:
                        st.warning("Goal seek failed—try adjusting the target or assumptions.")

            with st.expander("Tornado & spider diagnostics", expanded=False):
                tornado_df = _tornado_dataframe(portfolio, base_rnpv)
                if tornado_df.empty:
                    st.info("Unable to compute tornado deltas.")
                else:
                    st.dataframe(tornado_df.style.format({"rNPV": "{:.0f}", "Delta": "{:+,.0f}"}))
                    if go is not None:
                        tornado_fig = go.Figure()
                        pos = tornado_df[tornado_df["Delta"] >= 0]
                        neg = tornado_df[tornado_df["Delta"] < 0]
                        tornado_fig.add_trace(
                            go.Bar(
                                y=pos["Driver"],
                                x=pos["Delta"],
                                orientation="h",
                                name="Positive",
                            )
                        )
                        tornado_fig.add_trace(
                            go.Bar(
                                y=neg["Driver"],
                                x=neg["Delta"],
                                orientation="h",
                                name="Negative",
                            )
                        )
                        tornado_fig.update_layout(barmode="relative", title="Tornado impact")
                        st.plotly_chart(tornado_fig, use_container_width=True)

                        spider_fig = go.Figure()
                        pivot = tornado_df.pivot(index="Driver", columns="Change", values="rNPV").fillna(base_rnpv)
                        spider_fig.add_trace(
                            go.Scatterpolar(r=pivot.get("+20%", [base_rnpv]), theta=pivot.index, name="Upside")
                        )
                        spider_fig.add_trace(
                            go.Scatterpolar(r=pivot.get("-20%", [base_rnpv]), theta=pivot.index, name="Downside")
                        )
                        st.plotly_chart(spider_fig, use_container_width=True)

            with st.expander("Regression & classification models", expanded=False):
                reg_df = _run_linear_regressions(cons)
                if reg_df is not None:
                    st.table(reg_df.style.format({"Intercept": "{:.0f}", "Revenue beta": "{:.2f}", "R^2": "{:.2f}"}))
                else:
                    st.info("Install scikit-learn to unlock regression diagnostics.")

                seg_df = _build_segmentation_table(valuation_result)
                class_df = _run_classification_model(seg_df)
                if class_df is not None:
                    st.dataframe(class_df.style.format({
                        "Revenue share": "{:.1%}",
                        "EBITDA margin": "{:.1%}",
                        "High-margin probability": "{:.1%}",
                    }))
                else:
                    st.caption("Classification output requires scikit-learn and at least one product.")

            with st.expander("Time-series & ML forecasting", expanded=False):
                ts_metric = st.selectbox("Series to forecast", ["revenue", "ebitda"], key="forecast_metric")
                method = st.selectbox("Forecast model", ["ARIMA", "Prophet", "LSTM"], key="forecast_method")
                try:
                    horizon_years = int(model_cfg.n_years)
                except (TypeError, ValueError):
                    horizon_years = 5
                horizon_max = int(max(5, horizon_years))
                horizon_default = int(min(10, horizon_max))
                horizon_default = min(max(5, horizon_default), horizon_max)
                if horizon_max <= 5:
                    horizon = st.number_input(
                        "Forecast steps",
                        min_value=5,
                        max_value=5,
                        value=5,
                        step=1,
                    )
                else:
                    horizon = st.slider("Forecast steps", 5, horizon_max, horizon_default)
                if st.button("Run time-series model"):
                    fe = ForecastEngine(model_cfg)
                    period_index = pd.period_range(str(model_cfg.first_year), periods=len(cons), freq="Y")
                    series = pd.Series(cons[ts_metric].values, index=period_index)
                    series.index = series.index.to_timestamp()
                    try:
                        if method == "ARIMA":
                            forecast = fe.forecast_arima(series, steps=horizon)
                            st.line_chart(forecast)
                        elif method == "Prophet":
                            hist_df = pd.DataFrame({"ds": series.index, "y": series.values})
                            forecast = fe.forecast_prophet(hist_df, periods=horizon)
                            st.line_chart(forecast.set_index("ds")["yhat"])
                        else:
                            forecast = fe.forecast_lstm(series, steps_ahead=horizon)
                            st.line_chart(pd.Series(forecast))
                    except Exception as exc:
                        st.warning(f"Forecast failed: {exc}")

            with st.expander("Optimisation, portfolio design & real options", expanded=False):
                opt_df = _optimize_operations(cons)
                if opt_df is not None:
                    st.table(opt_df.style.format({"Value": "{:.2f}"}))
                else:
                    st.caption("Install SciPy to enable nonlinear optimisation.")

                mv_df = _mean_variance_portfolio(valuation_result)
                if mv_df is not None:
                    st.dataframe(
                        mv_df.style.format({"Mean": "{:.0f}", "Std": "{:.0f}", "Suggested weight": "{:.1%}"})
                    )

                option_val = _real_options_value(valuation_result)
                if option_val is not None:
                    st.write(f"Real option (deferral) value estimate: {option_val:,.0f}")
                else:
                    st.caption("Provide R&D cash flows and install SciPy to compute real options.")

            with st.expander("Risk, copulas, macro & ESG linkages", expanded=False):
                copula_df = _copula_simulation(cons)
                if copula_df is not None:
                    st.scatter_chart(copula_df)

                macro_cols = st.columns(4)
                inflation = macro_cols[0].slider("Inflation", 0.0, 0.15, 0.03)
                gdp = macro_cols[1].slider("GDP growth", -0.05, 0.1, 0.02)
                fx = macro_cols[2].slider("FX depreciation", -0.1, 0.2, 0.0)
                sentiment = macro_cols[3].slider("Market sentiment", -0.3, 0.3, 0.0)
                macro_revenue = cons["revenue"] * (1 + inflation + gdp + sentiment - fx)
                st.line_chart(pd.DataFrame({"Original": cons["revenue"], "Macro-adjusted": macro_revenue}))

                esg_cols = st.columns(3)
                carbon_price = esg_cols[0].slider("Carbon price ($/t)", 0, 200, 75)
                emissions = esg_cols[1].slider("Emissions (kt)", 0, 500, 120)
                renewable_share = esg_cols[2].slider("Renewable share", 0.0, 1.0, 0.35)
                esg_cost = carbon_price * emissions * (1 - renewable_share)
                st.write(f"ESG-adjusted annual carbon cost: {esg_cost:,.0f}")

                intel_score = st.slider("Market intelligence sentiment", -1.0, 1.0, 0.1)
                st.write(
                    f"Sentiment-adjusted revenue uplift: {(intel_score * 5):+.1f}% applied to TAM during scenario planning."
                )

            with st.expander("Comparative & ML-based valuation", expanded=False):
                cluster_df = _cluster_products(valuation_result)
                if cluster_df is not None:
                    st.dataframe(cluster_df)
                else:
                    st.caption("Need scikit-learn and multiple products for clustering.")

                ml_mult_df = _machine_learning_multiple(cons)
                if ml_mult_df is not None:
                    st.line_chart(ml_mult_df.set_index("Year"))
                else:
                    st.caption("Install scikit-learn to run ML-driven multiple predictions.")

    with vc_tab:
        st.subheader("VC method helper")
        if valuation_result is None or model_cfg is None:
            st.info("Configure the model and run a valuation before using VC analysis.")
        else:
            vc_col1, vc_col2, vc_col3, vc_col4 = st.columns(4)
            cons_index = valuation_result.consolidated.index
            exit_year_min = int(cons_index.min())
            exit_year_max = int(cons_index.max())
            exit_year = vc_col1.number_input(
                "Exit year",
                min_value=exit_year_min,
                max_value=exit_year_max,
                value=min(exit_year_max, model_cfg.first_year + 5),
            )
            target_irr = vc_col2.slider("Target IRR", 0.05, 0.6, 0.3)
            ownership = vc_col3.slider("Investor ownership at exit", 0.05, 0.9, 0.25)
            new_money = vc_col4.number_input(
                "New money ($)", min_value=1_000_000, value=50_000_000, step=5_000_000
            )
            exit_multiple = st.slider("Exit EV/EBITDA multiple", 2.0, 25.0, model_cfg.ev_ebitda_multiple)

            vc_inputs = VCInputs(
                exit_year=int(exit_year),
                target_irr=float(target_irr),
                investor_ownership_at_exit=float(ownership),
                new_money=float(new_money),
            )
            vc_valuator = VCValuator(valuation_result)
            try:
                vc_output = vc_valuator.vc_method(vc_inputs, exit_multiple=float(exit_multiple))
            except ValueError as exc:
                st.error(f"VC method failed: {exc}")
            else:
                vc_df = pd.DataFrame(
                    {
                        "Metric": list(vc_output.keys()),
                        "Value": [
                            f"{value:,.0f}" if "irr" not in key else f"{value:.2%}"
                            for key, value in vc_output.items()
                        ],
                    }
                )
                st.table(vc_df)

    with rag_tab:
        _render_rag_assistant_page()

    st.caption(
        "Tip: Upload a Prophet-ready dataframe (ds, y) and plug it into ForecastScenarioBridge for richer scenarios."
    )


# ---------------------------------------------------------------------------
# Scenario state hooks — called by the parent NumQuants shell.
# ---------------------------------------------------------------------------

# DataFrames stored directly in session_state (serialised by the parent app's
# _serialize() helper). Portfolio and ValuationResult are derived objects and
# are rebuilt from these tables on the next render — no need to save them.
_BIOTECH_DF_KEYS = [
    "product_table",
    "stage_schedule_mapping",
    "vaccine_sales_table",
    "vaccine_development_table",
    "vaccine_revenue_table",
    "vaccine_cost_table",
    "vaccine_rd_table",
    "vaccine_capex_table",
    "vaccine_royalty_table",
    "vaccine_market_share_table",
    "market_size_estimation",
    "uses_table",
    "sources_table",
    "shareholders_table",
    "debt_schedule_table",
]

_BIOTECH_SCALAR_KEYS = [
    "stage_mapping_auto_apply",
    "stage_mapping_overwrite",
    "scenario_basket",
    "scenario_rev_mult",
    "scenario_cost_mult",
    "scenario_dr_shift",
    "scenario_prob_mult",
    "vaccine_sales_first_year",
    "vaccine_sales_n_years",
    "debt_interest_rate",
    "debt_repayment_mode",
    "debt_grace_years",
    "debt_target_dscr",
    "minimum_cash_reserve",
    "funding_required",
    "planned_new_equity",
    "opening_nol_balance",
    "debt_schedule_first_year",
    "debt_schedule_n_years",
]


def get_state() -> dict:
    """Snapshot all user-editable inputs.

    ModelConfig is serialised via dataclasses.asdict() so it round-trips through
    JSON cleanly. Portfolio and ValuationResult are omitted — they are rebuilt
    from product_table + model_config on the next render.
    """
    import streamlit as _st
    state: dict = {}

    # ModelConfig dataclass → plain dict
    model_cfg = _st.session_state.get("model_config")
    if model_cfg is not None:
        from dataclasses import asdict as _asdict
        try:
            state["model_config"] = _asdict(model_cfg)
        except TypeError:
            pass  # not a dataclass, skip

    # DataFrames and scalars
    for key in _BIOTECH_DF_KEYS + _BIOTECH_SCALAR_KEYS:
        val = _st.session_state.get(key)
        if val is not None:
            state[key] = val

    return state


def set_state(state: dict) -> None:
    """Restore a previously saved state snapshot.

    Reconstructs ModelConfig from its saved dict. DataFrames are written
    directly to session_state; portfolio is rebuilt by main() on next render.
    """
    import streamlit as _st

    if "model_config" in state:
        try:
            _st.session_state["model_config"] = ModelConfig(**state["model_config"])
        except (TypeError, KeyError):
            pass  # schema changed; leave unset so main() uses defaults

    for key in _BIOTECH_DF_KEYS + _BIOTECH_SCALAR_KEYS:
        if key in state:
            _st.session_state[key] = state[key]


if __name__ == "__main__":
    main()
