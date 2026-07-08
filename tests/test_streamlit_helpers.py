import contextlib
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from streamlit.errors import StreamlitAPIException
from streamlit.testing.v1 import AppTest

from streamlit_app import (
    _apply_debt_schedule,
    _build_bankable_snapshot_payload,
    _build_stage_mapping_candidate_row,
    _build_financial_excel,
    _build_enterprise_to_equity_bridge,
    _build_investor_waterfall,
    _build_lender_metrics,
    _build_portfolio,
    _machine_learning_multiple,
    _build_probability_preview,
    _default_debt_schedule,
    _default_stage_schedule_mapping,
    _consume_pending_panel_state,
    _pending_panel_state_key,
    _render_row_selector,
    _resolve_valuation_context,
    _stage_mapping_row_warnings,
    _set_dataframe_cell,
)
from valuation_codex_package.core import ModelConfig, Product, ProductConfig, Portfolio, ValuationEngine
from valuation_codex_package.ui_state import _validate_selection


APP_PATH = Path(__file__).resolve().parents[1] / "streamlit_app.py"


class StreamlitHelperTests(unittest.TestCase):
    def test_default_debt_schedule_includes_tenor_column(self) -> None:
        schedule = _default_debt_schedule(2025, 5)

        self.assertIn("Loan tenor (years)", schedule.columns)
        self.assertEqual(schedule.loc[0, "Loan tenor (years)"], 5)
        self.assertEqual(schedule.loc[4, "Loan tenor (years)"], 0)

    def test_apply_debt_schedule_respects_facility_tenor_and_later_new_debt(self) -> None:
        cash_flow_df = pd.DataFrame(
            {
                "Net cash from operations": [0.0, 0.0, 0.0, 0.0],
                "Net cash from investing": [0.0, 0.0, 0.0, 0.0],
                "Equity issuance": [0.0, 0.0, 0.0, 0.0],
                "Beginning cash balance": [0.0, 0.0, 0.0, 0.0],
            },
            index=[2025, 2026, 2027, 2028],
        )
        debt_schedule = pd.DataFrame(
            [
                {
                    "Year": 2025,
                    "Debt drawdowns": 100.0,
                    "Loan tenor (years)": 2.0,
                    "Manual debt repayments": 0.0,
                },
                {
                    "Year": 2027,
                    "Debt drawdowns": 60.0,
                    "Loan tenor (years)": 1.0,
                    "Manual debt repayments": 0.0,
                },
            ]
        )

        updated = _apply_debt_schedule(cash_flow_df, debt_schedule, 0.10)

        self.assertIsNotNone(updated)
        assert updated is not None
        self.assertAlmostEqual(float(updated.loc[2025, "Debt opening balance"]), 0.0)
        self.assertAlmostEqual(float(updated.loc[2025, "Debt repayments"]), 50.0)
        self.assertAlmostEqual(float(updated.loc[2025, "Debt closing balance"]), 50.0)
        self.assertAlmostEqual(float(updated.loc[2026, "Interest paid"]), 5.0)
        self.assertAlmostEqual(float(updated.loc[2026, "Debt closing balance"]), 0.0)
        self.assertAlmostEqual(float(updated.loc[2027, "Debt opening balance"]), 0.0)
        self.assertAlmostEqual(float(updated.loc[2027, "Debt repayments"]), 60.0)
        self.assertAlmostEqual(float(updated.loc[2027, "Debt closing balance"]), 0.0)
        self.assertAlmostEqual(float(updated.loc[2028, "Debt opening balance"]), 0.0)
        self.assertAlmostEqual(float(updated.loc[2028, "Interest paid"]), 0.0)

    def test_apply_debt_schedule_forces_manual_maturity_paydown(self) -> None:
        cash_flow_df = pd.DataFrame(
            {
                "Net cash from operations": [0.0, 0.0, 0.0],
                "Net cash from investing": [0.0, 0.0, 0.0],
                "Equity issuance": [0.0, 0.0, 0.0],
                "Beginning cash balance": [0.0, 0.0, 0.0],
            },
            index=[2025, 2026, 2027],
        )
        debt_schedule = pd.DataFrame(
            [
                {
                    "Year": 2025,
                    "Debt drawdowns": 120.0,
                    "Loan tenor (years)": 2.0,
                    "Manual debt repayments": 0.0,
                },
                {
                    "Year": 2026,
                    "Debt drawdowns": 0.0,
                    "Loan tenor (years)": 0.0,
                    "Manual debt repayments": 10.0,
                },
            ]
        )

        updated = _apply_debt_schedule(
            cash_flow_df,
            debt_schedule,
            0.10,
            repayment_mode="manual",
        )

        self.assertIsNotNone(updated)
        assert updated is not None
        self.assertAlmostEqual(float(updated.loc[2025, "Debt repayments"]), 0.0)
        self.assertAlmostEqual(float(updated.loc[2025, "Debt closing balance"]), 120.0)
        self.assertAlmostEqual(float(updated.loc[2026, "Interest paid"]), 12.0)
        self.assertAlmostEqual(float(updated.loc[2026, "Debt repayments"]), 120.0)
        self.assertAlmostEqual(float(updated.loc[2026, "Debt closing balance"]), 0.0)
        self.assertAlmostEqual(float(updated.loc[2027, "Interest paid"]), 0.0)

    def test_consume_pending_panel_state_applies_deferred_checkbox_value(self) -> None:
        session_state = {
            "uses_table_add_open": True,
            _pending_panel_state_key("uses_table_add_open"): False,
        }

        with patch("streamlit_app.st.session_state", session_state):
            pending = _consume_pending_panel_state("uses_table_add_open")

        self.assertFalse(pending)
        self.assertFalse(session_state["uses_table_add_open"])
        self.assertNotIn(_pending_panel_state_key("uses_table_add_open"), session_state)

    def test_consume_pending_panel_state_requeues_when_checkbox_key_is_locked(self) -> None:
        class LockedState(dict):
            def __setitem__(self, key, value):
                if key == "debt_schedule_table_add_open":
                    raise StreamlitAPIException("locked")
                super().__setitem__(key, value)

        session_state = LockedState(
            {
                _pending_panel_state_key("debt_schedule_table_add_open"): False,
            }
        )

        with patch("streamlit_app.st.session_state", session_state):
            pending = _consume_pending_panel_state("debt_schedule_table_add_open")

        self.assertIsNone(pending)
        self.assertIn(_pending_panel_state_key("debt_schedule_table_add_open"), session_state)
        self.assertFalse(session_state[_pending_panel_state_key("debt_schedule_table_add_open")])
        self.assertNotIn("debt_schedule_table_add_open", session_state)

    def test_set_dataframe_cell_upcasts_for_incompatible_editor_value(self) -> None:
        df = pd.DataFrame({"Year": [2025]})

        _set_dataframe_cell(df, 0, "Year", None)

        self.assertIsNone(df.at[0, "Year"])
        self.assertEqual(str(df["Year"].dtype), "object")

    def test_render_row_selector_keeps_selected_vaccine_id(self) -> None:
        df = pd.DataFrame(
            [
                {"ID_vaccine": "VAC-001", "Vaccine name": "AgSeed-101"},
                {"ID_vaccine": "VAC-002", "Vaccine name": "BioYield-Plus"},
            ]
        )
        session_state = {}
        captured: dict[str, object] = {}
        select_key = "vaccine_revenue_table_row_select"

        def _fake_selectbox(_label, *, options, format_func, index, key):
            captured["options"] = list(options)
            captured["labels"] = [format_func(option) for option in options]
            captured["index"] = index
            session_state[key] = options[1]
            return options[1]

        with patch("streamlit_app.st.session_state", session_state), patch(
            "streamlit_app.st.selectbox", side_effect=_fake_selectbox
        ):
            selected_idx = _render_row_selector(df, select_key, "ID_vaccine", "Vaccine name")
            _validate_selection(df, select_key, "ID_vaccine")

        self.assertEqual(selected_idx, 1)
        self.assertEqual(captured["options"], ["VAC-001", "VAC-002"])
        self.assertEqual(captured["labels"], ["VAC-001 - AgSeed-101", "VAC-002 - BioYield-Plus"])
        self.assertEqual(captured["index"], 0)
        self.assertEqual(session_state[select_key], "VAC-002")
        self.assertNotIn(f"{select_key}_pending", session_state)

    def test_render_row_selector_migrates_name_based_legacy_index_state(self) -> None:
        df = pd.DataFrame(
            [
                {"name": "AgSeed-101", "stage": "Discovery"},
                {"name": "BioYield-Plus", "stage": "Preclinical"},
            ]
        )
        session_state = {"product_table_row_select": 0}
        captured: dict[str, object] = {}
        select_key = "product_table_row_select"

        def _fake_selectbox(_label, *, options, format_func, index, key):
            captured["options"] = list(options)
            captured["labels"] = [format_func(option) for option in options]
            captured["index"] = index
            session_state[key] = options[1]
            return options[1]

        with patch("streamlit_app.st.session_state", session_state), patch(
            "streamlit_app.st.selectbox", side_effect=_fake_selectbox
        ):
            selected_idx = _render_row_selector(df, select_key, None, "name")
            _validate_selection(df, select_key, None, "name")

        self.assertEqual(selected_idx, 1)
        self.assertEqual(captured["options"], ["AgSeed-101", "BioYield-Plus"])
        self.assertEqual(captured["labels"], ["AgSeed-101", "BioYield-Plus"])
        self.assertEqual(captured["index"], 0)
        self.assertEqual(session_state[select_key], "BioYield-Plus")
        self.assertNotIn(f"{select_key}_pending", session_state)

    def test_enterprise_to_equity_bridge_applies_cash_debt_and_new_equity(self) -> None:
        model_cfg = ModelConfig(first_year=2024, n_years=1, discount_rate=0.0, tax_rate=0.0, ev_ebitda_multiple=0.0)
        product_cfg = ProductConfig(
            name="BridgeAsset",
            stage="Commercial",
            success_prob=1.0,
            include_in_consolidation=True,
            preexisting_market=True,
            time_to_market=0,
            patent_years=5,
            patent_revenue_target=100.0,
            post_patent_revenue_target=0.0,
        )
        valuation_result = ValuationEngine(Portfolio([Product(product_cfg, model_cfg)], model_cfg)).run()
        cash_flow_df = pd.DataFrame(
            {
                "Ending cash balance": [15.0],
                "Debt closing balance": [40.0],
            },
            index=[2024],
        )

        bridge = _build_enterprise_to_equity_bridge(valuation_result, cash_flow_df, planned_new_equity=25.0)
        amounts = dict(zip(bridge["Component"], bridge["Amount"]))
        self.assertEqual(amounts["Enterprise value (DCF)"], valuation_result.enterprise_value)
        self.assertEqual(amounts["Less: debt outstanding"], -40.0)
        self.assertEqual(amounts["Add: cash / (cash deficit)"], 15.0)
        self.assertEqual(amounts["Post-money equity value"], amounts["Pre-money equity value"] + 25.0)

    def test_build_revenue_series_handles_readonly_series_values(self) -> None:
        model_cfg = ModelConfig(first_year=2024, n_years=6)
        product_cfg = ProductConfig(
            name="ReadonlyAsset",
            stage="Commercial",
            success_prob=1.0,
            include_in_consolidation=True,
            preexisting_market=True,
            time_to_market=0,
            patent_years=5,
            patent_revenue_target=100.0,
            post_patent_revenue_target=50.0,
            market_growth_patent=0.0,
            market_growth_post=0.0,
        )
        product = Product(product_cfg, model_cfg)

        def _readonly_values(series: pd.Series):
            values = series.to_numpy(copy=False)
            values.setflags(write=False)
            return values

        with patch.object(pd.Series, "values", new=property(_readonly_values)):
            revenue = product.build_revenue_series()

        self.assertEqual(len(revenue), model_cfg.n_years)
        self.assertTrue((revenue >= 0.0).all())

    def test_bankable_snapshot_payload_ignores_advisory_metric_overrides(self) -> None:
        model_cfg = ModelConfig(first_year=2024, n_years=5, discount_rate=0.10, tax_rate=0.20)
        product_cfg = ProductConfig(
            name="LockedSnapshotAsset",
            stage="Commercial",
            success_prob=1.0,
            include_in_consolidation=True,
            preexisting_market=True,
            time_to_market=0,
            patent_years=5,
            patent_revenue_target=100.0,
            post_patent_revenue_target=0.0,
        )
        portfolio = Portfolio([Product(product_cfg, model_cfg)], model_cfg)
        valuation_result = ValuationEngine(portfolio).run()
        advisory_inputs = {
            "npv": 999_999_999.0,
            "irr": 9.99,
            "dscr_min": 0.1,
            "scenarios": [{"name": "Manual override", "npv": -123.0, "irr": -0.5}],
            "notes": "Do not export this.",
        }

        payload = _build_bankable_snapshot_payload(
            "BIO-LOCK",
            model_cfg,
            valuation_result,
            portfolio,
            workbook_hash="hash-123",
            advisory_inputs=advisory_inputs,
        )

        self.assertEqual(payload["snapshot_source"], "live_model_only")
        self.assertEqual(payload["project_id"], "BIO-LOCK")
        self.assertEqual(payload["workbook_hash"], "hash-123")
        self.assertEqual(payload["advisory_inputs"]["npv"], advisory_inputs["npv"])
        self.assertEqual(payload["advisory_inputs"]["notes"], advisory_inputs["notes"])
        self.assertNotEqual(payload["financial_snapshot"]["npv"], advisory_inputs["npv"])
        self.assertNotEqual(payload["financial_snapshot"]["irr"], advisory_inputs["irr"])
        self.assertNotEqual(payload["financial_snapshot"]["dscr_min"], advisory_inputs["dscr_min"])
        self.assertNotEqual(
            payload["financial_snapshot"]["scenarios"],
            advisory_inputs["scenarios"],
        )

    def test_streamlit_app_smoke_renders_without_exceptions(self) -> None:
        app = AppTest.from_file(str(APP_PATH))

        app.run(timeout=120)

        self.assertEqual(len(app.exception), 0, app.exception)

    def test_streamlit_app_default_run_populates_outputs(self) -> None:
        app = AppTest.from_file(str(APP_PATH))

        app.run(timeout=120)

        self.assertEqual(len(app.exception), 0, app.exception)
        self.assertTrue(
            any("Run complete:" in element.value for element in app.success),
            [element.value for element in app.success],
        )
        self.assertFalse(
            any("Validation issues detected:" in element.value for element in app.error),
            [element.value for element in app.error],
        )

    def test_product_assumption_table_uses_fresh_editor_key_after_row_edit(self) -> None:
        old_df = pd.DataFrame(
            [
                {"Year": 2025, "Debt drawdowns": 0.0},
                {"Year": 2026, "Debt drawdowns": 100.0},
            ]
        )
        session_state = {
            "debt_schedule_table": old_df.copy(),
            "debt_schedule_table_edit_workflow_open": True,
        }
        action_cols = [contextlib.nullcontext() for _ in range(4)]
        editor_keys: list[str] = []

        def fake_selectbox(_label, *, options, **kwargs):
            value = options[1]
            key = kwargs.get("key")
            if key:
                session_state[key] = value
            return value

        def fake_data_editor(df, **kwargs):
            editor_keys.append(kwargs["key"])
            if kwargs["key"] == "debt_schedule_table_editor_0":
                return old_df.copy()
            return df.copy()

        with patch("streamlit_app.st.session_state", session_state):
            with patch("streamlit_app._render_row_selector", return_value=1):
                with patch("streamlit_app.st.columns", return_value=action_cols), patch(
                    "streamlit_app.st.checkbox", side_effect=[False, False]
                ), patch("streamlit_app.st.button", return_value=False), patch(
                    "streamlit_app.st.caption"
                ), patch("streamlit_app.st.success"), patch(
                    "streamlit_app.st.markdown"
                ), patch(
                    "streamlit_app.st.selectbox", side_effect=fake_selectbox
                ), patch(
                    "streamlit_app._render_row_form",
                    return_value=("save", {"Year": 2026, "Debt drawdowns": 250.0}),
                ), patch(
                    "streamlit_app.st.rerun"
                ), patch(
                    "streamlit_app.st.data_editor", side_effect=fake_data_editor
                ), patch(
                    "streamlit_app._validate_selection"
                ):
                    from streamlit_app import _render_product_assumption_table

                    result_df = _render_product_assumption_table(
                        session_key="debt_schedule_table",
                        default_factory=lambda: old_df.copy(),
                        blank_row_factory=lambda df: {"Year": 2027, "Debt drawdowns": 0.0},
                        id_column=None,
                        name_column="Year",
                    )

        self.assertEqual(editor_keys, ["debt_schedule_table_editor_1"])
        self.assertEqual(float(result_df.loc[1, "Debt drawdowns"]), 250.0)
        self.assertEqual(float(session_state["debt_schedule_table"].loc[1, "Debt drawdowns"]), 250.0)

    def test_machine_learning_multiple_handles_object_numeric_inputs(self) -> None:
        cons = pd.DataFrame(
            {
                "revenue": ["1000000", "1200000", "1500000", "1800000"],
                "ebitda": ["150000", "200000", "260000", "320000"],
            },
            index=[2024, 2025, 2026, 2027],
        )

        result = _machine_learning_multiple(cons)

        self.assertIsNotNone(result)
        assert result is not None
        self.assertEqual(list(result.columns), ["Year", "Predicted multiple"])
        self.assertEqual(len(result), 4)
        self.assertTrue(pd.api.types.is_numeric_dtype(result["Predicted multiple"]))

    def test_resolve_valuation_context_reuses_saved_outputs_on_validation_failure(self) -> None:
        saved_model_cfg = object()
        saved_portfolio = object()
        saved_result = object()
        candidate_model_cfg = object()
        candidate_portfolio = object()
        session_state = {
            "model_config": saved_model_cfg,
            "portfolio": saved_portfolio,
            "valuation_result": saved_result,
        }

        with patch("streamlit_app.st.session_state", session_state), patch(
            "streamlit_app.validate_portfolio", return_value=["invalid weights"]
        ):
            model_cfg, portfolio, valuation_result, validation_issues, used_saved_outputs = (
                _resolve_valuation_context(candidate_model_cfg, candidate_portfolio)
            )

        self.assertIs(model_cfg, saved_model_cfg)
        self.assertIs(portfolio, saved_portfolio)
        self.assertIs(valuation_result, saved_result)
        self.assertEqual(validation_issues, ["invalid weights"])
        self.assertTrue(used_saved_outputs)

    def test_resolve_valuation_context_persists_successful_run(self) -> None:
        candidate_model_cfg = object()
        candidate_portfolio = object()
        valuation_result = object()
        session_state = {}

        with patch("streamlit_app.st.session_state", session_state), patch(
            "streamlit_app.validate_portfolio", return_value=[]
        ), patch("streamlit_app.ValuationEngine") as engine_cls:
            engine_cls.return_value.run.return_value = valuation_result
            model_cfg, portfolio, resolved_result, validation_issues, used_saved_outputs = (
                _resolve_valuation_context(candidate_model_cfg, candidate_portfolio)
            )

        self.assertIs(model_cfg, candidate_model_cfg)
        self.assertIs(portfolio, candidate_portfolio)
        self.assertIs(resolved_result, valuation_result)
        self.assertEqual(validation_issues, [])
        self.assertFalse(used_saved_outputs)
        self.assertIs(session_state["model_config"], candidate_model_cfg)
        self.assertIs(session_state["portfolio"], candidate_portfolio)
        self.assertIs(session_state["valuation_result"], valuation_result)

    def test_render_row_selector_preserves_selected_vaccine_row_across_reruns(self) -> None:
        df = pd.DataFrame(
            [
                {"ID_vaccine": "VAC-001", "Vaccine name": "AgSeed-101"},
                {"ID_vaccine": "VAC-002", "Vaccine name": "BioYield-Plus"},
            ]
        )
        session_state = {"vaccine_development_table_row_select": 0}
        select_key = "vaccine_development_table_row_select"
        captured_indexes: list[int] = []

        def _first_selectbox(_label, *, options, format_func, index, key):
            captured_indexes.append(index)
            session_state[key] = options[1]
            return options[1]

        def _second_selectbox(_label, *, options, format_func, index, key):
            captured_indexes.append(index)
            session_state[key] = options[index]
            return options[index]

        with patch("streamlit_app.st.session_state", session_state), patch(
            "streamlit_app.st.selectbox", side_effect=_first_selectbox
        ):
            first_selected_idx = _render_row_selector(df, select_key, "ID_vaccine", "Vaccine name")
            _validate_selection(df, select_key, "ID_vaccine", "Vaccine name")

        with patch("streamlit_app.st.session_state", session_state), patch(
            "streamlit_app.st.selectbox", side_effect=_second_selectbox
        ):
            second_selected_idx = _render_row_selector(df, select_key, "ID_vaccine", "Vaccine name")
            _validate_selection(df, select_key, "ID_vaccine", "Vaccine name")

        self.assertEqual(first_selected_idx, 1)
        self.assertEqual(second_selected_idx, 1)
        self.assertEqual(captured_indexes, [0, 1])
        self.assertEqual(session_state[select_key], "VAC-002")
        self.assertNotIn(f"{select_key}_pending", session_state)

    def test_probability_preview_flags_stage_transition_source(self) -> None:
        product_df = pd.DataFrame(
            [
                {
                    "name": "AgSeed-101",
                    "stage": "Phase 2",
                    "success_prob": 0.35,
                    "include_in_consolidation": True,
                    "time_to_market": 4,
                    "patent_years": 15,
                    "patent_revenue_target": 120_000_000,
                    "post_patent_revenue_target": 60_000_000,
                }
            ]
        )
        preview = _build_probability_preview(
            product_df,
            ModelConfig(first_year=2024, n_years=10),
            _default_stage_schedule_mapping(),
            overwrite_defaults=False,
            detail_tables={},
        )
        self.assertEqual(preview.iloc[0]["Stage"], "Phase II")
        self.assertEqual(preview.iloc[0]["Probability source"], "Stage-transition path")
        self.assertEqual(preview.iloc[0]["Probability path used"], "Stage-transition path")
        self.assertGreater(preview.iloc[0]["Effective cumulative success probability"], 0.0)

    def test_probability_preview_normalizes_market_alias_to_commercial(self) -> None:
        product_df = pd.DataFrame(
            [
                {
                    "name": "CommercialAsset",
                    "stage": "Market",
                    "success_prob": 1.0,
                    "include_in_consolidation": True,
                    "time_to_market": 0,
                    "patent_years": 10,
                    "patent_revenue_target": 50_000_000,
                    "post_patent_revenue_target": 25_000_000,
                }
            ]
        )
        preview = _build_probability_preview(
            product_df,
            ModelConfig(first_year=2024, n_years=10),
            _default_stage_schedule_mapping(),
            overwrite_defaults=False,
            detail_tables={},
        )
        self.assertEqual(preview.iloc[0]["Stage"], "Commercial")
        self.assertEqual(preview.iloc[0]["Probability source"], "Stage-transition path")

    def test_stage_generated_milestones_do_not_embed_transition_probability_twice(self) -> None:
        mapping_df = _default_stage_schedule_mapping().copy()
        phase_mask = mapping_df["Stage"] == "Phase II"
        mapping_df.loc[phase_mask, "Phase II duration (years)"] = 1
        mapping_df.loc[phase_mask, "Phase III duration (years)"] = 1
        mapping_df.loc[phase_mask, "Approval duration (years)"] = 0
        mapping_df.loc[phase_mask, "Time to market (years)"] = 2
        mapping_df.loc[phase_mask, "Phase II->Phase III"] = 50.0
        mapping_df.loc[phase_mask, "Phase III->Approval"] = 50.0
        mapping_df.loc[phase_mask, "Phase II->Phase III annual success %"] = 50.0
        mapping_df.loc[phase_mask, "Phase III->Approval annual success %"] = 50.0
        for col in [c for c in mapping_df.columns if c.endswith("completion milestone (USD)")]:
            mapping_df.loc[phase_mask, col] = 0.0
        mapping_df.loc[phase_mask, "Phase II completion milestone (USD)"] = 100.0

        product_df = pd.DataFrame(
            [
                {
                    "name": "StageMilestone",
                    "stage": "Phase II",
                    "success_prob": 0.9,
                    "include_in_consolidation": True,
                    "time_to_market": 2,
                    "patent_years": 5,
                    "patent_revenue_target": 0.0,
                    "post_patent_revenue_target": 0.0,
                }
            ]
        )
        model_cfg = ModelConfig(
            first_year=2024,
            n_years=10,
            discount_rate=0.0,
            tax_rate=0.0,
            ev_ebitda_multiple=0.0,
        )
        portfolio = _build_portfolio(
            product_df,
            model_cfg,
            stage_mapping=mapping_df,
            overwrite_defaults=False,
            detail_tables={},
        )
        self.assertIsNotNone(portfolio)
        product = portfolio.products[0]
        normalized_milestones = product._normalized_milestones()
        self.assertEqual(len(normalized_milestones), 1)
        self.assertEqual(normalized_milestones[0].probability, 1.0)

        valuation_result = ValuationEngine(portfolio).run()
        milestone_year = model_cfg.first_year + normalized_milestones[0].year_offset
        expected_weighted_milestone = (
            float(normalized_milestones[0].amount)
            * float(product._success_prob_schedule().loc[milestone_year])
        )
        self.assertEqual(valuation_result.per_product["StageMilestone"].loc[milestone_year, "milestones"], 100.0)
        self.assertEqual(
            valuation_result.consolidated.loc[milestone_year, "milestones"],
            expected_weighted_milestone,
        )

    def test_stage_mapping_candidate_row_derives_time_to_market_from_durations(self) -> None:
        mapping_df = _default_stage_schedule_mapping().copy()
        base_row = mapping_df.loc[mapping_df["Stage"] == "Phase II"].iloc[0]

        candidate = _build_stage_mapping_candidate_row(
            base_row,
            {
                "Stage": "Phase II",
                "Time to market (years)": 99,
                "Phase II duration (years)": 3,
                "Phase III duration (years)": 2,
                "Approval duration (years)": 1,
            },
        )

        self.assertEqual(candidate["Time to market (years)"], 6)

    def test_stage_mapping_row_warnings_preview_inconsistent_approval_row(self) -> None:
        mapping_df = _default_stage_schedule_mapping().copy()
        row_idx = mapping_df.index[mapping_df["Stage"] == "Approval"][0]
        base_row = mapping_df.loc[row_idx]

        candidate = _build_stage_mapping_candidate_row(
            base_row,
            {
                "Stage": "Approval",
                "Discovery duration (years)": 1,
                "Preclinical duration (years)": 1,
                "Phase I duration (years)": 1,
                "Phase II duration (years)": 2,
                "Phase III duration (years)": 1,
                "Approval duration (years)": 1,
            },
        )
        warnings = _stage_mapping_row_warnings(mapping_df, int(row_idx), candidate)

        self.assertTrue(any("time-to-market" in warning for warning in warnings))
        self.assertTrue(any("early-stage durations" in warning for warning in warnings))

    def test_debt_schedule_builds_bullet_profile_and_lender_metrics(self) -> None:
        cash_flow_df = pd.DataFrame(
            {
                "Net cash from operations": [80.0, 90.0, 100.0],
                "Net cash from investing": [-10.0, -10.0, -10.0],
                "Equity issuance": [0.0, 0.0, 0.0],
                "Net cash from financing": [0.0, 0.0, 0.0],
                "Net change in cash": [70.0, 80.0, 90.0],
                "Beginning cash balance": [0.0, 0.0, 0.0],
                "Ending cash balance": [70.0, 150.0, 240.0],
            },
            index=[2024, 2025, 2026],
        )
        debt_schedule = pd.DataFrame(
            {
                "Year": [2024, 2025, 2026],
                "Debt drawdowns": [120.0, 0.0, 0.0],
                "Manual debt repayments": [0.0, 0.0, 0.0],
            }
        )

        updated = _apply_debt_schedule(
            cash_flow_df,
            debt_schedule,
            interest_rate=0.10,
            repayment_mode="bullet",
            grace_years=1,
            minimum_cash_reserve=5.0,
        )
        self.assertEqual(updated.loc[2024, "Debt repayments"], 0.0)
        self.assertEqual(updated.loc[2025, "Debt repayments"], 0.0)
        self.assertEqual(updated.loc[2026, "Debt repayments"], 120.0)
        self.assertEqual(updated.loc[2025, "Beginning cash balance"], updated.loc[2024, "Ending cash balance"])

        metrics = _build_lender_metrics(
            updated,
            discount_rate=0.10,
            minimum_cash_reserve=5.0,
            target_dscr=1.2,
        )
        self.assertAlmostEqual(metrics.loc[2025, "DSCR"], 80.0 / 12.0)
        self.assertAlmostEqual(metrics.loc[2026, "Debt service"], 132.0)
        self.assertEqual(metrics.loc[2025, "Covenant status"], "Pass")

    def test_investor_waterfall_applies_preference_before_common(self) -> None:
        shareholders_df = pd.DataFrame(
            [
                {
                    "Shareholder": "Founders",
                    "Security": "Common",
                    "Seniority": 3,
                    "Ownership %": 0.5,
                    "Investment": 20.0,
                    "Liquidation preference (x)": 0.0,
                    "Participating preferred": False,
                },
                {
                    "Shareholder": "Series A",
                    "Security": "Preferred",
                    "Seniority": 1,
                    "Ownership %": 0.5,
                    "Investment": 40.0,
                    "Liquidation preference (x)": 1.0,
                    "Participating preferred": False,
                },
            ]
        )

        waterfall = _build_investor_waterfall(shareholders_df, exit_equity_value=60.0).set_index("Shareholder")
        self.assertEqual(waterfall.loc["Series A", "Decision"], "Take preference")
        self.assertEqual(waterfall.loc["Series A", "Preference paid"], 40.0)
        self.assertEqual(waterfall.loc["Founders", "Total proceeds"], 20.0)

    def test_financial_excel_includes_debt_and_waterfall_sheets(self) -> None:
        cons = pd.DataFrame(
            {
                "revenue": [100.0],
                "cogs": [-40.0],
                "ebitda": [20.0],
                "nopat": [12.0],
                "rd_cash": [-5.0],
                "capex_cash": [-3.0],
                "fcff_after_wc": [10.0],
            },
            index=[2024],
        )
        perf_df = pd.DataFrame({"Revenue": [100.0]}, index=[2024])
        position_df = pd.DataFrame({"Total equity": [50.0]}, index=[2024])
        cash_flow_df = pd.DataFrame({"Net change in cash": [10.0]}, index=[2024])
        lender_metrics = pd.DataFrame(
            {"CFADS": [10.0], "Debt service": [5.0], "DSCR": [2.0]},
            index=[2024],
        )
        investor_waterfall = pd.DataFrame(
            {
                "Shareholder": ["Founders"],
                "Total proceeds": [25.0],
            }
        )

        workbook = load_workbook(
            BytesIO(
                _build_financial_excel(
                    cons,
                    perf_df,
                    position_df,
                    cash_flow_df,
                    ModelConfig(first_year=2024, n_years=1),
                    lender_metrics=lender_metrics,
                    investor_waterfall=investor_waterfall,
                )
            )
        )
        self.assertIn("Debt metrics", workbook.sheetnames)
        self.assertIn("Investor waterfall", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()

